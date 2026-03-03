# -*- coding: utf-8 -*-
"""
テスト項目書 比較分析スクリプト v2
参考例（ユーザー改善版） vs 成果物v7（AI改善版）
"""
import openpyxl
import re
import sys
import io
import warnings

warnings.filterwarnings('ignore')

# Windows環境でのUTF-8出力対応
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

FILE_A = r"C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\テスト項目書_ログイン方法でID・PWとTOKIUM IDが存在し、TOKIUM IDだけ利用したい場合にフラグでログイン制限できるように (2).xlsx"
FILE_B = r"C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\成果物\テスト項目書_改善後_v7.xlsx"

# 曖昧表現パターン
AMBIGUOUS_PATTERNS = [
    r'正常であること',
    r'問題ないこと',
    r'仕様通り',
    r'適切に',
    r'正しく',
    r'期待通り',
    r'想定通り',
]

# 良好末尾表現
GOOD_ENDINGS = [
    r'されること\s*$',
    r'されていること\s*$',
    r'であること\s*$',
    r'できること\s*$',
    r'ないこと\s*$',
    r'なること\s*$',
    r'いること\s*$',
    r'れること\s*$',
    r'すること\s*$',
    r'あること\s*$',
    r'していること\s*$',
]


def cell_str(ws, row, col):
    """セルの値を文字列で取得"""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def extract_testcases_from_sheet(ws, sheet_name):
    """テストケースシートからケースを抽出（ヘッダー2行目固定、列構造固定）"""
    # 列マッピング（2行目がヘッダー）
    NO_COL = 2
    GROUP_COL = 3  # 確認すること
    SCREEN_COL = 4  # 画面
    TARGET_COL = 5  # 確認対象
    DETAIL_COL = 6  # 詳細
    PROC_COL = 7    # テスト実行手順
    EXPECT_COL = 8  # 期待値
    REMARK_COL = 9  # 備考

    testcases = []
    current_group = ""
    current_screen = ""
    current_procedure = ""

    data_start = 3  # 3行目からデータ

    for row_idx in range(data_start, ws.max_row + 1):
        expected = cell_str(ws, row_idx, EXPECT_COL)
        if not expected:
            continue

        no = cell_str(ws, row_idx, NO_COL)
        group = cell_str(ws, row_idx, GROUP_COL)
        screen = cell_str(ws, row_idx, SCREEN_COL)
        target = cell_str(ws, row_idx, TARGET_COL)
        detail = cell_str(ws, row_idx, DETAIL_COL)
        procedure = cell_str(ws, row_idx, PROC_COL)
        remark = cell_str(ws, row_idx, REMARK_COL)

        # グループ情報の継承（結合セル対応）
        if group:
            current_group = group
        if screen:
            current_screen = screen
        if procedure and procedure != "-":
            current_procedure = procedure

        tc = {
            'row': row_idx,
            'sheet': sheet_name,
            'no': no,
            'group': current_group if not group else group,
            'screen': current_screen if not screen else screen,
            'target': target,
            'detail': detail,
            'procedure': procedure if (procedure and procedure != "-") else "",
            'procedure_inherited': current_procedure if not (procedure and procedure != "-") else procedure,
            'expected': expected,
            'remark': remark,
        }
        testcases.append(tc)

    return testcases


def analyze_testcases(testcases, label):
    """テストケースの統計情報を分析・出力"""
    total = len(testcases)
    if total == 0:
        print(f"  テストケース数: 0件")
        return {}

    # 期待値の文字数
    expected_lengths = [len(tc['expected']) for tc in testcases]
    avg_len = sum(expected_lengths) / total
    min_len = min(expected_lengths)
    max_len = max(expected_lengths)

    # 手順記載率（直接記載 + 継承）
    has_procedure_direct = sum(1 for tc in testcases if tc['procedure'])
    has_procedure_inherited = sum(1 for tc in testcases if tc['procedure_inherited'])

    # 「こと」の回数分析
    koto_0 = sum(1 for tc in testcases if tc['expected'].count('こと') == 0)
    koto_1 = sum(1 for tc in testcases if tc['expected'].count('こと') == 1)
    koto_2plus = sum(1 for tc in testcases if tc['expected'].count('こと') >= 2)

    # 曖昧表現チェック
    ambiguous_count = 0
    ambiguous_examples = []
    for tc in testcases:
        for pat in AMBIGUOUS_PATTERNS:
            if re.search(pat, tc['expected']):
                ambiguous_count += 1
                if len(ambiguous_examples) < 5:
                    ambiguous_examples.append((tc['no'], tc['expected'][:80], pat))
                break

    # 良好末尾表現チェック
    good_ending_count = 0
    for tc in testcases:
        for pat in GOOD_ENDINGS:
            if re.search(pat, tc['expected']):
                good_ending_count += 1
                break

    # 番号付き手順の割合
    numbered_procedure = 0
    total_with_proc = 0
    for tc in testcases:
        proc = tc['procedure_inherited']
        if proc:
            total_with_proc += 1
            if re.search(r'(^|\n)\s*([\d]+[\.\)）]|①|【手順|手順\s*[\d]|Step\s*[\d]|【セットアップ|【前提)', proc):
                numbered_procedure += 1

    # 期待値の長さ分布
    short_count = sum(1 for l in expected_lengths if l < 10)
    long_count = sum(1 for l in expected_lengths if l > 80)
    normal_count = sum(1 for l in expected_lengths if 10 <= l <= 80)

    stats = {
        'total': total,
        'avg_len': avg_len,
        'min_len': min_len,
        'max_len': max_len,
        'has_procedure_direct': has_procedure_direct,
        'has_procedure_inherited': has_procedure_inherited,
        'procedure_direct_rate': has_procedure_direct / total * 100,
        'procedure_inherited_rate': has_procedure_inherited / total * 100,
        'koto_0': koto_0,
        'koto_1': koto_1,
        'koto_2plus': koto_2plus,
        'koto_1_rate': koto_1 / total * 100,
        'koto_2plus_rate': koto_2plus / total * 100,
        'ambiguous_count': ambiguous_count,
        'ambiguous_rate': ambiguous_count / total * 100,
        'good_ending_count': good_ending_count,
        'good_ending_rate': good_ending_count / total * 100,
        'numbered_procedure': numbered_procedure,
        'numbered_rate': numbered_procedure / total_with_proc * 100 if total_with_proc else 0,
        'short_count': short_count,
        'long_count': long_count,
        'normal_count': normal_count,
    }

    print(f"\n--- テストケース統計 [{label}] ---")
    print(f"  総テストケース数: {total}件")
    print()
    print(f"  [期待値の文字数]")
    print(f"    平均: {avg_len:.1f}字")
    print(f"    最小: {min_len}字 / 最大: {max_len}字")
    print(f"    適正範囲(10-80字): {normal_count}件 ({normal_count/total*100:.1f}%)")
    print(f"    短すぎ(<10字): {short_count}件 ({short_count/total*100:.1f}%)")
    print(f"    長すぎ(>80字): {long_count}件 ({long_count/total*100:.1f}%)")
    print()
    print(f"  [手順記載率]")
    print(f"    直接記載: {has_procedure_direct}件 / {total}件 ({has_procedure_direct/total*100:.1f}%)")
    print(f"    継承含む: {has_procedure_inherited}件 / {total}件 ({has_procedure_inherited/total*100:.1f}%)")
    print(f"    番号付き/構造化手順: {numbered_procedure}件 / {total_with_proc}件 ({stats['numbered_rate']:.1f}%)")
    print()
    print(f"  [「こと」の出現回数]")
    print(f"    0回: {koto_0}件 ({koto_0/total*100:.1f}%)")
    print(f"    1回（適正）: {koto_1}件 ({koto_1/total*100:.1f}%)")
    print(f"    2回以上（複数混在）: {koto_2plus}件 ({koto_2plus/total*100:.1f}%)")
    if koto_2plus > 0:
        print(f"    -- 2回以上の例 --")
        count = 0
        for tc in testcases:
            if tc['expected'].count('こと') >= 2:
                print(f"      No.{tc['no']}: {tc['expected'][:80]}")
                count += 1
                if count >= 3:
                    break
    print()
    print(f"  [曖昧表現]")
    print(f"    曖昧表現を含むケース: {ambiguous_count}件 ({ambiguous_count/total*100:.1f}%)")
    if ambiguous_examples:
        for no, exp, pat in ambiguous_examples:
            print(f"      No.{no}: 「{exp}」 (検出: {pat})")
    print()
    print(f"  [良好末尾表現]")
    print(f"    良好な末尾表現: {good_ending_count}件 ({good_ending_count/total*100:.1f}%)")

    return stats


def main():
    print("=" * 80)
    print("テスト項目書 比較分析レポート")
    print("=" * 80)

    label_a = "参考例（ユーザー改善版）"
    label_b = "成果物v7（AI改善版）"

    # ==== ファイルA: 参考例 ====
    wb_a = openpyxl.load_workbook(FILE_A, data_only=True)
    print(f"\n{'='*80}")
    print(f"【{label_a}】")
    short_name_a = FILE_A.split("\\")[-1]
    print(f"ファイル: {short_name_a}")
    print(f"{'='*80}")

    print(f"\n--- シート名一覧 ---")
    for i, name in enumerate(wb_a.sheetnames, 1):
        ws = wb_a[name]
        print(f"  {i}. 「{name}」 - {ws.max_row}行 x {ws.max_column}列")

    # テストケースシートを特定
    tc_sheets_a = [n for n in wb_a.sheetnames if n.startswith('テストケース_目次')]
    print(f"\n  テストケースシート: {tc_sheets_a}")

    # ヘッダー表示
    if tc_sheets_a:
        ws_first = wb_a[tc_sheets_a[0]]
        print(f"\n--- ヘッダー行（2行目）---")
        for c in range(1, 22):
            v = ws_first.cell(row=2, column=c).value
            if v is not None:
                print(f"    列{c}: {str(v).strip()}")

    # 全テストケースを抽出
    all_tc_a = []
    for sname in tc_sheets_a:
        ws = wb_a[sname]
        tcs = extract_testcases_from_sheet(ws, sname)
        print(f"    シート「{sname}」: {len(tcs)}件")
        all_tc_a.extend(tcs)

    print(f"\n  合計テストケース数: {len(all_tc_a)}件")
    stats_a = analyze_testcases(all_tc_a, label_a)

    # ==== ファイルB: 成果物v7 ====
    wb_b = openpyxl.load_workbook(FILE_B, data_only=True)
    print(f"\n\n{'='*80}")
    print(f"【{label_b}】")
    short_name_b = FILE_B.split("\\")[-1]
    print(f"ファイル: {short_name_b}")
    print(f"{'='*80}")

    print(f"\n--- シート名一覧 ---")
    for i, name in enumerate(wb_b.sheetnames, 1):
        ws = wb_b[name]
        print(f"  {i}. 「{name}」 - {ws.max_row}行 x {ws.max_column}列")

    # テストケースシート: 「テストケース_改善後」
    tc_sheet_b = 'テストケース_改善後'
    ws_b = wb_b[tc_sheet_b]

    print(f"\n--- ヘッダー行（2行目）---")
    for c in range(1, 22):
        v = ws_b.cell(row=2, column=c).value
        if v is not None:
            print(f"    列{c}: {str(v).strip()}")

    all_tc_b = extract_testcases_from_sheet(ws_b, tc_sheet_b)
    print(f"\n  テストケース数: {len(all_tc_b)}件")
    stats_b = analyze_testcases(all_tc_b, label_b)

    # ==== 差分比較 ====
    print(f"\n\n{'='*80}")
    print(f"【差分比較】 {label_a} vs {label_b}")
    print(f"{'='*80}")

    print(f"\n--- ケース数の増減 ---")
    diff = len(all_tc_a) - len(all_tc_b)
    print(f"  {label_a}: {len(all_tc_a)}件")
    print(f"  {label_b}: {len(all_tc_b)}件")
    if diff > 0:
        print(f"  差分: {label_a}が {diff}件 多い")
    elif diff < 0:
        print(f"  差分: {label_b}が {abs(diff)}件 多い")
    else:
        print(f"  差分: 同数")

    # Noベースの対応付け
    # 参考例は複数シートにまたがるため、各シート内のNoを使う
    # v7は元No.列（列10）がある場合はそれを使い、なければNo列を使う

    # v7の元No列を確認
    print(f"\n--- No.ベースの対応付け ---")

    # 参考例: シートごとにNoをユニーク化
    # 全体でNoが重複する可能性があるため、シート内で一意のキーを作成
    # しかしv7は1シートに統合されている
    # v7の「元No.」列（列10）で対応を試みる

    # v7の元No取得
    v7_orig_nos = {}
    for tc in all_tc_b:
        ws_cell = ws_b.cell(row=tc['row'], column=10)
        orig_no = ws_cell.value
        if orig_no is not None:
            tc['orig_no'] = str(orig_no).strip()
        else:
            tc['orig_no'] = ""

    # v7のNoをキーにしたマップ
    no_map_b = {}
    for tc in all_tc_b:
        if tc['no']:
            no_map_b[tc['no']] = tc

    # 参考例のNoをキーにしたマップ（シートごとに分離）
    no_map_a_by_sheet = {}
    for tc in all_tc_a:
        key = f"{tc['sheet']}_{tc['no']}"
        no_map_a_by_sheet[key] = tc

    # 簡易比較: 期待値の完全一致
    expected_set_a = set(tc['expected'] for tc in all_tc_a)
    expected_set_b = set(tc['expected'] for tc in all_tc_b)

    common_expected = expected_set_a & expected_set_b
    only_a = expected_set_a - expected_set_b
    only_b = expected_set_b - expected_set_a

    print(f"\n--- 期待値ベースの比較 ---")
    print(f"  期待値が完全一致: {len(common_expected)}件")
    print(f"  {label_a}にのみ存在する期待値: {len(only_a)}件")
    print(f"  {label_b}にのみ存在する期待値: {len(only_b)}件")

    # v7の元Noと参考例_目次1のNoで対応を試みる
    # 参考例の目次1のNoリスト
    tc_a_sheet1 = [tc for tc in all_tc_a if tc['sheet'] == 'テストケース_目次1']
    a_no_map = {tc['no']: tc for tc in tc_a_sheet1 if tc['no']}

    # v7のorig_noで目次1のNoとマッチング
    matched_pairs = []
    for tc_b_item in all_tc_b:
        orig = tc_b_item.get('orig_no', '')
        if orig and orig in a_no_map:
            matched_pairs.append((a_no_map[orig], tc_b_item))

    print(f"\n--- 元No.による対応付け（v7の元No. <-> 参考例_目次1のNo.）---")
    print(f"  マッチしたペア: {len(matched_pairs)}件")

    # 期待値が変わったケース
    changed_expected = []
    changed_procedure = []

    for tc_a_item, tc_b_item in matched_pairs:
        if tc_a_item['expected'] != tc_b_item['expected']:
            changed_expected.append((tc_a_item, tc_b_item))
        proc_a = tc_a_item['procedure_inherited']
        proc_b = tc_b_item['procedure_inherited']
        if proc_a != proc_b and (proc_a or proc_b):
            changed_procedure.append((tc_a_item, tc_b_item))

    print(f"\n--- 期待値が変更されたケース: {len(changed_expected)}件 ---")
    if changed_expected:
        show_count = min(5, len(changed_expected))
        print(f"  (代表例 {show_count}件)")
        for i, (tca, tcb) in enumerate(changed_expected[:show_count]):
            print(f"\n  [{i+1}] 参考例No.{tca['no']} / v7No.{tcb['no']}")
            print(f"    参考例: {tca['expected'][:100]}")
            print(f"    v7    : {tcb['expected'][:100]}")
    else:
        print(f"  変更なし")

    print(f"\n--- 手順が変更されたケース: {len(changed_procedure)}件 ---")
    if changed_procedure:
        show_count = min(5, len(changed_procedure))
        print(f"  (代表例 {show_count}件)")
        for i, (tca, tcb) in enumerate(changed_procedure[:show_count]):
            proc_a_show = tca['procedure_inherited'][:100] if tca['procedure_inherited'] else "(空)"
            proc_b_show = tcb['procedure_inherited'][:100] if tcb['procedure_inherited'] else "(空)"
            print(f"\n  [{i+1}] 参考例No.{tca['no']} / v7No.{tcb['no']}")
            print(f"    参考例: {proc_a_show}")
            print(f"    v7    : {proc_b_show}")
    else:
        print(f"  変更なし")

    # 参考例にのみ存在するケース（v7で削除されたもの）
    matched_a_nos = set(tc_a_item['no'] for tc_a_item, _ in matched_pairs)
    unmatched_a = [tc for tc in tc_a_sheet1 if tc['no'] and tc['no'] not in matched_a_nos]

    # v7にのみ存在するケース（新規追加されたもの）
    matched_b_orig_nos = set(tc_b_item.get('orig_no', '') for _, tc_b_item in matched_pairs)
    unmatched_b = [tc for tc in all_tc_b if tc.get('orig_no', '') and tc['orig_no'] not in [a['no'] for a in tc_a_sheet1]]

    if unmatched_a:
        print(f"\n--- 参考例_目次1に存在するがv7で対応なしのケース: {len(unmatched_a)}件 ---")
        for i, tc in enumerate(unmatched_a[:5]):
            print(f"  {i+1}. No.{tc['no']}: {tc['expected'][:80]}")

    # 参考例で追加されたシート（目次1続き、目次2、目次3）
    additional_sheets = [n for n in tc_sheets_a if n != 'テストケース_目次1']
    if additional_sheets:
        print(f"\n--- 参考例で追加されたシート ---")
        for sname in additional_sheets:
            tcs = [tc for tc in all_tc_a if tc['sheet'] == sname]
            print(f"  シート「{sname}」: {len(tcs)}件")
            for i, tc in enumerate(tcs[:3]):
                print(f"    例{i+1}. No.{tc['no']} [{tc['group']}] {tc['expected'][:60]}")

    # v7で元Noがないケース（新規追加）
    no_orig_in_b = [tc for tc in all_tc_b if not tc.get('orig_no')]
    if no_orig_in_b:
        print(f"\n--- v7で元No.がないケース（新規追加の可能性）: {len(no_orig_in_b)}件 ---")
        for i, tc in enumerate(no_orig_in_b[:5]):
            print(f"  {i+1}. No.{tc['no']}: {tc['expected'][:80]}")

    # ==== 統計比較サマリー ====
    print(f"\n\n{'='*80}")
    print(f"【統計比較サマリー】")
    print(f"{'='*80}")

    if stats_a and stats_b:
        items = [
            ("テストケース数", f"{stats_a['total']}件", f"{stats_b['total']}件", f"{stats_a['total']-stats_b['total']:+d}"),
            ("期待値平均文字数", f"{stats_a['avg_len']:.1f}字", f"{stats_b['avg_len']:.1f}字", f"{stats_a['avg_len']-stats_b['avg_len']:+.1f}"),
            ("適正範囲(10-80字)率", f"{stats_a['normal_count']/stats_a['total']*100:.1f}%", f"{stats_b['normal_count']/stats_b['total']*100:.1f}%", ""),
            ("手順記載率(直接)", f"{stats_a['procedure_direct_rate']:.1f}%", f"{stats_b['procedure_direct_rate']:.1f}%", f"{stats_a['procedure_direct_rate']-stats_b['procedure_direct_rate']:+.1f}"),
            ("手順記載率(継承含む)", f"{stats_a['procedure_inherited_rate']:.1f}%", f"{stats_b['procedure_inherited_rate']:.1f}%", f"{stats_a['procedure_inherited_rate']-stats_b['procedure_inherited_rate']:+.1f}"),
            ("番号付き手順率", f"{stats_a['numbered_rate']:.1f}%", f"{stats_b['numbered_rate']:.1f}%", f"{stats_a['numbered_rate']-stats_b['numbered_rate']:+.1f}"),
            ("「こと」1回率", f"{stats_a['koto_1_rate']:.1f}%", f"{stats_b['koto_1_rate']:.1f}%", f"{stats_a['koto_1_rate']-stats_b['koto_1_rate']:+.1f}"),
            ("「こと」2回以上率", f"{stats_a['koto_2plus_rate']:.1f}%", f"{stats_b['koto_2plus_rate']:.1f}%", f"{stats_a['koto_2plus_rate']-stats_b['koto_2plus_rate']:+.1f}"),
            ("曖昧表現率", f"{stats_a['ambiguous_rate']:.1f}%", f"{stats_b['ambiguous_rate']:.1f}%", f"{stats_a['ambiguous_rate']-stats_b['ambiguous_rate']:+.1f}"),
            ("良好末尾表現率", f"{stats_a['good_ending_rate']:.1f}%", f"{stats_b['good_ending_rate']:.1f}%", f"{stats_a['good_ending_rate']-stats_b['good_ending_rate']:+.1f}"),
        ]

        header = f"{'指標':<25s} {'参考例':>15s} {'v7':>15s} {'差':>10s}"
        print(header)
        print("-" * 70)
        for name, va, vb, d in items:
            print(f"{name:<25s} {va:>15s} {vb:>15s} {d:>10s}")

    # グループ別ケース数比較
    print(f"\n\n--- グループ（確認すること）別ケース数 ---")
    groups_a = {}
    for tc in all_tc_a:
        g = tc['group']
        groups_a[g] = groups_a.get(g, 0) + 1

    groups_b = {}
    for tc in all_tc_b:
        g = tc['group']
        groups_b[g] = groups_b.get(g, 0) + 1

    all_groups = sorted(set(list(groups_a.keys()) + list(groups_b.keys())))
    print(f"  {'グループ名':<50s} {'参考例':>6s} {'v7':>6s}")
    print(f"  {'-'*65}")
    for g in all_groups:
        ca = groups_a.get(g, 0)
        cb = groups_b.get(g, 0)
        marker = ""
        if ca > 0 and cb == 0:
            marker = " << 参考例のみ"
        elif ca == 0 and cb > 0:
            marker = " << v7のみ"
        gname = g[:48] if len(g) > 48 else g
        print(f"  {gname:<50s} {ca:>6d} {cb:>6d}{marker}")

    print(f"\n\n分析完了。")


if __name__ == '__main__':
    main()
