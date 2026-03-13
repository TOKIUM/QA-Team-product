#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""テストケース自動評価スクリプト

運用ルールに基づいてテスト項目書（Excel）を100点満点で評価する。
減点項目を検出し、修正推奨事項を出力する。

使い方:
    python evaluate_testcases.py <テスト項目書.xlsx> [--keywords キーワードファイル.txt] [--sheet シート名]

引数:
    テスト項目書.xlsx  : 評価対象のExcelファイル
    --keywords         : 我流単語チェック用のNGワードリスト（1行1パターン: NG単語<TAB>正式名称）
    --sheet            : テストケースシート名（デフォルト: 自動検出・全シート走査）
    --kanten           : 観点シート名（デフォルト: 自動検出）

キーワードファイルの例（TSV形式）:
    ID/PW	メールアドレス・パスワード
    PW	パスワード
    リカバリー	パスワード再設定
"""
import os, sys, re, json, argparse
from collections import Counter
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl でインストールしてください。")
    sys.exit(1)

from common import (
    HEADER_KEYWORDS, SKIP_SHEET_PATTERNS, VAGUE_EXCEPTIONS,
    SCREEN_NAME_URL_PATTERN, SCREEN_NAME_IMPL_TERMS, SCREEN_NAME_VAGUE,
    FORMAT_PRECONDITION_KEYWORDS, FORMAT_EXECUTION_KEYWORDS,
    SECURITY_KEYWORDS,
    detect_header_row, detect_columns, find_sheet,
    should_skip_sheet, strip_parens, is_vague_exception,
)


# =====================================================================
# 設定
# =====================================================================

# 曖昧表現のNGパターン（期待値列に含まれていたらNG）
# 外部JSON（品質基準/vague_patterns.json）があればそちらを優先。なければハードコードを使用
VAGUE_PATTERNS_DEFAULT = [
    (r'正常であること', 5, '「正常であること」は具体的な期待値に書き換えてください'),
    (r'正常に動作すること', 5, '「正常に動作すること」は具体的な挙動を記載してください'),
    (r'問題ないこと', 5, '「問題ないこと」は具体的な期待値に書き換えてください'),
    (r'問題なく[^\s]*こと', 5, '「問題なく〜」は具体的な期待値に書き換えてください'),
    (r'仕様通りであること', 5, '「仕様通りであること」は具体的な期待値に書き換えてください'),
    (r'仕様通りに動作すること', 5, '「仕様通りに動作すること」は具体的な挙動を記載してください'),
    (r'正しく動作すること', 3, '「正しく動作すること」は具体的な挙動を記載してください'),
    (r'適切に表示されること', 3, '「適切に表示されること」は表示内容を具体的に記載してください'),
    (r'適切であること', 3, '「適切であること」は具体的な基準を記載してください'),
    (r'正しく表示されること', 3, '「正しく表示されること」は表示内容を具体的に記載してください'),
    (r'期待通り', 3, '「期待通り」は具体的な内容を記載してください'),
    (r'想定通り', 3, '「想定通り」は具体的な内容を記載してください'),
]

def _load_vague_patterns():
    """品質基準/vague_patterns.json から曖昧表現パターンを読み込む"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, '..', '品質基準', 'vague_patterns.json')
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [(p[0], p[1], p[2]) for p in data.get('patterns', [])]
        except (json.JSONDecodeError, IndexError, KeyError):
            pass
    return VAGUE_PATTERNS_DEFAULT

VAGUE_PATTERNS = _load_vague_patterns()

# 期待値の複数「こと」検出
MULTI_KOTO_PENALTY = 2  # 1件あたりの減点

# 我流単語の減点
CUSTOM_WORD_PENALTY = 1  # 1件あたりの減点

# 手順省略の減点
MISSING_STEPS_PENALTY = 2  # 1件あたりの減点

# 上限設定
MAX_VAGUE_PENALTY = 15
MAX_MULTI_KOTO_PENALTY = 10
MAX_CUSTOM_WORD_PENALTY = 10
MAX_MISSING_STEPS_PENALTY = 10

# =====================================================================
# 過去参考例から算出した基準値（ベンチマーク）
# =====================================================================
# デフォルト値（benchmark.json が無い場合に使用）
# learn_benchmark.py で参考例から自動生成した benchmark.json があれば上書きされる

BENCHMARK_DEFAULT = {
    'expected_length': {
        'ideal_min': 8,     # 深掘り分析: 5%tile=8字
        'ideal_max': 80,
    },
    'steps_coverage': {
        'good': 0.80,
        'warning': 0.60,
    },
    'good_endings': {
        'good': 0.80,
        'patterns': [
            # 深掘り分析TOP16（出現率順）
            'ないこと',         # 14.9% — 否定確認
            'になること',       # 13.6% — 状態変化
            'されていること',   # 13.3% — 状態確認
            'できること',       # 12.1% — 操作可否
            'されること',       # 11.0% — 処理結果
            'していること',     #  8.5% — 動作確認
            'すること',         #  6.0% — 動作指示
            'であること',       #  4.3% — 状態断定
            'になっていること', #  1.5% — 状態変化確認
            'れること',         #  1.1% — 可能形
            'くこと',           #  0.9% — 動作（届く等）
            'むこと',           #  0.5% — 動作（進む等）
            'できていること',   #  0.0% — 過去完了
            'されないこと',     #        — 否定受動
        ],
    },
    'numbered_steps': {
        'good': 0.20,
    },
}

# 手順の構造化キーワード（深掘り分析で抽出）
STRUCTURED_STEP_KEYWORDS = ['【前提条件】', '【セットアップ】', '【実行手順】', '【前提】']

# 粒度チェックの減点上限
MAX_GRANULARITY_PENALTY = 15

# ガードレールチェック
MAX_GUARDRAIL_PENALTY = 10
MAX_MONEY_P1_PENALTY = 5
MAX_KANTEN_COVERAGE_PENALTY = 5

# チェック7: 画面名・フォーマット
MAX_SCREEN_FORMAT_PENALTY = 15
SCREEN_URL_PENALTY = 2
SCREEN_IMPL_PENALTY = 1
SCREEN_VAGUE_PENALTY = 2
FORMAT_NO_SECTION_PENALTY = 1
MAX_FORMAT_NO_SECTION_PENALTY = 3
FORMAT_SAME_LINE_PENALTY = 1
MAX_FORMAT_SAME_LINE_PENALTY = 3

# チェック8: 構造化
MAX_STRUCTURE_PENALTY = 10
GROUPING_PENALTY = 1       # グループ化は推奨（必須ではない）
MAX_GROUPING_PENALTY = 3
NO_DT_PENALTY = 2          # 小規模チケットではDT不要の場合もある
COMPLEX_PRECONDITION_PENALTY = 2

# チェック9: 観点シート品質
MAX_PERSPECTIVE_QUALITY_PENALTY = 10

# 金額関連キーワード（ガードレール qa-perspectives.md 1-13 準拠）
# 注意: "金額"単体は画面カラム名として頻出するため除外。金額「計算」に関するキーワードに限定
MONEY_KEYWORDS = [
    '金額計算', '金額変更', '按分', '消費税', '源泉', '丸め', '切捨', '切上', '四捨五入',
    '外貨', '為替レート', '外貨レート', '税抜', '税込', '課金', '請求額', '税額', '小数点',
    '端数', '仕訳', '定期支払', '金額確認',
]


def load_benchmark(benchmark_path=None):
    """benchmark.json を読み込み、BENCHMARKを返す。無ければデフォルト値を返す"""
    benchmark = dict(BENCHMARK_DEFAULT)

    if benchmark_path and os.path.exists(benchmark_path):
        with open(benchmark_path, 'r', encoding='utf-8') as f:
            loaded = json.load(f)
        # JSONから基準値を上書き
        if 'expected_length' in loaded:
            benchmark['expected_length'] = {
                'ideal_min': loaded['expected_length'].get('ideal_min', 10),
                'ideal_max': loaded['expected_length'].get('ideal_max', 80),
            }
        if 'steps_coverage' in loaded:
            benchmark['steps_coverage'] = {
                'good': loaded['steps_coverage'].get('good', 0.80),
                'warning': loaded['steps_coverage'].get('warning', 0.60),
            }
        if 'good_endings' in loaded:
            benchmark['good_endings'] = {
                'good': loaded['good_endings'].get('good', 0.80),
                'patterns': loaded['good_endings'].get('patterns', BENCHMARK_DEFAULT['good_endings']['patterns']),
            }
        if 'numbered_steps' in loaded:
            benchmark['numbered_steps'] = {
                'good': loaded['numbered_steps'].get('good', 0.20),
            }
        # メタ情報を保持
        benchmark['_meta'] = loaded.get('_meta', {})
        benchmark['per_file'] = loaded.get('per_file', [])
        return benchmark, True

    # デフォルト: スクリプトと同じフォルダの benchmark.json を探す
    script_dir = os.path.dirname(os.path.abspath(__file__))
    auto_path = os.path.join(script_dir, 'benchmark.json')
    if os.path.exists(auto_path):
        return load_benchmark(auto_path)

    return benchmark, False


# =====================================================================
# ユーティリティ
# =====================================================================

def read_cases(ws, col_map, data_start_row=3):
    """テストケースを読み込む"""
    cases = []
    no_col = col_map.get('no', 2)
    for row in range(data_start_row, ws.max_row + 1):
        no = ws.cell(row=row, column=no_col).value
        if no is None:
            continue
        case = {'no': no, 'row': row}
        for key, col in col_map.items():
            if key == 'no':
                continue
            case[key] = str(ws.cell(row=row, column=col).value or '')
        cases.append(case)
    return cases


def load_keywords(filepath):
    """NGワードファイルを読み込む（TSV形式: NG単語<TAB>正式名称）"""
    keywords = []
    if not filepath or not os.path.exists(filepath):
        return keywords
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = line.split('\t')
            if len(parts) >= 2:
                keywords.append((parts[0], parts[1]))
            else:
                keywords.append((parts[0], '（正式名称未定義）'))
    return keywords


# =====================================================================
# チェック関数
# =====================================================================

def check_vague_expressions(cases, col_map):
    """曖昧表現チェック"""
    issues = []
    if 'expected' not in col_map:
        return issues, 0

    for case in cases:
        exp = case.get('expected', '')
        if not exp or exp == '-':
            continue
        for pattern, penalty, message in VAGUE_PATTERNS:
            if re.search(pattern, exp):
                # 例外パターンに該当する場合は偽陽性としてスキップ
                if is_vague_exception(exp):
                    continue
                issues.append({
                    'no': case['no'],
                    'row': case['row'],
                    'field': '期待値',
                    'value': exp[:80],
                    'penalty': penalty,
                    'message': message,
                })

    total_penalty = min(sum(i['penalty'] for i in issues), MAX_VAGUE_PENALTY)
    return issues, total_penalty


def check_multi_koto(cases, col_map):
    """期待値の複数「こと」チェック"""
    issues = []
    if 'expected' not in col_map:
        return issues, 0

    for case in cases:
        exp = case.get('expected', '')
        if not exp or exp == '-':
            continue

        # 括弧内の補足コメントを除外してから「こと」をカウント
        exp_without_parens = strip_parens(exp)
        matches = re.findall(r'こと', exp_without_parens)
        if len(matches) >= 2:
            issues.append({
                'no': case['no'],
                'row': case['row'],
                'field': '期待値',
                'value': exp[:80],
                'count': len(matches),
                'penalty': MULTI_KOTO_PENALTY,
            })

    total_penalty = min(sum(i['penalty'] for i in issues), MAX_MULTI_KOTO_PENALTY)
    return issues, total_penalty


def check_custom_words(cases, col_map, keywords, ws_kanten=None):
    """我流単語チェック"""
    issues = []
    if not keywords:
        return issues, 0

    text_fields = [k for k in col_map.keys() if k != 'no']

    # テストケースシートチェック
    for case in cases:
        for field in text_fields:
            val = case.get(field, '')
            if not val or val == '-':
                continue
            for ng_word, correct_word in keywords:
                if ng_word in val:
                    issues.append({
                        'no': case['no'],
                        'row': case['row'],
                        'sheet': 'テストケース',
                        'field': field,
                        'ng_word': ng_word,
                        'correct_word': correct_word,
                        'value': val[:60],
                        'penalty': CUSTOM_WORD_PENALTY,
                    })

    # 観点シートチェック
    if ws_kanten:
        for row in range(1, ws_kanten.max_row + 1):
            for col in range(1, ws_kanten.max_column + 1):
                val = ws_kanten.cell(row=row, column=col).value
                if not val or not isinstance(val, str):
                    continue
                for ng_word, correct_word in keywords:
                    if ng_word in val:
                        issues.append({
                            'no': f'観点Row{row}',
                            'row': row,
                            'sheet': '観点',
                            'field': f'Col{col}',
                            'ng_word': ng_word,
                            'correct_word': correct_word,
                            'value': val[:60],
                            'penalty': CUSTOM_WORD_PENALTY,
                        })

    total_penalty = min(sum(i['penalty'] for i in issues), MAX_CUSTOM_WORD_PENALTY)
    return issues, total_penalty


def check_missing_steps(cases, col_map):
    """手順省略チェック

    グループ構造を考慮:
    - 同じ「確認すること」カテゴリ内で、最初のケースに手順があり
      後続のケースが空欄の場合は「グループ共有手順」として許容する
    - カテゴリ内の全ケースが手順空欄の場合のみ検出する
    """
    issues = []
    if 'steps' not in col_map:
        return issues, 0

    # カテゴリごとにグループ化
    groups = {}  # confirm -> [cases]
    for case in cases:
        confirm = case.get('confirm', '').strip()
        if confirm not in groups:
            groups[confirm] = []
        groups[confirm].append(case)

    for confirm, group_cases in groups.items():
        # グループ内に手順ありのケースがあるか
        has_steps_in_group = any(
            (c.get('steps', '') or '').strip() not in ('', '-', 'None')
            for c in group_cases
        )

        if has_steps_in_group:
            # グループ内に手順ありケースがある場合、空欄は許容（グループ共有）
            continue

        # グループ全体で手順なし → 全件を検出
        for case in group_cases:
            steps = case.get('steps', '')
            if not steps or steps.strip() == '' or steps.strip() == '-':
                issues.append({
                    'no': case['no'],
                    'row': case['row'],
                    'field': '手順',
                    'penalty': MISSING_STEPS_PENALTY,
                    'message': f'手順が空欄です（カテゴリ「{confirm[:20]}」内に手順ありケースなし）',
                })

    total_penalty = min(sum(i['penalty'] for i in issues), MAX_MISSING_STEPS_PENALTY)
    return issues, total_penalty


def check_granularity(cases, col_map, benchmark=None):
    """粒度・品質チェック（過去参考例の基準値と比較して自動採点）

    チェック項目（計15点満点）:
    - 期待値の具体性:  期待値が短すぎ/長すぎのケース比率（5点）
    - 手順の詳細度:    番号付きステップの使用率（5点）
    - 期待値の末尾表現: 参考例で頻出の良好パターンとの一致率（5点）
    """
    if benchmark is None:
        benchmark = BENCHMARK_DEFAULT
    stats = {
        'total': len(cases),
        'with_steps': 0,
        'with_preconditions': 0,
        'with_numbered_steps': 0,
        'with_structured_steps': 0,  # 構造化キーワードを含む手順
        'avg_steps_length': 0,
        'avg_expected_length': 0,
        'short_expected': 0,       # 期待値が短すぎるケース数
        'long_expected': 0,        # 期待値が長すぎるケース数
        'good_ending_count': 0,    # 良好な末尾表現のケース数
        'expected_count': 0,       # 期待値ありのケース数
    }
    issues = []

    steps_lengths = []
    expected_lengths = []

    for case in cases:
        steps = case.get('steps', '')
        expected = case.get('expected', '')

        # 手順の統計
        if steps and steps.strip() not in ('', '-', 'None'):
            stats['with_steps'] += 1
            steps_lengths.append(len(steps))
        if '前提条件' in steps or '【前提】' in steps or '【前提条件】' in steps:
            stats['with_preconditions'] += 1
        if any(kw in steps for kw in STRUCTURED_STEP_KEYWORDS):
            stats['with_structured_steps'] += 1
        if re.search(r'(?:^|\n)\s*[1１①⑴\(1\)][\.\．\s、）\)]', steps):
            stats['with_numbered_steps'] += 1

        # 期待値の統計 — 「同上」は除外（末尾パターン不一致の誤減点防止）
        if expected and expected.strip() not in ('', '-', 'None', '同上'):
            stats['expected_count'] += 1
            exp_len = len(expected)
            expected_lengths.append(exp_len)

            # 短すぎ/長すぎ判定
            if exp_len < benchmark['expected_length']['ideal_min']:
                stats['short_expected'] += 1
                if exp_len <= 5:
                    issues.append({
                        'no': case['no'],
                        'type': '期待値が短すぎ',
                        'detail': f'{exp_len}字「{expected[:30]}」→ 具体性が不足している可能性',
                    })
            if exp_len > benchmark['expected_length']['ideal_max']:
                stats['long_expected'] += 1
                issues.append({
                    'no': case['no'],
                    'type': '期待値が長すぎ',
                    'detail': f'{exp_len}字 → 複数の期待値が混在している可能性',
                })

            # 良好な末尾表現チェック
            good_patterns = benchmark.get('good_endings', {})
            if isinstance(good_patterns, dict):
                ending_list = good_patterns.get('patterns', [])
            else:
                ending_list = good_patterns
            for ending in ending_list:
                if expected.rstrip().endswith(ending):
                    stats['good_ending_count'] += 1
                    break

    if steps_lengths:
        stats['avg_steps_length'] = sum(steps_lengths) / len(steps_lengths)
    if expected_lengths:
        stats['avg_expected_length'] = sum(expected_lengths) / len(expected_lengths)

    # --- 採点 ---
    total = max(stats['total'], 1)
    exp_count = max(stats['expected_count'], 1)
    penalty = 0

    # (1) 期待値の具体性（5点）: 短すぎ+長すぎが全体の10%以上で減点
    bad_length_rate = (stats['short_expected'] + stats['long_expected']) / exp_count
    if bad_length_rate > 0.20:
        penalty += 5
        issues.append({'no': '-', 'type': '期待値の長さ', 'detail': f'短すぎ/長すぎが{bad_length_rate*100:.0f}%（基準: 10〜60字）'})
    elif bad_length_rate > 0.10:
        penalty += 3
        issues.append({'no': '-', 'type': '期待値の長さ', 'detail': f'短すぎ/長すぎが{bad_length_rate*100:.0f}%（基準: 10〜60字）'})
    elif bad_length_rate > 0.05:
        penalty += 1

    # (2) 手順の詳細度（5点）: 番号付きステップの使用率
    # 参考例: 大規模26% / バグ修正100% / v7 43% → 20%以上を良好とする
    numbered_rate = stats['with_numbered_steps'] / total
    # 手順記載率はグループ構造を考慮する
    # （同カテゴリ内で1件でも手順があれば、グループ全体を「手順あり」とみなす）
    if 'confirm' in col_map:
        groups = {}
        for case in cases:
            confirm = case.get('confirm', '').strip()
            if confirm not in groups:
                groups[confirm] = {'total': 0, 'has_steps': False}
            groups[confirm]['total'] += 1
            steps_val = (case.get('steps', '') or '').strip()
            if steps_val and steps_val not in ('-', 'None'):
                groups[confirm]['has_steps'] = True
        covered = sum(g['total'] for g in groups.values() if g['has_steps'])
        steps_rate = covered / total
    else:
        steps_rate = stats['with_steps'] / total

    if steps_rate < benchmark['steps_coverage']['warning']:
        penalty += 5
        issues.append({'no': '-', 'type': '手順の記載率', 'detail': f'手順あり{steps_rate*100:.0f}%（基準: 80%以上、グループ共有考慮済み）'})
    elif steps_rate < benchmark['steps_coverage']['good']:
        penalty += 3
        issues.append({'no': '-', 'type': '手順の記載率', 'detail': f'手順あり{steps_rate*100:.0f}%（基準: 80%以上、グループ共有考慮済み）'})
    elif numbered_rate < 0.15:
        penalty += 2
        issues.append({'no': '-', 'type': '番号付きステップ', 'detail': f'番号付き{numbered_rate*100:.0f}%（基準: 20%以上）'})

    # (3) 期待値の末尾表現（5点）: 良好パターンの一致率
    good_ending_threshold = benchmark.get('good_endings', {})
    if isinstance(good_ending_threshold, dict):
        ge_good = good_ending_threshold.get('good', 0.80)
    else:
        ge_good = 0.80
    good_ending_rate = stats['good_ending_count'] / exp_count
    if good_ending_rate < ge_good - 0.10:
        penalty += 5
        issues.append({'no': '-', 'type': '末尾表現', 'detail': f'良好パターン{good_ending_rate*100:.0f}%（基準: {ge_good*100:.0f}%以上）'})
    elif good_ending_rate < ge_good:
        penalty += 3
        issues.append({'no': '-', 'type': '末尾表現', 'detail': f'良好パターン{good_ending_rate*100:.0f}%（基準: {ge_good*100:.0f}%以上）'})
    elif good_ending_rate < ge_good + 0.10:
        penalty += 1

    penalty = min(penalty, MAX_GRANULARITY_PENALTY)
    stats['issues'] = issues
    stats['penalty'] = penalty
    stats['numbered_rate'] = numbered_rate
    stats['good_ending_rate'] = good_ending_rate
    stats['_benchmark_per_file'] = benchmark.get('per_file', [])

    return stats


# =====================================================================
# ガードレールチェック（qa-perspectives.md 準拠）
# =====================================================================

def check_money_priority(cases, col_map):
    """金額・セキュリティ関連テストのP1優先度チェック

    ガードレール基準:
    - 金額計算に関するテストは原則P1とする
    - セキュリティ関連テスト（認証/権限/脆弱性）も原則P1とする
    """
    issues = []
    priority_col = col_map.get('priority')

    # 金額関連ケースを検出
    money_cases = []
    security_cases = []
    text_fields = [k for k in ('confirm', 'target', 'detail', 'expected') if k in col_map]
    for case in cases:
        is_money = False
        is_security = False
        for field in text_fields:
            val = case.get(field, '')
            if any(kw in val for kw in MONEY_KEYWORDS):
                is_money = True
            if any(kw in val for kw in SECURITY_KEYWORDS):
                is_security = True
        if is_money:
            money_cases.append(case)
        elif is_security:
            security_cases.append(case)

    p1_required_cases = money_cases + security_cases
    if not p1_required_cases:
        return {'money_cases': 0, 'security_cases': 0, 'issues': [], 'penalty': 0}

    if priority_col is None:
        msg_parts = []
        if money_cases:
            msg_parts.append(f'金額{len(money_cases)}件')
        if security_cases:
            msg_parts.append(f'セキュリティ{len(security_cases)}件')
        issues.append({
            'type': 'no_priority_column',
            'message': f'{"+".join(msg_parts)}検出。優先度列がないためP1設定を確認できません',
        })
        return {'money_cases': len(money_cases), 'security_cases': len(security_cases),
                'issues': issues, 'penalty': 2}

    # P1でないケースを検出
    non_p1 = []
    for case in p1_required_cases:
        priority = str(case.get('priority', '')).strip().upper()
        if priority not in ('P1', '1', 'HIGH', '高'):
            case_type = '金額' if case in money_cases else 'セキュリティ'
            non_p1.append((case, case_type))

    if non_p1:
        for case, case_type in non_p1[:5]:
            pri_val = case.get('priority', '').strip() or '(空)'
            issues.append({
                'type': 'priority_not_p1',
                'no': case['no'],
                'row': case['row'],
                'priority': pri_val,
                'message': f'No.{case["no"]}: {case_type}関連テストがP1ではありません (現在: {pri_val})',
            })
        if len(non_p1) > 5:
            issues.append({
                'type': 'priority_not_p1_more',
                'message': f'... 他 {len(non_p1) - 5}件',
            })

    penalty = min(len(non_p1), MAX_MONEY_P1_PENALTY)
    return {'money_cases': len(money_cases), 'security_cases': len(security_cases),
            'non_p1': len(non_p1), 'issues': issues, 'penalty': penalty}


def check_kanten_tc_coverage(cases, col_map, ws_kanten):
    """観点シートの○項目とTCの整合性チェック

    観点シートで○（対象）とマークされた項目に対応するTCが存在するか確認する。
    """
    issues = []
    if not ws_kanten:
        return {'checked': False, 'issues': [], 'penalty': 0}

    # 観点シートから○のついた行を抽出
    # 行1-26はヘッダ領域（環境/OS/ブラウザ/アプリ/権限が混在）のためスキップ
    # 行27以降の観点セクション（リスト/マトリクス/DT/チェックリスト）のみ対象
    PERSPECTIVE_SECTION_START = 27
    marked_perspectives = []
    for row in range(PERSPECTIVE_SECTION_START, ws_kanten.max_row + 1):
        row_has_mark = False
        row_texts = []
        for col in range(1, min(ws_kanten.max_column + 1, 20)):
            val = ws_kanten.cell(row=row, column=col).value
            if val is None:
                continue
            val_str = str(val).strip()
            if val_str in ('○', '◯', '〇', 'O'):
                row_has_mark = True
            elif len(val_str) > 2:
                row_texts.append(val_str)

        if row_has_mark and row_texts:
            combined = ' '.join(row_texts)
            marked_perspectives.append({
                'row': row,
                'text': combined,
            })

    if not marked_perspectives:
        return {'checked': True, 'marked_count': 0, 'uncovered_count': 0, 'issues': [], 'penalty': 0}

    # TCの全テキストを連結（検索用）
    tc_text_parts = []
    text_fields = [k for k in ('confirm', 'target', 'detail', 'expected', 'steps') if k in col_map]
    for case in cases:
        for field in text_fields:
            val = case.get(field, '')
            if val:
                tc_text_parts.append(val)
    tc_text = ' '.join(tc_text_parts)

    # 各○観点がTCでカバーされているか確認
    uncovered = []
    for persp in marked_perspectives:
        # 観点テキストからプレフィックス（TH_/WDL_等）を除去し、2文字以上の日本語単語を抽出
        clean_text = re.sub(r'[A-Z]{2,}_', '', persp['text'])
        words = re.findall(r'[\u3040-\u9fff]{2,}', clean_text)
        if not words:
            continue
        # いずれかのキーワードがTCテキストに含まれていればカバー済み
        covered = any(word in tc_text for word in words)
        if not covered:
            uncovered.append(persp)

    for persp in uncovered[:5]:
        issues.append({
            'type': 'uncovered_perspective',
            'row': persp['row'],
            'text': persp['text'][:60],
            'message': f'観点シートRow{persp["row"]}: ○だがTCに対応ケースなし「{persp["text"][:40]}」',
        })
    if len(uncovered) > 5:
        issues.append({
            'type': 'uncovered_more',
            'message': f'... 他 {len(uncovered) - 5}件',
        })

    penalty = min(len(uncovered) * 2, MAX_KANTEN_COVERAGE_PENALTY)
    return {
        'checked': True,
        'marked_count': len(marked_perspectives),
        'uncovered_count': len(uncovered),
        'issues': issues,
        'penalty': penalty,
    }


# =====================================================================
# チェック7: 画面名・フォーマット
# =====================================================================

def check_screen_and_format(cases, col_map):
    """画面名とフォーマットのチェック

    7a. 画面名フォーマット（上限-8点）:
        - URL混入: 画面列に /path が含まれる → -2点/件
        - 実装用語: モーダル/ドロワー等 → -1点/件
        - 曖昧画面名: 全画面共通等 → -2点/件

    7b. 手順フォーマット（上限-7点）:
        - 【前提条件】【実行手順】分離なし → -1点/件（上限-4点）
        - 同一行に【前提条件】【実行手順】混在 → -1点/件（上限-3点）
    """
    issues = []
    penalty_7a = 0
    penalty_7b_nosection = 0
    penalty_7b_sameline = 0

    # --- 7a: 画面名チェック ---
    if 'screen' in col_map:
        for case in cases:
            screen = case.get('screen', '')
            if not screen or screen.strip() in ('', '-', 'None'):
                continue

            # URL混入
            if re.search(SCREEN_NAME_URL_PATTERN, screen):
                penalty_7a += SCREEN_URL_PENALTY
                issues.append({
                    'type': 'screen_url',
                    'no': case['no'],
                    'row': case['row'],
                    'value': screen[:60],
                    'message': f'No.{case["no"]}: 画面名にURLパスが混入「{screen[:40]}」→ 画面名のみ記載',
                })

            # 実装用語
            for term in SCREEN_NAME_IMPL_TERMS:
                if term in screen:
                    penalty_7a += SCREEN_IMPL_PENALTY
                    issues.append({
                        'type': 'screen_impl',
                        'no': case['no'],
                        'row': case['row'],
                        'value': screen[:60],
                        'message': f'No.{case["no"]}: 実装用語「{term}」→ ユーザー視点の画面名に変更',
                    })
                    break  # 1件につき1回のみ

            # 曖昧画面名
            for vague in SCREEN_NAME_VAGUE:
                if vague in screen:
                    penalty_7a += SCREEN_VAGUE_PENALTY
                    issues.append({
                        'type': 'screen_vague',
                        'no': case['no'],
                        'row': case['row'],
                        'value': screen[:60],
                        'message': f'No.{case["no"]}: 曖昧な画面名「{vague}」→ 具体的な画面名を記載',
                    })
                    break

    penalty_7a = min(penalty_7a, 8)  # 7a上限

    # --- 7b: 手順フォーマットチェック ---
    if 'steps' in col_map:
        for case in cases:
            steps = case.get('steps', '')
            if not steps or steps.strip() in ('', '-', 'None'):
                continue

            has_any_section = any(kw in steps for kw in FORMAT_PRECONDITION_KEYWORDS + FORMAT_EXECUTION_KEYWORDS)

            # 手順に内容があるが【】セクション分離なし
            # 30文字以上かつ改行を含む（複数行手順）の場合のみ検出
            if not has_any_section and len(steps.strip()) > 30 and '\n' in steps:
                penalty_7b_nosection += FORMAT_NO_SECTION_PENALTY
                issues.append({
                    'type': 'format_no_section',
                    'no': case['no'],
                    'row': case['row'],
                    'message': f'No.{case["no"]}: 手順に【前提条件】【実行手順】のセクション分離なし',
                })

            # 同一行に【前提条件】と【実行手順】が混在（改行なし連結）
            has_precond = any(kw in steps for kw in FORMAT_PRECONDITION_KEYWORDS)
            has_exec = any(kw in steps for kw in FORMAT_EXECUTION_KEYWORDS)
            if has_precond and has_exec:
                # 改行で分離されているか確認
                lines = steps.split('\n')
                for line in lines:
                    if any(kw in line for kw in FORMAT_PRECONDITION_KEYWORDS) and \
                       any(kw in line for kw in FORMAT_EXECUTION_KEYWORDS):
                        penalty_7b_sameline += FORMAT_SAME_LINE_PENALTY
                        issues.append({
                            'type': 'format_same_line',
                            'no': case['no'],
                            'row': case['row'],
                            'message': f'No.{case["no"]}: 【前提条件】と【実行手順】が同一行に混在→改行で分離',
                        })
                        break

    penalty_7b_nosection = min(penalty_7b_nosection, MAX_FORMAT_NO_SECTION_PENALTY)
    penalty_7b_sameline = min(penalty_7b_sameline, MAX_FORMAT_SAME_LINE_PENALTY)
    penalty_7b = penalty_7b_nosection + penalty_7b_sameline
    penalty_7b = min(penalty_7b, 7)  # 7b上限

    total_penalty = min(penalty_7a + penalty_7b, MAX_SCREEN_FORMAT_PENALTY)
    return {
        'issues': issues,
        'penalty': total_penalty,
        'penalty_7a': min(penalty_7a, 8),
        'penalty_7b': min(penalty_7b, 7),
    }


# =====================================================================
# チェック8: 構造化（グループTC・DT・シナリオ）
# =====================================================================

def check_structure(cases, col_map, wb):
    """構造化チェック

    8a. グループ化TC検証（上限-5点）:
        同一「確認すること」内で期待値3件以上あるのにグループ化（手順「-」）なし → -2点/グループ

    8b. デシジョンテーブル存在確認（-3点）:
        シート一覧に「デシジョン」「DT」「条件表」がなく、観点シートにもDT領域なし → -3点

    8c. テストシナリオシート確認（-2点）:
        複雑な前提条件がある場合にシナリオシートなし → -2点
    """
    issues = []
    penalty_8a = 0
    penalty_8b = 0
    penalty_8c = 0

    # --- 8a: グループ化TC検証 ---
    # 同一「確認すること」内で手順が類似している（共通化すべき）ケースのみ検知
    if 'confirm' in col_map:
        groups = {}
        for case in cases:
            confirm = case.get('confirm', '').strip()
            if not confirm:
                continue
            if confirm not in groups:
                groups[confirm] = []
            groups[confirm].append(case)

        for confirm, group_cases in groups.items():
            if len(group_cases) < 5:  # 5件以上のグループのみ対象
                continue
            # 全行に手順が記載されている（グループ化されていない）
            steps_list = [
                (c.get('steps', '') or '').strip()
                for c in group_cases
                if (c.get('steps', '') or '').strip() not in ('', '-', 'None')
            ]
            if len(steps_list) < 5:
                continue
            # 手順の類似度チェック: 最初の1行が同じ手順が50%以上なら共通化すべき
            first_lines = [s.split('\n')[0].strip() for s in steps_list]
            most_common_count = Counter(first_lines).most_common(1)[0][1]
            similarity_rate = most_common_count / len(first_lines)
            if similarity_rate >= 0.50:
                penalty_8a += GROUPING_PENALTY
                issues.append({
                    'type': 'no_grouping',
                    'confirm': confirm[:30],
                    'count': len(group_cases),
                    'message': f'「{confirm[:30]}」: {len(group_cases)}件で手順の{similarity_rate*100:.0f}%が類似→ 共通手順は親行にまとめ子行は「-」',
                })

    penalty_8a = min(penalty_8a, MAX_GROUPING_PENALTY)

    # --- 8b: デシジョンテーブル存在確認 ---
    dt_sheet_keywords = ['デシジョン', 'DT', '条件表', 'decision']
    has_dt_sheet = False
    if wb:
        for sheet_name in wb.sheetnames:
            if any(kw in sheet_name for kw in dt_sheet_keywords):
                has_dt_sheet = True
                break

    if not has_dt_sheet:
        # 観点シート内のDT領域も確認
        has_dt_in_kanten = False
        kanten_candidates = ['観点', '観点シート', 'テスト観点']
        ws_kanten = find_sheet(wb, kanten_candidates) if wb else None
        if ws_kanten:
            for row in range(1, min(ws_kanten.max_row + 1, 200)):
                for col in range(1, min(ws_kanten.max_column + 1, 20)):
                    val = ws_kanten.cell(row=row, column=col).value
                    if val and isinstance(val, str):
                        if any(kw in val for kw in ['デシジョン', 'DT', '条件表', '条件/動作']):
                            has_dt_in_kanten = True
                            break
                if has_dt_in_kanten:
                    break

        if not has_dt_in_kanten:
            penalty_8b = NO_DT_PENALTY
            issues.append({
                'type': 'no_dt',
                'message': 'デシジョンテーブルが見つかりません（シート/観点シート内ともになし）',
            })

    # --- 8c: シナリオシート確認 ---
    scenario_keywords = ['シナリオ', '開発側連携']
    has_scenario = False
    if wb:
        for sheet_name in wb.sheetnames:
            if any(kw in sheet_name for kw in scenario_keywords):
                has_scenario = True
                break

    if not has_scenario:
        # 複雑な前提条件の有無を推定（Railsコマンド、SQL、API呼び出し等）
        complex_keywords = ['rails', 'rake', 'SQL', 'curl', 'API', 'コンソール', 'DB操作', 'マイグレーション']
        has_complex = False
        for case in cases:
            steps = case.get('steps', '')
            if steps and any(kw.lower() in steps.lower() for kw in complex_keywords):
                has_complex = True
                break

        if has_complex:
            penalty_8c = COMPLEX_PRECONDITION_PENALTY
            issues.append({
                'type': 'no_scenario',
                'message': '複雑な前提条件（Rails/SQL/API等）がありますがシナリオシートがありません→ 開発側連携用のシナリオシートの追加を推奨',
            })

    total_penalty = min(penalty_8a + penalty_8b + penalty_8c, MAX_STRUCTURE_PENALTY)
    return {
        'issues': issues,
        'penalty': total_penalty,
        'penalty_8a': penalty_8a,
        'penalty_8b': penalty_8b,
        'penalty_8c': penalty_8c,
    }


# =====================================================================
# チェック9: 観点シート品質
# =====================================================================

def check_perspective_quality(ws_kanten):
    """観点シート品質チェック

    9a. 記入密度（上限-6点）:
        L3（画面列）空欄率20%超→-3点、40%超→-6点
        L5（詳細列）空欄率40%超→-2点、60%超→-4点

    9b. 対象外理由チェック（上限-2点）:
        「対象外」「×」「-」マーク行で理由列が空欄 → -1点/件

    9c. パンくず形式チェック（上限-2点）:
        リグレッション行でパンくず率50%未満 → -2点
    """
    issues = []
    if not ws_kanten:
        return {'checked': False, 'issues': [], 'penalty': 0}

    penalty_9a = 0
    penalty_9b = 0
    penalty_9c = 0

    PERSPECTIVE_SECTION_START = 27

    # --- 観点シートの列位置を推定 ---
    # L3=画面列, L5=詳細列を検出
    # ヘッダー行（Row26付近）から列位置を特定
    screen_col = None
    detail_col = None
    reason_col = None
    mark_col = None

    for header_row in range(24, 28):
        for col in range(1, min(ws_kanten.max_column + 1, 30)):
            val = ws_kanten.cell(row=header_row, column=col).value
            if not val:
                continue
            val_str = str(val).strip()
            if '画面' in val_str and screen_col is None:
                screen_col = col
            if ('詳細' in val_str or '内容' in val_str) and detail_col is None:
                detail_col = col
            if '理由' in val_str and reason_col is None:
                reason_col = col
            if ('対象' in val_str and '対象外' not in val_str) and mark_col is None:
                mark_col = col

    # --- 9a: 記入密度 ---
    total_rows = 0
    screen_empty = 0
    detail_empty = 0

    for row in range(PERSPECTIVE_SECTION_START, min(ws_kanten.max_row + 1, 300)):
        # 空行スキップ（全セルが空なら終了と見なす）
        row_has_data = False
        for col in range(1, min(ws_kanten.max_column + 1, 20)):
            if ws_kanten.cell(row=row, column=col).value:
                row_has_data = True
                break
        if not row_has_data:
            continue

        total_rows += 1

        if screen_col:
            val = ws_kanten.cell(row=row, column=screen_col).value
            if not val or str(val).strip() == '':
                screen_empty += 1

        if detail_col:
            val = ws_kanten.cell(row=row, column=detail_col).value
            if not val or str(val).strip() == '':
                detail_empty += 1

    if total_rows > 0:
        if screen_col:
            screen_empty_rate = screen_empty / total_rows
            if screen_empty_rate > 0.40:
                penalty_9a += 6
                issues.append({
                    'type': 'screen_density',
                    'message': f'観点シート画面列: 空欄率{screen_empty_rate*100:.0f}%（基準: 20%以下）',
                })
            elif screen_empty_rate > 0.20:
                penalty_9a += 3
                issues.append({
                    'type': 'screen_density',
                    'message': f'観点シート画面列: 空欄率{screen_empty_rate*100:.0f}%（基準: 20%以下）',
                })

        if detail_col:
            detail_empty_rate = detail_empty / total_rows
            if detail_empty_rate > 0.60:
                penalty_9a += 4
                issues.append({
                    'type': 'detail_density',
                    'message': f'観点シート詳細列: 空欄率{detail_empty_rate*100:.0f}%（基準: 40%以下）',
                })
            elif detail_empty_rate > 0.40:
                penalty_9a += 2
                issues.append({
                    'type': 'detail_density',
                    'message': f'観点シート詳細列: 空欄率{detail_empty_rate*100:.0f}%（基準: 40%以下）',
                })

    penalty_9a = min(penalty_9a, 6)

    # --- 9b: 対象外理由チェック ---
    exclusion_markers = ['対象外', '×', '✕', '☓']
    missing_reason_count = 0
    for row in range(PERSPECTIVE_SECTION_START, min(ws_kanten.max_row + 1, 300)):
        row_has_exclusion = False
        for col in range(1, min(ws_kanten.max_column + 1, 20)):
            val = ws_kanten.cell(row=row, column=col).value
            if val and str(val).strip() in exclusion_markers:
                row_has_exclusion = True
                break
            if val and isinstance(val, str) and val.strip() == '-':
                # 「-」はマーク列（mark_col近辺）でのみ対象外と判定
                if mark_col and abs(col - mark_col) <= 2:
                    row_has_exclusion = True
                    break

        if row_has_exclusion and reason_col:
            reason_val = ws_kanten.cell(row=row, column=reason_col).value
            if not reason_val or str(reason_val).strip() == '':
                missing_reason_count += 1
                if missing_reason_count <= 3:
                    issues.append({
                        'type': 'no_exclusion_reason',
                        'row': row,
                        'message': f'観点シートRow{row}: 対象外だが理由が空欄',
                    })

    if missing_reason_count > 3:
        issues.append({
            'type': 'no_exclusion_reason_more',
            'message': f'... 他 {missing_reason_count - 3}件の対象外理由なし',
        })
    penalty_9b = min(missing_reason_count, 2)

    # --- 9c: パンくず形式チェック ---
    regression_keywords = ['影響', '既存', 'リグレッション', '回帰']
    regression_rows = 0
    breadcrumb_count = 0

    for row in range(PERSPECTIVE_SECTION_START, min(ws_kanten.max_row + 1, 300)):
        row_text = ''
        for col in range(1, min(ws_kanten.max_column + 1, 20)):
            val = ws_kanten.cell(row=row, column=col).value
            if val:
                row_text += str(val)

        if any(kw in row_text for kw in regression_keywords):
            regression_rows += 1
            if screen_col:
                screen_val = ws_kanten.cell(row=row, column=screen_col).value
                if screen_val and isinstance(screen_val, str):
                    if '>' in screen_val or '→' in screen_val:
                        breadcrumb_count += 1

    if regression_rows > 0:
        breadcrumb_rate = breadcrumb_count / regression_rows
        if breadcrumb_rate < 0.50:
            penalty_9c = 2
            issues.append({
                'type': 'breadcrumb',
                'message': f'リグレッション行{regression_rows}件中パンくず形式{breadcrumb_count}件（{breadcrumb_rate*100:.0f}%、基準: 50%以上）',
            })

    total_penalty = min(penalty_9a + penalty_9b + penalty_9c, MAX_PERSPECTIVE_QUALITY_PENALTY)
    return {
        'checked': True,
        'issues': issues,
        'penalty': total_penalty,
        'penalty_9a': penalty_9a,
        'penalty_9b': penalty_9b,
        'penalty_9c': penalty_9c,
        'total_rows': total_rows,
    }


# =====================================================================
# レポート生成
# =====================================================================

def generate_report(filepath, sheet_results, benchmark, screen_format_result=None,
                     structure_result=None, perspective_quality_result=None):
    """評価レポートを生成（複数シート対応）

    sheet_results: list of dict, 各要素は以下のキーを持つ:
        sheet_name, cases, vague_issues, vague_penalty,
        multi_issues, multi_penalty, word_issues, word_penalty,
        steps_issues, steps_penalty, granularity_stats,
        guardrail_money, guardrail_kanten
    screen_format_result: check_screen_and_format() の結果（全シート合算済み）
    structure_result: check_structure() の結果
    perspective_quality_result: check_perspective_quality() の結果
    """

    # 全シート合算
    total_cases = sum(len(sr['cases']) for sr in sheet_results)
    total_vague_penalty = 0
    total_multi_penalty = 0
    total_word_penalty = 0
    total_steps_penalty = 0
    total_granularity_penalty = 0
    total_guardrail_money_penalty = 0
    total_guardrail_kanten_penalty = 0
    all_vague_issues = []
    all_multi_issues = []
    all_word_issues = []
    all_steps_issues = []
    all_granularity_issues = []
    all_guardrail_money_issues = []
    all_guardrail_kanten_issues = []

    for sr in sheet_results:
        total_vague_penalty += sr['vague_penalty']
        total_multi_penalty += sr['multi_penalty']
        total_word_penalty += sr['word_penalty']
        total_steps_penalty += sr['steps_penalty']
        total_granularity_penalty += sr['granularity_stats'].get('penalty', 0)
        total_guardrail_money_penalty += sr.get('guardrail_money', {}).get('penalty', 0)
        total_guardrail_kanten_penalty += sr.get('guardrail_kanten', {}).get('penalty', 0)
        # シート名をissueに付与
        for issue in sr['vague_issues']:
            issue['_sheet'] = sr['sheet_name']
            all_vague_issues.append(issue)
        for issue in sr['multi_issues']:
            issue['_sheet'] = sr['sheet_name']
            all_multi_issues.append(issue)
        for issue in sr['word_issues']:
            issue['_sheet'] = sr['sheet_name']
            all_word_issues.append(issue)
        for issue in sr['steps_issues']:
            issue['_sheet'] = sr['sheet_name']
            all_steps_issues.append(issue)
        for issue in sr['granularity_stats'].get('issues', []):
            issue['_sheet'] = sr['sheet_name']
            all_granularity_issues.append(issue)
        for issue in sr.get('guardrail_money', {}).get('issues', []):
            issue['_sheet'] = sr['sheet_name']
            all_guardrail_money_issues.append(issue)
        for issue in sr.get('guardrail_kanten', {}).get('issues', []):
            issue['_sheet'] = sr['sheet_name']
            all_guardrail_kanten_issues.append(issue)

    # 上限適用（全シート合算後）
    total_vague_penalty = min(total_vague_penalty, MAX_VAGUE_PENALTY)
    total_multi_penalty = min(total_multi_penalty, MAX_MULTI_KOTO_PENALTY)
    total_word_penalty = min(total_word_penalty, MAX_CUSTOM_WORD_PENALTY)
    total_steps_penalty = min(total_steps_penalty, MAX_MISSING_STEPS_PENALTY)
    total_granularity_penalty = min(total_granularity_penalty, MAX_GRANULARITY_PENALTY)
    total_guardrail_money_penalty = min(total_guardrail_money_penalty, MAX_MONEY_P1_PENALTY)
    total_guardrail_kanten_penalty = min(total_guardrail_kanten_penalty, MAX_KANTEN_COVERAGE_PENALTY)
    total_guardrail_penalty = min(
        total_guardrail_money_penalty + total_guardrail_kanten_penalty,
        MAX_GUARDRAIL_PENALTY
    )

    # 新チェック7-9のペナルティ
    total_screen_format_penalty = screen_format_result.get('penalty', 0) if screen_format_result else 0
    total_structure_penalty = structure_result.get('penalty', 0) if structure_result else 0
    total_perspective_quality_penalty = perspective_quality_result.get('penalty', 0) if perspective_quality_result else 0

    total_penalty = (total_vague_penalty + total_multi_penalty +
                     total_word_penalty + total_steps_penalty +
                     total_granularity_penalty + total_guardrail_penalty +
                     total_screen_format_penalty + total_structure_penalty +
                     total_perspective_quality_penalty)
    score = max(0, 100 - total_penalty)

    multi_sheet = len(sheet_results) > 1

    lines = []
    lines.append("=" * 70)
    lines.append("テストケース評価レポート")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"ファイル: {os.path.basename(filepath)}")
    if multi_sheet:
        lines.append(f"評価シート数: {len(sheet_results)}シート")
        sheet_names = [sr['sheet_name'] for sr in sheet_results]
        lines.append(f"評価シート: {', '.join(sheet_names)}")
    else:
        lines.append(f"評価シート: {sheet_results[0]['sheet_name']}")
    lines.append(f"総ケース数: {total_cases}件")
    lines.append(f"評価日: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append(f"★ 総合スコア: {score}/100点")
    lines.append("")

    # シート別ケース数サマリー（複数シートの場合）
    if multi_sheet:
        lines.append("-" * 70)
        lines.append("シート別ケース数")
        lines.append("-" * 70)
        for sr in sheet_results:
            lines.append(f"  {sr['sheet_name']}: {len(sr['cases'])}件")
        lines.append("")

    # 減点サマリー
    lines.append("-" * 70)
    lines.append("減点サマリー")
    lines.append("-" * 70)
    lines.append(f"  1. 期待値の曖昧表現:     -{total_vague_penalty}点 ({len(all_vague_issues)}件検出, 上限{MAX_VAGUE_PENALTY})")
    lines.append(f"  2. 期待値の複数混在:     -{total_multi_penalty}点 ({len(all_multi_issues)}件検出, 上限{MAX_MULTI_KOTO_PENALTY})")
    lines.append(f"  3. 我流の単語:           -{total_word_penalty}点 ({len(all_word_issues)}件検出, 上限{MAX_CUSTOM_WORD_PENALTY})")
    lines.append(f"  4. 手順の省略:           -{total_steps_penalty}点 ({len(all_steps_issues)}件検出, 上限{MAX_MISSING_STEPS_PENALTY})")
    lines.append(f"  5. 粒度・品質:           -{total_granularity_penalty}点 (過去参考例との比較, 上限{MAX_GRANULARITY_PENALTY})")
    guardrail_issue_count = len(all_guardrail_money_issues) + len(all_guardrail_kanten_issues)
    lines.append(f"  6. ガードレール準拠:     -{total_guardrail_penalty}点 ({guardrail_issue_count}件検出, 上限{MAX_GUARDRAIL_PENALTY})")
    sf_issue_count = len(screen_format_result.get('issues', [])) if screen_format_result else 0
    lines.append(f"  7. 画面名・フォーマット: -{total_screen_format_penalty}点 ({sf_issue_count}件検出, 上限{MAX_SCREEN_FORMAT_PENALTY})")
    st_issue_count = len(structure_result.get('issues', [])) if structure_result else 0
    lines.append(f"  8. 構造化:               -{total_structure_penalty}点 ({st_issue_count}件検出, 上限{MAX_STRUCTURE_PENALTY})")
    pq_issue_count = len(perspective_quality_result.get('issues', [])) if perspective_quality_result else 0
    lines.append(f"  9. 観点シート品質:       -{total_perspective_quality_penalty}点 ({pq_issue_count}件検出, 上限{MAX_PERSPECTIVE_QUALITY_PENALTY})")
    lines.append(f"  ---")
    lines.append(f"  合計減点: -{total_penalty}点")
    lines.append("")

    # 詳細: 曖昧表現
    if all_vague_issues:
        lines.append("-" * 70)
        lines.append(f"【1】期待値の曖昧表現 (-{total_vague_penalty}点 / {len(all_vague_issues)}件)")
        lines.append("-" * 70)
        for issue in all_vague_issues:
            prefix = f"[{issue['_sheet']}] " if multi_sheet else ""
            lines.append(f"  {prefix}No.{issue['no']}: {issue['message']}")
            lines.append(f"    原文: {issue['value']}")
        lines.append("")

    # 詳細: 複数こと
    if all_multi_issues:
        lines.append("-" * 70)
        lines.append(f"【2】期待値の複数混在 (-{total_multi_penalty}点 / {len(all_multi_issues)}件)")
        lines.append("-" * 70)
        for issue in all_multi_issues:
            prefix = f"[{issue['_sheet']}] " if multi_sheet else ""
            lines.append(f"  {prefix}No.{issue['no']}: 「こと」が{issue['count']}回")
            lines.append(f"    原文: {issue['value']}")
        lines.append("")

    # 詳細: 我流の単語
    if all_word_issues:
        lines.append("-" * 70)
        lines.append(f"【3】我流の単語 (-{total_word_penalty}点 / {len(all_word_issues)}件)")
        lines.append("-" * 70)
        # NG単語ごとに集計
        word_summary = {}
        for issue in all_word_issues:
            key = issue['ng_word']
            if key not in word_summary:
                word_summary[key] = {'correct': issue['correct_word'], 'tc': 0, 'kanten': 0, 'examples': []}
            if issue['sheet'] == 'テストケース':
                word_summary[key]['tc'] += 1
            else:
                word_summary[key]['kanten'] += 1
            if len(word_summary[key]['examples']) < 3:
                word_summary[key]['examples'].append(f"No.{issue['no']} {issue['field']}")

        for ng_word, info in word_summary.items():
            lines.append(f"  「{ng_word}」→「{info['correct']}」: TC {info['tc']}件 + 観点 {info['kanten']}件")
            for ex in info['examples']:
                lines.append(f"    例: {ex}")
        lines.append("")

    # 詳細: 手順省略
    if all_steps_issues:
        lines.append("-" * 70)
        lines.append(f"【4】手順の省略 (-{total_steps_penalty}点 / {len(all_steps_issues)}件)")
        lines.append("-" * 70)
        for issue in all_steps_issues[:10]:
            prefix = f"[{issue['_sheet']}] " if multi_sheet else ""
            lines.append(f"  {prefix}No.{issue['no']}: {issue['message']}")
        if len(all_steps_issues) > 10:
            lines.append(f"  ... 他 {len(all_steps_issues) - 10}件")
        lines.append("")

    # 粒度・品質（全シート統合サマリー）
    lines.append("-" * 70)
    lines.append(f"【5】粒度・品質（過去参考例との比較） (-{total_granularity_penalty}点)")
    lines.append("-" * 70)

    # 全シートの統計を集約
    agg_total = 0
    agg_with_steps = 0
    agg_with_preconditions = 0
    agg_with_numbered = 0
    agg_with_structured = 0
    agg_expected_count = 0
    agg_good_ending = 0
    agg_short = 0
    agg_long = 0
    for sr in sheet_results:
        gs = sr['granularity_stats']
        agg_total += gs['total']
        agg_with_steps += gs['with_steps']
        agg_with_preconditions += gs['with_preconditions']
        agg_with_numbered += gs.get('with_numbered_steps', 0)
        agg_with_structured += gs.get('with_structured_steps', 0)
        agg_expected_count += gs.get('expected_count', 0)
        agg_good_ending += gs.get('good_ending_count', 0)
        agg_short += gs.get('short_expected', 0)
        agg_long += gs.get('long_expected', 0)

    safe_total = max(agg_total, 1)
    safe_exp = max(agg_expected_count, 1)

    lines.append(f"  総ケース数:        {agg_total}件")
    # 平均文字数は各シートのstatsから再計算
    total_exp_len_sum = sum(
        gs['avg_expected_length'] * gs.get('expected_count', 0)
        for sr in sheet_results for gs in [sr['granularity_stats']]
    )
    total_steps_len_sum = sum(
        gs['avg_steps_length'] * gs['with_steps']
        for sr in sheet_results for gs in [sr['granularity_stats']]
    )
    avg_exp = total_exp_len_sum / safe_exp if agg_expected_count else 0
    avg_steps = total_steps_len_sum / max(agg_with_steps, 1) if agg_with_steps else 0
    lines.append(f"  期待値 平均文字数:  {avg_exp:.0f}字（参考例基準: 10〜60字）")
    lines.append(f"  手順 平均文字数:    {avg_steps:.0f}字")
    lines.append(f"  手順あり:          {agg_with_steps}件 ({agg_with_steps/safe_total*100:.0f}%)")
    lines.append(f"  番号付きステップ:  {agg_with_numbered}件 ({agg_with_numbered/safe_total*100:.0f}%)")
    lines.append(f"  構造化手順:        {agg_with_structured}件（【前提条件】【セットアップ】【実行手順】等）")
    lines.append(f"  前提条件あり:      {agg_with_preconditions}件")
    lines.append(f"  良好な末尾表現:    {agg_good_ending}件 ({agg_good_ending/safe_exp*100:.0f}%)")
    lines.append(f"  期待値 短すぎ:     {agg_short}件")
    lines.append(f"  期待値 長すぎ:     {agg_long}件")

    if all_granularity_issues:
        lines.append("")
        lines.append("  指摘事項:")
        for issue in all_granularity_issues:
            prefix = f"[{issue['_sheet']}] " if multi_sheet else ""
            if issue['no'] == '-':
                lines.append(f"    {prefix}{issue['type']}: {issue['detail']}")
            else:
                lines.append(f"    {prefix}No.{issue['no']}: {issue['type']} - {issue['detail']}")

    lines.append("")
    per_file = benchmark.get('per_file', [])
    if per_file:
        lines.append("  ※ ベンチマーク元の参考例:")
        for pf in per_file:
            lines.append(f"    {pf['file']}({pf['cases']}件):"
                         f" 期待値平均{pf['expected_avg']}字"
                         f" / 手順{pf['steps_coverage']*100:.0f}%"
                         f" / こと1回{pf['koto_1_rate']*100:.0f}%")
    else:
        lines.append("  ※ デフォルト基準値を使用（learn_benchmark.py で参考例から生成可能）")
    lines.append("")

    # ガードレール準拠チェック
    lines.append("-" * 70)
    lines.append(f"【6】ガードレール準拠（qa-perspectives.md基準） (-{total_guardrail_penalty}点)")
    lines.append("-" * 70)

    # 6a. 金額・セキュリティテストP1チェック
    total_money_cases = sum(sr.get('guardrail_money', {}).get('money_cases', 0) for sr in sheet_results)
    total_security_cases = sum(sr.get('guardrail_money', {}).get('security_cases', 0) for sr in sheet_results)
    lines.append(f"  [6a] 金額・セキュリティ関連テストのP1設定 (-{total_guardrail_money_penalty}点)")
    p1_parts = []
    if total_money_cases > 0:
        p1_parts.append(f"金額{total_money_cases}件")
    if total_security_cases > 0:
        p1_parts.append(f"セキュリティ{total_security_cases}件")
    if p1_parts:
        lines.append(f"    P1推奨テスト検出: {' + '.join(p1_parts)}")
    if all_guardrail_money_issues:
        for issue in all_guardrail_money_issues:
            prefix = f"[{issue['_sheet']}] " if multi_sheet else ""
            lines.append(f"    {prefix}{issue['message']}")
    elif total_money_cases + total_security_cases > 0:
        lines.append(f"    全件P1設定済み")
    else:
        lines.append(f"    金額・セキュリティ関連テストなし（チェック対象外）")

    lines.append("")

    # 6b. 観点シートTC整合性チェック
    total_marked = sum(sr.get('guardrail_kanten', {}).get('marked_count', 0) for sr in sheet_results)
    total_uncovered = sum(sr.get('guardrail_kanten', {}).get('uncovered_count', 0) for sr in sheet_results)
    kanten_checked = any(sr.get('guardrail_kanten', {}).get('checked', False) for sr in sheet_results)
    lines.append(f"  [6b] 観点シートとTCの整合性 (-{total_guardrail_kanten_penalty}点)")
    if kanten_checked:
        lines.append(f"    観点シート○項目: {total_marked}件 / TC未対応: {total_uncovered}件")
        if all_guardrail_kanten_issues:
            for issue in all_guardrail_kanten_issues:
                prefix = f"[{issue['_sheet']}] " if multi_sheet else ""
                lines.append(f"    {prefix}{issue['message']}")
        elif total_marked > 0:
            lines.append(f"    全観点にTCが対応済み")
    else:
        lines.append(f"    観点シートなし（チェック対象外）")

    lines.append("")

    # チェック7: 画面名・フォーマット
    if screen_format_result:
        lines.append("-" * 70)
        lines.append(f"【7】画面名・フォーマット (-{total_screen_format_penalty}点)")
        lines.append("-" * 70)
        sf_issues = screen_format_result.get('issues', [])
        if sf_issues:
            lines.append(f"  [7a] 画面名 (-{screen_format_result.get('penalty_7a', 0)}点)")
            for issue in sf_issues:
                if issue['type'].startswith('screen_'):
                    lines.append(f"    {issue['message']}")
            lines.append(f"  [7b] 手順フォーマット (-{screen_format_result.get('penalty_7b', 0)}点)")
            for issue in sf_issues:
                if issue['type'].startswith('format_'):
                    lines.append(f"    {issue['message']}")
        else:
            lines.append(f"  指摘なし")
        lines.append("")

    # チェック8: 構造化
    if structure_result:
        lines.append("-" * 70)
        lines.append(f"【8】構造化 (-{total_structure_penalty}点)")
        lines.append("-" * 70)
        st_issues = structure_result.get('issues', [])
        if st_issues:
            for issue in st_issues:
                lines.append(f"    {issue['message']}")
        else:
            lines.append(f"  指摘なし")
        lines.append("")

    # チェック9: 観点シート品質
    if perspective_quality_result and perspective_quality_result.get('checked'):
        lines.append("-" * 70)
        lines.append(f"【9】観点シート品質 (-{total_perspective_quality_penalty}点)")
        lines.append("-" * 70)
        pq_issues = perspective_quality_result.get('issues', [])
        if pq_issues:
            for issue in pq_issues:
                lines.append(f"    {issue['message']}")
        else:
            lines.append(f"  指摘なし")
        lines.append("")

    # 修正推奨
    if total_penalty > 0:
        lines.append("=" * 70)
        lines.append("修正推奨事項（優先度順）")
        lines.append("=" * 70)
        priority = 1
        if all_guardrail_money_issues:
            p1_total = total_money_cases + total_security_cases
            lines.append(f"  {priority}. P1推奨テスト{p1_total}件の優先度をP1に設定（ガードレール基準）")
            priority += 1
        if all_vague_issues:
            lines.append(f"  {priority}. 曖昧な期待値 {len(all_vague_issues)}件を具体的な表現に書き換え")
            priority += 1
        if all_multi_issues:
            lines.append(f"  {priority}. 複数期待値 {len(all_multi_issues)}件を1ケース1期待値に分割または統合")
            priority += 1
        if all_word_issues:
            ng_words = list(set(i['ng_word'] for i in all_word_issues))
            lines.append(f"  {priority}. 我流単語を正式名称に置換: {', '.join(ng_words)}")
            priority += 1
        if all_steps_issues:
            lines.append(f"  {priority}. 手順が空欄の {len(all_steps_issues)}件に手順を追加")
            priority += 1
        if total_granularity_penalty > 0:
            details = [i['type'] for i in all_granularity_issues if i['no'] == '-']
            lines.append(f"  {priority}. 粒度・品質の改善: {', '.join(details)}")
            priority += 1
        if all_guardrail_kanten_issues:
            lines.append(f"  {priority}. 観点シート○の{total_uncovered}件にTCを追加（観点カバレッジ不足）")
            priority += 1
        if screen_format_result and screen_format_result.get('issues'):
            lines.append(f"  {priority}. 画面名・フォーマットの修正（{len(screen_format_result['issues'])}件）")
            priority += 1
        if structure_result and structure_result.get('issues'):
            lines.append(f"  {priority}. 構造化の改善（{len(structure_result['issues'])}件）")
            priority += 1
        if perspective_quality_result and perspective_quality_result.get('issues'):
            lines.append(f"  {priority}. 観点シート品質の改善（{len(perspective_quality_result['issues'])}件）")
        lines.append("")
    else:
        lines.append("=" * 70)
        lines.append("全チェック項目をクリアしています！")
        lines.append("=" * 70)

    return '\n'.join(lines), score


# =====================================================================
# メイン
# =====================================================================

def main():
    parser = argparse.ArgumentParser(description='テストケース自動評価')
    parser.add_argument('excel_file', help='評価対象のテスト項目書（.xlsx）')
    parser.add_argument('--keywords', '-k', help='NGワードリスト（TSV形式）', default=None)
    parser.add_argument('--sheet', '-s', help='テストケースシート名（指定時は1シートのみ評価）', default=None)
    parser.add_argument('--kanten', help='観点シート名', default=None)
    parser.add_argument('--header-row', type=int, help='ヘッダー行番号（手動上書き用）', default=None)
    parser.add_argument('--data-start-row', type=int, help='データ開始行番号（手動上書き用）', default=None)
    parser.add_argument('--output', '-o', help='レポート出力先（デフォルト: 標準出力）', default=None)
    parser.add_argument('--benchmark', '-b', help='ベンチマークJSONファイル（learn_benchmark.pyで生成）', default=None)
    args = parser.parse_args()

    if not os.path.exists(args.excel_file):
        print(f"ERROR: ファイルが見つかりません: {args.excel_file}")
        sys.exit(1)

    # ベンチマーク読み込み
    benchmark, benchmark_loaded = load_benchmark(args.benchmark)
    if benchmark_loaded:
        meta = benchmark.get('_meta', {})
        src_count = len(meta.get('source_files', []))
        total_cases = meta.get('total_cases', '?')
        print(f"ベンチマーク: {src_count}ファイル / {total_cases}ケースから算出")
    else:
        print("ベンチマーク: デフォルト値を使用（--benchmark で指定可能）")

    print(f"読み込み中: {args.excel_file}")
    wb = openpyxl.load_workbook(args.excel_file, data_only=True)
    print(f"シート一覧: {wb.sheetnames}")

    # キーワード読み込み
    keywords = load_keywords(args.keywords)
    if keywords:
        print(f"NGワード: {len(keywords)}件")
    else:
        print("NGワードファイル未指定。我流単語チェックはスキップします。")

    # 観点シート検出
    ws_kanten = None
    if args.kanten:
        ws_kanten = wb[args.kanten]
    else:
        kanten_candidates = ['観点', '観点シート', 'テスト観点']
        ws_kanten = find_sheet(wb, kanten_candidates)
    if ws_kanten:
        print(f"観点シート: {ws_kanten.title}")
    else:
        print("WARNING: 観点シートが見つかりません。観点シートのチェックはスキップします。")

    # =====================================================================
    # シート検出と評価実行
    # =====================================================================
    sheet_results = []

    if args.sheet:
        # --sheet 指定時: 1シートのみ評価
        ws_tc = wb[args.sheet]
        header_row = args.header_row if args.header_row else (detect_header_row(ws_tc) or 2)
        data_start_row = args.data_start_row if args.data_start_row else (header_row + 1)
        col_map = detect_columns(ws_tc, header_row)
        print(f"テストケースシート: {ws_tc.title} (ヘッダー行: {header_row})")
        print(f"検出カラム: {col_map}")

        cases = read_cases(ws_tc, col_map, data_start_row)
        print(f"ケース数: {len(cases)}件")

        if cases:
            vague_issues, vague_penalty = check_vague_expressions(cases, col_map)
            multi_issues, multi_penalty = check_multi_koto(cases, col_map)
            word_issues, word_penalty = check_custom_words(cases, col_map, keywords, ws_kanten)
            steps_issues, steps_penalty = check_missing_steps(cases, col_map)
            granularity_stats = check_granularity(cases, col_map, benchmark)
            guardrail_money = check_money_priority(cases, col_map)
            guardrail_kanten = check_kanten_tc_coverage(cases, col_map, ws_kanten)

            sheet_results.append({
                'sheet_name': ws_tc.title,
                'cases': cases,
                'col_map': col_map,
                'vague_issues': vague_issues,
                'vague_penalty': vague_penalty,
                'multi_issues': multi_issues,
                'multi_penalty': multi_penalty,
                'word_issues': word_issues,
                'word_penalty': word_penalty,
                'steps_issues': steps_issues,
                'steps_penalty': steps_penalty,
                'granularity_stats': granularity_stats,
                'guardrail_money': guardrail_money,
                'guardrail_kanten': guardrail_kanten,
            })
    else:
        # --sheet 未指定: 全シートを走査
        print()
        print("全シート走査モード:")
        for ws_name in wb.sheetnames:
            # スキップ対象シートをフィルタ
            if should_skip_sheet(ws_name):
                print(f"  SKIP: {ws_name} (除外パターンに一致)")
                continue

            ws = wb[ws_name]

            # ヘッダー行を自動検出（手動上書きがあればそちらを使用）
            if args.header_row:
                header_row = args.header_row
            else:
                header_row = detect_header_row(ws)
            if header_row is None:
                continue

            # カラム検出
            col_map = detect_columns(ws, header_row)
            if 'expected' not in col_map and 'steps' not in col_map:
                continue

            # データ開始行
            data_start_row = args.data_start_row if args.data_start_row else (header_row + 1)

            # テストケース読み込み
            cases = read_cases(ws, col_map, data_start_row)
            if not cases:
                continue

            print(f"  評価中: {ws_name} ({len(cases)}件, ヘッダー行: {header_row})")

            # 各チェック実行
            vague_issues, vague_penalty = check_vague_expressions(cases, col_map)
            multi_issues, multi_penalty = check_multi_koto(cases, col_map)
            word_issues, word_penalty = check_custom_words(cases, col_map, keywords, ws_kanten)
            steps_issues, steps_penalty = check_missing_steps(cases, col_map)
            granularity_stats = check_granularity(cases, col_map, benchmark)
            guardrail_money = check_money_priority(cases, col_map)
            guardrail_kanten = check_kanten_tc_coverage(cases, col_map, ws_kanten)

            sheet_results.append({
                'sheet_name': ws_name,
                'cases': cases,
                'col_map': col_map,
                'vague_issues': vague_issues,
                'vague_penalty': vague_penalty,
                'multi_issues': multi_issues,
                'multi_penalty': multi_penalty,
                'word_issues': word_issues,
                'word_penalty': word_penalty,
                'steps_issues': steps_issues,
                'steps_penalty': steps_penalty,
                'granularity_stats': granularity_stats,
                'guardrail_money': guardrail_money,
                'guardrail_kanten': guardrail_kanten,
            })

    if not sheet_results:
        print("ERROR: テストケースを含むシートが見つかりません。--sheet オプションで指定してください。")
        sys.exit(1)

    print()
    total_cases = sum(len(sr['cases']) for sr in sheet_results)
    print(f"合計: {len(sheet_results)}シート / {total_cases}件")

    # 新チェック7-9（全シート合算のケースで実行）
    all_cases = []
    all_col_map = {}
    for sr in sheet_results:
        all_cases.extend(sr['cases'])
        if not all_col_map:
            all_col_map = sr['col_map']

    # チェック7: 画面名・フォーマット
    screen_format_result = check_screen_and_format(all_cases, all_col_map)
    # チェック8: 構造化
    structure_result = check_structure(all_cases, all_col_map, wb)
    # チェック9: 観点シート品質
    perspective_quality_result = check_perspective_quality(ws_kanten)

    print()

    # レポート生成
    report, score = generate_report(
        args.excel_file, sheet_results, benchmark,
        screen_format_result=screen_format_result,
        structure_result=structure_result,
        perspective_quality_result=perspective_quality_result,
    )

    print(report)

    # ファイル出力
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"\nレポートを保存しました: {args.output}")

    return score


if __name__ == '__main__':
    main()
