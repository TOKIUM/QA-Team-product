#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""参考例フォルダからベンチマーク（基準値）を自動生成する

使い方:
    python learn_benchmark.py <参考例フォルダ> [--output benchmark.json]

参考例フォルダに .xlsx ファイルを置くだけで、全ファイルを分析して
基準値 (benchmark.json) を自動生成する。

生成された benchmark.json は evaluate_testcases.py が自動読み込みする。
"""
import os, sys, re, json, argparse, shutil
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl")
    sys.exit(1)

from common import (
    HEADER_KEYWORDS, SKIP_SHEET_PATTERNS, VAGUE_EXCEPTIONS,
    detect_header_row, detect_columns, merge_stats,
    should_skip_sheet, strip_parens, is_vague_exception,
)


GOOD_ENDINGS = [
    'されていること', 'されること', 'していること',
    'できていること', 'できること', 'であること',
    'すること', 'になること', 'になっていること',
    'ないこと', 'されないこと', 'れること',
    'いること', 'くこと',
]

VAGUE_PATTERNS = [
    r'正常であること', r'正常に動作すること',
    r'問題ないこと', r'問題なく',
    r'仕様通りであること', r'仕様通りに動作すること',
    r'正しく動作すること', r'適切に表示されること',
    r'適切であること', r'正しく表示されること',
    r'期待通り', r'想定通り',
]


def analyze_sheet(ws, col_map, data_start_row):
    """1シートを分析して統計情報を返す"""
    no_col = col_map.get('no')
    stats = {
        'case_count': 0,
        'expected_lengths': [],
        'steps_lengths': [],
        'with_steps': 0,
        'with_numbered_steps': 0,
        'with_preconditions': 0,
        'koto_1': 0,
        'koto_2plus': 0,
        'vague_count': 0,
        'good_ending_count': 0,
        'expected_count': 0,
        'endings': {},
    }

    for row in range(data_start_row, ws.max_row + 1):
        # No列があればNoneの行はスキップ
        if no_col:
            no_val = ws.cell(row=row, column=no_col).value
            if no_val is None:
                continue

        stats['case_count'] += 1

        # 手順
        steps_col = col_map.get('steps')
        steps = ''
        if steps_col:
            steps = str(ws.cell(row=row, column=steps_col).value or '')
        if steps and steps.strip() not in ('', '-', 'None'):
            stats['with_steps'] += 1
            stats['steps_lengths'].append(len(steps))
            if re.search(r'(?:^|\n)\s*[1１①⑴\(1\)][\.\．\s、）\)]', steps):
                stats['with_numbered_steps'] += 1
            if '前提条件' in steps or '【前提】' in steps or '【前提条件】' in steps:
                stats['with_preconditions'] += 1

        # 期待値
        exp_col = col_map.get('expected')
        expected = ''
        if exp_col:
            expected = str(ws.cell(row=row, column=exp_col).value or '')
        if expected and expected.strip() not in ('', '-', 'None', '同上'):
            stats['expected_count'] += 1
            stats['expected_lengths'].append(len(expected))

            # 「こと」カウント（括弧内の補足は除外）
            exp_without_parens = strip_parens(expected)
            koto_count = len(re.findall(r'こと', exp_without_parens))
            if koto_count == 1:
                stats['koto_1'] += 1
            elif koto_count >= 2:
                stats['koto_2plus'] += 1

            # 曖昧表現（例外パターンは偽陽性として除外）
            for pattern in VAGUE_PATTERNS:
                if re.search(pattern, expected):
                    if not is_vague_exception(expected):
                        stats['vague_count'] += 1
                    break

            # 末尾表現
            for ending in GOOD_ENDINGS:
                if expected.rstrip().endswith(ending):
                    stats['good_ending_count'] += 1
                    stats['endings'][ending] = stats['endings'].get(ending, 0) + 1
                    break

    return stats


# =====================================================================
# ベンチマーク生成
# =====================================================================

def compute_benchmark(file_stats_list):
    """全ファイルの分析結果からベンチマーク基準値を算出"""

    # 全ファイル統合
    all_expected_lengths = []
    all_steps_lengths = []
    total_cases = 0
    total_with_steps = 0
    total_with_numbered = 0
    total_expected = 0
    total_koto_1 = 0
    total_koto_2plus = 0
    total_vague = 0
    total_good_ending = 0
    all_endings = {}

    for fs in file_stats_list:
        stats = fs['stats']
        all_expected_lengths.extend(stats['expected_lengths'])
        all_steps_lengths.extend(stats['steps_lengths'])
        total_cases += stats['case_count']
        total_with_steps += stats['with_steps']
        total_with_numbered += stats['with_numbered_steps']
        total_expected += stats['expected_count']
        total_koto_1 += stats['koto_1']
        total_koto_2plus += stats['koto_2plus']
        total_vague += stats['vague_count']
        total_good_ending += stats['good_ending_count']
        for k, v in stats['endings'].items():
            all_endings[k] = all_endings.get(k, 0) + v

    # 統計値算出
    avg_expected = sum(all_expected_lengths) / max(len(all_expected_lengths), 1)
    avg_steps = sum(all_steps_lengths) / max(len(all_steps_lengths), 1)

    # パーセンタイル（期待値の長さ）
    sorted_exp = sorted(all_expected_lengths)
    p5 = sorted_exp[int(len(sorted_exp) * 0.05)] if sorted_exp else 0
    p95 = sorted_exp[int(len(sorted_exp) * 0.95)] if sorted_exp else 100

    # 末尾表現TOP
    top_endings = sorted(all_endings.items(), key=lambda x: -x[1])[:15]

    safe_expected = max(total_expected, 1)

    benchmark = {
        '_meta': {
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'source_files': [fs['file'] for fs in file_stats_list],
            'total_cases': total_cases,
            'total_expected': total_expected,
        },
        'expected_length': {
            'avg': round(avg_expected, 1),
            'p5': p5,
            'p95': p95,
            'ideal_min': max(5, p5),
            'ideal_max': max(80, p95),
        },
        'steps_length': {
            'avg': round(avg_steps, 1),
        },
        'steps_coverage': {
            'rate': round(total_with_steps / max(total_cases, 1), 3),
            'good': 0.80,
            'warning': 0.60,
        },
        'numbered_steps': {
            'rate': round(total_with_numbered / max(total_cases, 1), 3),
            'good': 0.20,
        },
        'single_koto_rate': {
            'rate': round(total_koto_1 / safe_expected, 4),
            'good': 0.98,
            'warning': 0.95,
        },
        'vague_rate': {
            'rate': round(total_vague / safe_expected, 4),
            'good': 0.01,
            'acceptable': 0.03,
        },
        'good_endings': {
            'rate': round(total_good_ending / safe_expected, 3),
            'good': 0.80,
            'patterns': [e[0] for e in top_endings],
        },
        'per_file': [],
    }

    # ファイル別サマリー
    for fs in file_stats_list:
        s = fs['stats']
        se = max(s['expected_count'], 1)
        sc = max(s['case_count'], 1)
        benchmark['per_file'].append({
            'file': fs['file'],
            'cases': s['case_count'],
            'expected_avg': round(sum(s['expected_lengths']) / max(len(s['expected_lengths']), 1), 1),
            'steps_avg': round(sum(s['steps_lengths']) / max(len(s['steps_lengths']), 1), 1),
            'steps_coverage': round(s['with_steps'] / sc, 3),
            'koto_1_rate': round(s['koto_1'] / se, 3),
            'vague_rate': round(s['vague_count'] / se, 4),
            'good_ending_rate': round(s['good_ending_count'] / se, 3),
        })

    return benchmark


# =====================================================================
# 学習レポート生成
# =====================================================================

def load_previous_benchmark(report_base_dir):
    """直前の学習レポートからbenchmark.jsonを読み込む（差分比較用）"""
    if not os.path.isdir(report_base_dir):
        return None
    dates = sorted([d for d in os.listdir(report_base_dir)
                    if os.path.isdir(os.path.join(report_base_dir, d))
                    and re.match(r'\d{4}-\d{2}-\d{2}', d)])
    if not dates:
        return None
    prev_path = os.path.join(report_base_dir, dates[-1], 'benchmark.json')
    if os.path.isfile(prev_path):
        with open(prev_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def fmt_diff(new_val, old_val, fmt='.1f', suffix='', invert=False):
    """差分を +/- 表記でフォーマット。invertがTrueなら減少が改善方向"""
    if old_val is None:
        return '（前回なし）'
    diff = new_val - old_val
    if diff == 0:
        return '±0'
    sign = '+' if diff > 0 else ''
    direction = ''
    if invert:
        direction = ' (改善)' if diff < 0 else ' (悪化)' if diff > 0 else ''
    else:
        direction = ' (改善)' if diff > 0 else ' (悪化)' if diff < 0 else ''
    return f'{sign}{diff:{fmt}}{suffix}{direction}'


def generate_report(benchmark, prev_benchmark, report_dir):
    """学習レポートMarkdownを生成して保存"""
    today = datetime.now().strftime('%Y-%m-%d')
    bm = benchmark
    prev = prev_benchmark

    # ファイル別テーブル
    file_rows = []
    for i, pf in enumerate(bm['per_file'], 1):
        file_rows.append(
            f"| {i} | {pf['file'][:60]} | {pf['cases']:,}件 "
            f"| {pf['expected_avg']}字 | {pf['koto_1_rate']*100:.0f}% "
            f"| {pf['vague_rate']*100:.1f}% | {pf['good_ending_rate']*100:.0f}% |"
        )

    # 差分テーブル
    def get_prev(key1, key2):
        if prev and key1 in prev and key2 in prev[key1]:
            return prev[key1][key2]
        return None

    diff_rows = [
        ('総ケース数', f"{bm['_meta']['total_cases']:,}",
         fmt_diff(bm['_meta']['total_cases'],
                  prev['_meta']['total_cases'] if prev else None, ',d', '件')),
        ('期待値平均文字数', f"{bm['expected_length']['avg']}字",
         fmt_diff(bm['expected_length']['avg'],
                  get_prev('expected_length', 'avg'), '.1f', '字')),
        ('期待値 適正範囲', f"{bm['expected_length']['ideal_min']}〜{bm['expected_length']['ideal_max']}字",
         '—'),
        ('手順平均文字数', f"{bm['steps_length']['avg']}字",
         fmt_diff(bm['steps_length']['avg'],
                  get_prev('steps_length', 'avg'), '.1f', '字')),
        ('手順記載率', f"{bm['steps_coverage']['rate']*100:.1f}%",
         fmt_diff(bm['steps_coverage']['rate']*100,
                  get_prev('steps_coverage', 'rate') * 100 if get_prev('steps_coverage', 'rate') else None, '.1f', 'pt')),
        ('番号付き手順率', f"{bm['numbered_steps']['rate']*100:.1f}%",
         fmt_diff(bm['numbered_steps']['rate']*100,
                  get_prev('numbered_steps', 'rate') * 100 if get_prev('numbered_steps', 'rate') else None, '.1f', 'pt')),
        ('こと1回率', f"{bm['single_koto_rate']['rate']*100:.2f}%",
         fmt_diff(bm['single_koto_rate']['rate']*100,
                  get_prev('single_koto_rate', 'rate') * 100 if get_prev('single_koto_rate', 'rate') else None, '.2f', 'pt')),
        ('曖昧表現率', f"{bm['vague_rate']['rate']*100:.2f}%",
         fmt_diff(bm['vague_rate']['rate']*100,
                  get_prev('vague_rate', 'rate') * 100 if get_prev('vague_rate', 'rate') else None, '.2f', 'pt', invert=True)),
        ('良好末尾率', f"{bm['good_endings']['rate']*100:.1f}%",
         fmt_diff(bm['good_endings']['rate']*100,
                  get_prev('good_endings', 'rate') * 100 if get_prev('good_endings', 'rate') else None, '.1f', 'pt')),
    ]

    diff_table = '\n'.join(
        f"| {name} | {val} | {diff} |" for name, val, diff in diff_rows
    )
    file_table = '\n'.join(file_rows)

    top5 = ', '.join(bm['good_endings']['patterns'][:5])

    report = f"""# 学習レポート {today}

## 概要

- 参考例ファイル数: **{len(bm['per_file'])}件**
- 総テストケース数: **{bm['_meta']['total_cases']:,}件**（期待値あり {bm['_meta']['total_expected']:,}件）
- 生成日時: {bm['_meta']['generated_at']}

---

## ファイル別サマリー

| # | ファイル | ケース数 | 期待値平均 | こと1回率 | 曖昧率 | 良好末尾率 |
|---|---|---|---|---|---|---|
{file_table}

---

## 基準値サマリー（前回との差分）

| 指標 | 今回の値 | 前回比 |
|---|---|---|
{diff_table}

---

## 末尾表現パターン TOP5

1. {top5.split(', ')[0] if len(bm['good_endings']['patterns']) > 0 else '—'}
2. {top5.split(', ')[1] if len(bm['good_endings']['patterns']) > 1 else '—'}
3. {top5.split(', ')[2] if len(bm['good_endings']['patterns']) > 2 else '—'}
4. {top5.split(', ')[3] if len(bm['good_endings']['patterns']) > 3 else '—'}
5. {top5.split(', ')[4] if len(bm['good_endings']['patterns']) > 4 else '—'}

---

## 判定閾値

| 指標 | 良好 | 警告 |
|---|---|---|
| 手順記載率 | {bm['steps_coverage']['good']*100:.0f}%以上 | {bm['steps_coverage']['warning']*100:.0f}%以下 |
| こと1回率 | {bm['single_koto_rate']['good']*100:.0f}%以上 | {bm['single_koto_rate']['warning']*100:.0f}%以下 |
| 曖昧表現率 | {bm['vague_rate']['good']*100:.0f}%以下 | {bm['vague_rate']['acceptable']*100:.0f}%超 |
| 良好末尾率 | {bm['good_endings']['good']*100:.0f}%以上 | — |
| 番号付き手順率 | {bm['numbered_steps']['good']*100:.0f}%以上 | — |
"""

    report_path = os.path.join(report_dir, '学習レポート.md')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)

    return report_path


# =====================================================================
# メイン
# =====================================================================

def main():
    parser = argparse.ArgumentParser(description='参考例からベンチマーク基準値を自動生成')
    parser.add_argument('reference_dir', help='参考例フォルダのパス（.xlsxファイルを含む）')
    parser.add_argument('--output', '-o', default=None,
                        help='出力先（デフォルト: 参考例フォルダ内の benchmark.json）')
    parser.add_argument('--verbose', '-v', action='store_true', help='詳細出力')
    parser.add_argument('--include-all', action='store_true',
                        help='シート名フィルタを無効化し、全シートを学習対象にする')
    args = parser.parse_args()

    ref_dir = args.reference_dir
    if not os.path.isdir(ref_dir):
        print(f"ERROR: フォルダが見つかりません: {ref_dir}")
        sys.exit(1)

    # .xlsxファイルを列挙
    xlsx_files = [f for f in os.listdir(ref_dir)
                  if f.endswith('.xlsx') and not f.startswith('~$')]
    if not xlsx_files:
        print(f"ERROR: {ref_dir} に .xlsx ファイルがありません")
        sys.exit(1)

    print(f"参考例フォルダ: {ref_dir}")
    print(f"対象ファイル: {len(xlsx_files)}件")
    print()

    file_stats_list = []

    for fname in xlsx_files:
        fpath = os.path.join(ref_dir, fname)
        print(f"分析中: {fname}")

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            print(f"  SKIP: 読み込みエラー ({e})")
            continue

        # テストケースを含むシートを探す
        sheet_stats = []
        skipped_sheets = []
        for ws_name in wb.sheetnames:
            # シート名フィルタ（改善前・メモ・振り返り等を除外）
            if not args.include_all and should_skip_sheet(ws_name):
                skipped_sheets.append(ws_name)
                continue

            ws = wb[ws_name]
            header_row = detect_header_row(ws)
            if header_row is None:
                continue

            col_map = detect_columns(ws, header_row)
            if 'expected' not in col_map and 'steps' not in col_map:
                continue

            data_start_row = header_row + 1
            stats = analyze_sheet(ws, col_map, data_start_row)

            if stats['case_count'] == 0:
                continue

            sheet_stats.append(stats)
            if args.verbose:
                print(f"  シート「{ws_name}」: {stats['case_count']}件"
                      f" (期待値あり{stats['expected_count']}件)")

        if skipped_sheets and args.verbose:
            print(f"  除外シート: {', '.join(skipped_sheets)}")

        if not sheet_stats:
            print(f"  SKIP: テストケースシートが見つかりません")
            continue

        merged = merge_stats(sheet_stats)
        print(f"  → {merged['case_count']}件（{len(sheet_stats)}シート）")
        file_stats_list.append({'file': fname, 'stats': merged})

    if not file_stats_list:
        print("\nERROR: 分析可能なファイルがありませんでした")
        sys.exit(1)

    # ベンチマーク生成
    print()
    benchmark = compute_benchmark(file_stats_list)

    # 出力
    output_path = args.output or os.path.join(ref_dir, 'benchmark.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(benchmark, f, ensure_ascii=False, indent=2)

    print("=" * 60)
    print("ベンチマーク生成完了")
    print("=" * 60)
    print(f"  ファイル数:    {len(file_stats_list)}件")
    print(f"  総ケース数:    {benchmark['_meta']['total_cases']}件")
    print(f"  総期待値数:    {benchmark['_meta']['total_expected']}件")
    print(f"  期待値平均:    {benchmark['expected_length']['avg']}字"
          f"（5%tile:{benchmark['expected_length']['p5']}字"
          f" / 95%tile:{benchmark['expected_length']['p95']}字）")
    print(f"  手順平均:      {benchmark['steps_length']['avg']}字")
    print(f"  手順記載率:    {benchmark['steps_coverage']['rate']*100:.1f}%")
    print(f"  番号付き率:    {benchmark['numbered_steps']['rate']*100:.1f}%")
    print(f"  こと1回率:     {benchmark['single_koto_rate']['rate']*100:.1f}%")
    print(f"  曖昧表現率:    {benchmark['vague_rate']['rate']*100:.2f}%")
    print(f"  良好末尾率:    {benchmark['good_endings']['rate']*100:.1f}%")
    print(f"  末尾TOP5:      {', '.join(benchmark['good_endings']['patterns'][:5])}")
    print()

    # ファイル別サマリー
    print("--- ファイル別サマリー ---")
    for pf in benchmark['per_file']:
        print(f"  {pf['file']}: {pf['cases']}件"
              f" / 期待値{pf['expected_avg']}字"
              f" / 手順{pf['steps_coverage']*100:.0f}%"
              f" / こと1回{pf['koto_1_rate']*100:.0f}%"
              f" / 曖昧{pf['vague_rate']*100:.1f}%")

    print(f"\n保存先: {output_path}")
    print("\nevaluate_testcases.py に --benchmark オプションで読み込ませてください:")
    print(f"  python evaluate_testcases.py <対象.xlsx> --benchmark \"{output_path}\"")

    # 学習レポート生成（日付フォルダ）
    script_dir = os.path.dirname(os.path.abspath(__file__))
    report_base = os.path.join(script_dir, '学習レポート')
    today = datetime.now().strftime('%Y-%m-%d')
    report_dir = os.path.join(report_base, today)
    os.makedirs(report_dir, exist_ok=True)

    # 前回のbenchmarkを読み込み（差分比較用）
    prev_benchmark = load_previous_benchmark(report_base)
    # 今日のフォルダが前回と同じ場合はその前を使う
    dates = sorted([d for d in os.listdir(report_base)
                    if os.path.isdir(os.path.join(report_base, d))
                    and re.match(r'\d{4}-\d{2}-\d{2}', d) and d != today])
    if dates:
        prev_path = os.path.join(report_base, dates[-1], 'benchmark.json')
        if os.path.isfile(prev_path):
            with open(prev_path, 'r', encoding='utf-8') as f:
                prev_benchmark = json.load(f)
        else:
            prev_benchmark = None
    else:
        prev_benchmark = None

    # benchmark.jsonのスナップショットを保存
    snapshot_path = os.path.join(report_dir, 'benchmark.json')
    shutil.copy2(output_path, snapshot_path)

    # レポート生成
    report_path = generate_report(benchmark, prev_benchmark, report_dir)

    print(f"\n学習レポート: {report_path}")
    print(f"スナップショット: {snapshot_path}")


if __name__ == '__main__':
    main()
