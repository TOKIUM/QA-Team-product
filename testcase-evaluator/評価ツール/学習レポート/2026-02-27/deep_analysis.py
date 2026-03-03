#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""高品質テスト項目書4件の深掘り分析スクリプト

5つの分析を実施し、Markdownレポートとして出力する。
"""
import os, sys, re, collections
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl")
    sys.exit(1)

# =====================================================================
# 設定
# =====================================================================

FILES = [
    {
        'short_name': 'クリッピング',
        'path': r'C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\【テスト項目書】クリッピング練習問題の複数ページ対応（画像の複数枚アップロード対応）.xlsx',
    },
    {
        'short_name': 'ログイン制限',
        'path': r'C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\テスト項目書_ログイン方法でID・PWとTOKIUM IDが存在し、TOKIUM IDだけ利用したい場合にフラグでログイン制限できるように (2).xlsx',
    },
    {
        'short_name': '取引先マスタ',
        'path': r'C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\テスト項目書_取引先マスタのテーブル構造変更（企業マスタ追加）.xlsx',
    },
    {
        'short_name': '改善後v7',
        'path': r'C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\テスト項目書_改善後_v7.xlsx',
    },
]

EXCLUDE_PATTERNS = ['改善前', '質問', '振り返り', 'レビュー指摘', 'memo']

# 曖昧表現パターン
VAGUE_PATTERNS = [
    ('正常であること', '正常であること'),
    ('正常に動作すること', '正常に動作すること'),
    ('問題ないこと', '問題ないこと'),
    ('問題なく', '問題なく〜こと'),
    ('仕様通りであること', '仕様通りであること'),
    ('仕様通りに動作', '仕様通りに動作'),
    ('正しく動作', '正しく動作'),
    ('適切に表示', '適切に表示'),
    ('適切であること', '適切であること'),
    ('正しく表示', '正しく表示'),
    ('期待通り', '期待通り'),
    ('想定通り', '想定通り'),
]

# =====================================================================
# ユーティリティ
# =====================================================================

def should_exclude_sheet(sheet_name):
    """除外すべきシートかどうか"""
    for pat in EXCLUDE_PATTERNS:
        if pat in sheet_name:
            return True
    # テストケースに関連しないシート
    non_tc_sheets = [
        '運用ルール', '観点', '進捗', 'チェック表', 'バグ一覧',
        'old_チェック表', '設定', 'スクショ', 'レビュー＆質問',
        '起票まとめ', '新機能について', 'テストシナリオ',
        '石川_ケースに不記載で確認したことめも',
    ]
    for ns in non_tc_sheets:
        if sheet_name == ns:
            return True
    if sheet_name.startswith('【修正中】'):
        return True
    return False


def detect_columns(ws, header_row=2):
    """ヘッダー行からカラムマッピングを自動検出"""
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if not val:
            continue
        val = str(val).strip()
        if val in ('No', 'NO', 'no'):
            col_map.setdefault('no', col)
        elif '確認すること' in val or '確認項目' in val:
            col_map.setdefault('confirm', col)
        elif val == '画面':
            col_map.setdefault('screen', col)
        elif '確認対象' in val:
            col_map.setdefault('target', col)
        elif '詳細' in val:
            col_map.setdefault('detail', col)
        elif '手順' in val:
            col_map.setdefault('steps', col)
        elif '期待値' in val or '期待結果' in val:
            col_map.setdefault('expected', col)
        elif val == '備考':
            col_map.setdefault('notes', col)
    return col_map


def read_all_cases(file_info):
    """1ファイルから全テストケースシートのケースを読み込む"""
    wb = openpyxl.load_workbook(file_info['path'], data_only=True)
    all_cases = []

    for sheet_name in wb.sheetnames:
        if should_exclude_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        col_map = detect_columns(ws, header_row=2)

        # 期待値カラムが無ければスキップ
        if 'expected' not in col_map:
            # header_row=1 も試す
            col_map = detect_columns(ws, header_row=1)
            if 'expected' not in col_map:
                continue

        no_col = col_map.get('no', 2)
        data_start = 3  # デフォルト

        for row in range(data_start, ws.max_row + 1):
            no_val = ws.cell(row=row, column=no_col).value
            if no_val is None:
                continue

            case = {
                'file': file_info['short_name'],
                'sheet': sheet_name,
                'row': row,
                'no': no_val,
            }
            for key, col in col_map.items():
                if key == 'no':
                    continue
                raw = ws.cell(row=row, column=col).value
                case[key] = str(raw).strip() if raw is not None else ''
            all_cases.append(case)

    return all_cases


# =====================================================================
# 分析1: 取引先マスタの曖昧表現の具体例
# =====================================================================

def analysis_1(all_data):
    """取引先マスタから曖昧表現を全件リストアップ"""
    lines = []
    lines.append("## 分析1: 取引先マスタの曖昧表現の具体例\n")

    torihikisaki_cases = [c for c in all_data if c['file'] == '取引先マスタ']
    total = len(torihikisaki_cases)
    expected_cases = [c for c in torihikisaki_cases if c.get('expected', '').strip() and c.get('expected', '') != '-']
    total_expected = len(expected_cases)

    found = []
    for case in expected_cases:
        exp = case.get('expected', '')
        matched_patterns = []
        for pattern_str, label in VAGUE_PATTERNS:
            if pattern_str in exp:
                matched_patterns.append(label)
        if matched_patterns:
            found.append({
                'sheet': case['sheet'],
                'row': case['row'],
                'expected': exp,
                'patterns': matched_patterns,
            })

    lines.append(f"- 取引先マスタの総ケース数: {total}件")
    lines.append(f"- 期待値ありのケース数: {total_expected}件")
    lines.append(f"- 曖昧表現を含むケース数: {len(found)}件")
    if total_expected > 0:
        lines.append(f"- 曖昧表現率: {len(found)/total_expected*100:.1f}%")
    lines.append("")

    if found:
        lines.append("| # | シート名 | 行番号 | 期待値（全文） | 該当パターン |")
        lines.append("|---|----------|--------|----------------|--------------|")
        for i, item in enumerate(found, 1):
            exp_escaped = item['expected'].replace('\n', ' ').replace('|', '\\|')
            patterns = ', '.join(item['patterns'])
            lines.append(f"| {i} | {item['sheet']} | {item['row']} | {exp_escaped} | {patterns} |")
    else:
        lines.append("曖昧表現は検出されませんでした。")

    lines.append("")
    return '\n'.join(lines)


# =====================================================================
# 分析2: 期待値の良い書き方パターンTOP20
# =====================================================================

def analysis_2(all_data):
    """末尾表現ごとにグループ化し、例文を抽出"""
    lines = []
    lines.append("## 分析2: 期待値の良い書き方パターンTOP20\n")

    # 期待値を収集
    expected_list = []
    for case in all_data:
        exp = case.get('expected', '').strip()
        if exp and exp != '-' and exp != 'None':
            expected_list.append({
                'text': exp.rstrip(),
                'file': case['file'],
                'sheet': case['sheet'],
                'row': case['row'],
            })

    total_expected = len(expected_list)
    lines.append(f"全4ファイルの期待値 合計: {total_expected}件\n")

    # 末尾パターンを検出
    ending_patterns = [
        'されていること', 'されること', 'れること',
        'していること', 'できること', 'であること',
        'すること', 'できていること', 'になること',
        'になっていること', 'ないこと', 'されないこと',
        'でないこと', 'いないこと', 'ていること',
        'れないこと', 'えること', 'くこと',
        'むこと', 'うこと', 'ぬこと',
        'つこと', 'ること',
    ]

    pattern_data = {}  # pattern -> {'count': N, 'examples': []}

    for item in expected_list:
        text = item['text']
        matched = False
        for pat in ending_patterns:
            if text.endswith(pat):
                if pat not in pattern_data:
                    pattern_data[pat] = {'count': 0, 'examples': []}
                pattern_data[pat]['count'] += 1
                pattern_data[pat]['examples'].append(item)
                matched = True
                break
        if not matched:
            other = 'その他'
            if other not in pattern_data:
                pattern_data[other] = {'count': 0, 'examples': []}
            pattern_data[other]['count'] += 1
            pattern_data[other]['examples'].append(item)

    # 出現回数でソート
    sorted_patterns = sorted(pattern_data.items(), key=lambda x: x[1]['count'], reverse=True)

    lines.append("| # | 末尾パターン | 出現回数 | 出現率 | 具体的な例文（最大3件） |")
    lines.append("|---|-------------|---------|--------|------------------------|")

    for i, (pat, data) in enumerate(sorted_patterns[:20], 1):
        rate = data['count'] / total_expected * 100

        # 異なるファイルからできるだけ3件選ぶ
        examples = data['examples']
        selected = []
        seen_files = set()
        # まず各ファイルから1件ずつ
        for ex in examples:
            if ex['file'] not in seen_files and len(selected) < 3:
                selected.append(ex)
                seen_files.add(ex['file'])
        # 3件に満たない場合は追加
        for ex in examples:
            if len(selected) >= 3:
                break
            if ex not in selected:
                selected.append(ex)

        example_texts = []
        for ex in selected[:3]:
            text_short = ex['text'].replace('\n', ' ').replace('|', '\\|')
            if len(text_short) > 60:
                text_short = text_short[:57] + '...'
            example_texts.append(f"[{ex['file']}] {text_short}")

        examples_str = '<br>'.join(example_texts)
        lines.append(f"| {i} | 〜{pat} | {data['count']} | {rate:.1f}% | {examples_str} |")

    lines.append("")
    return '\n'.join(lines)


# =====================================================================
# 分析3: 「こと」が2回以上含まれるケースの全リスト
# =====================================================================

def analysis_3(all_data):
    """「こと」が2回以上含まれる期待値を全件抽出"""
    lines = []
    lines.append("## 分析3: 「こと」が2回以上含まれるケースの全リスト\n")

    found = []
    total_expected = 0
    for case in all_data:
        exp = case.get('expected', '').strip()
        if not exp or exp == '-' or exp == 'None':
            continue
        total_expected += 1
        count = exp.count('こと')
        if count >= 2:
            found.append({
                'file': case['file'],
                'sheet': case['sheet'],
                'row': case['row'],
                'expected': exp,
                'count': count,
            })

    lines.append(f"- 全期待値数: {total_expected}件")
    lines.append(f"- 「こと」2回以上: {len(found)}件")
    if total_expected > 0:
        lines.append(f"- 該当率: {len(found)/total_expected*100:.1f}%")
    lines.append("")

    if found:
        lines.append("| # | ファイル | シート名 | 行番号 | 期待値（全文） | 「こと」回数 |")
        lines.append("|---|---------|----------|--------|----------------|-------------|")
        for i, item in enumerate(found, 1):
            exp_escaped = item['expected'].replace('\n', ' ').replace('|', '\\|')
            lines.append(f"| {i} | {item['file']} | {item['sheet']} | {item['row']} | {exp_escaped} | {item['count']}回 |")
    else:
        lines.append("「こと」が2回以上含まれるケースは検出されませんでした。")

    lines.append("")
    return '\n'.join(lines)


# =====================================================================
# 分析4: 期待値の文字数分布
# =====================================================================

def analysis_4(all_data):
    """期待値の文字数分布をヒストグラム的に集計"""
    lines = []
    lines.append("## 分析4: 期待値の文字数分布\n")

    bins = [
        (0, 5, '0-5字'),
        (6, 10, '6-10字'),
        (11, 15, '11-15字'),
        (16, 20, '16-20字'),
        (21, 25, '21-25字'),
        (26, 30, '26-30字'),
        (31, 40, '31-40字'),
        (41, 50, '41-50字'),
        (51, 60, '51-60字'),
        (61, 80, '61-80字'),
        (81, 99999, '81字以上'),
    ]

    # ファイル別に集計
    file_stats = {}
    for file_info in FILES:
        name = file_info['short_name']
        file_stats[name] = {b[2]: 0 for b in bins}
        file_stats[name]['_total'] = 0
        file_stats[name]['_lengths'] = []

    all_bin_counts = {b[2]: 0 for b in bins}
    all_lengths = []

    for case in all_data:
        exp = case.get('expected', '').strip()
        if not exp or exp == '-' or exp == 'None':
            continue
        # 改行を除去して文字数計算
        length = len(exp.replace('\n', '').replace('\r', '').strip())
        fname = case['file']

        if fname not in file_stats:
            continue

        file_stats[fname]['_total'] += 1
        file_stats[fname]['_lengths'].append(length)
        all_lengths.append(length)

        for low, high, label in bins:
            if low <= length <= high:
                file_stats[fname][label] += 1
                all_bin_counts[label] += 1
                break

    # テーブルヘッダー
    file_names = [f['short_name'] for f in FILES]
    header = "| 文字数範囲 | " + " | ".join(file_names) + " | 全体 |"
    separator = "|-----------|" + "|".join(["-----" for _ in file_names]) + "|------|"

    lines.append(header)
    lines.append(separator)

    for _, _, label in bins:
        row_vals = []
        for name in file_names:
            cnt = file_stats[name][label]
            total = file_stats[name]['_total']
            if total > 0:
                pct = cnt / total * 100
                row_vals.append(f"{cnt} ({pct:.0f}%)")
            else:
                row_vals.append("0")
        all_cnt = all_bin_counts[label]
        all_total = len(all_lengths) if all_lengths else 1
        all_pct = all_cnt / all_total * 100
        row_vals.append(f"{all_cnt} ({all_pct:.0f}%)")
        lines.append(f"| {label} | " + " | ".join(row_vals) + " |")

    # 合計行
    total_row = []
    for name in file_names:
        total_row.append(f"{file_stats[name]['_total']}")
    total_row.append(f"{len(all_lengths)}")
    lines.append(f"| **合計** | " + " | ".join(total_row) + " |")

    lines.append("")

    # 統計サマリー
    lines.append("### 文字数の統計サマリー\n")
    lines.append("| ファイル | 件数 | 平均 | 中央値 | 最小 | 最大 | 標準偏差 |")
    lines.append("|---------|------|------|--------|------|------|---------|")

    for name in file_names:
        lengths = file_stats[name]['_lengths']
        if lengths:
            avg = sum(lengths) / len(lengths)
            sorted_l = sorted(lengths)
            mid = len(sorted_l) // 2
            median = sorted_l[mid] if len(sorted_l) % 2 == 1 else (sorted_l[mid-1] + sorted_l[mid]) / 2
            min_l = min(lengths)
            max_l = max(lengths)
            variance = sum((x - avg) ** 2 for x in lengths) / len(lengths)
            std = variance ** 0.5
            lines.append(f"| {name} | {len(lengths)} | {avg:.1f}字 | {median:.0f}字 | {min_l}字 | {max_l}字 | {std:.1f} |")
        else:
            lines.append(f"| {name} | 0 | - | - | - | - | - |")

    if all_lengths:
        avg = sum(all_lengths) / len(all_lengths)
        sorted_l = sorted(all_lengths)
        mid = len(sorted_l) // 2
        median = sorted_l[mid] if len(sorted_l) % 2 == 1 else (sorted_l[mid-1] + sorted_l[mid]) / 2
        min_l = min(all_lengths)
        max_l = max(all_lengths)
        variance = sum((x - avg) ** 2 for x in all_lengths) / len(all_lengths)
        std = variance ** 0.5
        lines.append(f"| **全体** | {len(all_lengths)} | {avg:.1f}字 | {median:.0f}字 | {min_l}字 | {max_l}字 | {std:.1f} |")

    lines.append("")

    # テキストヒストグラム
    lines.append("### 視覚的な文字数分布（全体）\n")
    lines.append("```")
    max_count = max(all_bin_counts.values()) if all_bin_counts else 1
    for _, _, label in bins:
        cnt = all_bin_counts[label]
        bar_len = int(cnt / max_count * 40) if max_count > 0 else 0
        bar = '█' * bar_len
        lines.append(f"  {label:>8s} | {bar} {cnt}")
    lines.append("```")
    lines.append("")

    return '\n'.join(lines)


# =====================================================================
# 分析5: 手順の書き方パターン分析
# =====================================================================

def analysis_5(all_data):
    """手順の書き方パターン分析"""
    lines = []
    lines.append("## 分析5: 手順の書き方パターン分析\n")

    # 構造化キーワード
    structure_keywords = [
        '【前提条件】', '【前提】', '【セットアップ】',
        '【実行手順】', '【確認手順】', '【事前準備】',
        '【手順】',
    ]

    # ファイル別統計
    file_names = [f['short_name'] for f in FILES]
    file_step_stats = {}
    for name in file_names:
        file_step_stats[name] = {
            'total': 0,
            'with_steps': 0,
            'with_numbered': 0,
            'with_structure': 0,
            'structure_counts': collections.Counter(),
            'step_lengths': [],
            'good_examples': [],
        }

    all_total = 0
    all_with_steps = 0
    all_with_numbered = 0
    all_with_structure = 0
    all_structure_counts = collections.Counter()
    all_good_examples = []

    for case in all_data:
        steps = case.get('steps', '').strip()
        fname = case['file']
        if fname not in file_step_stats:
            continue

        file_step_stats[fname]['total'] += 1
        all_total += 1

        if not steps or steps == '-' or steps == 'None':
            continue

        file_step_stats[fname]['with_steps'] += 1
        all_with_steps += 1
        file_step_stats[fname]['step_lengths'].append(len(steps))

        # 番号付き手順チェック
        if re.search(r'(?:^|\n)\s*[1１①⑴\(1\)][\.\．\s、）\)]', steps):
            file_step_stats[fname]['with_numbered'] += 1
            all_with_numbered += 1

        # 構造化キーワードチェック
        found_kw = False
        for kw in structure_keywords:
            if kw in steps:
                file_step_stats[fname]['structure_counts'][kw] += 1
                all_structure_counts[kw] += 1
                found_kw = True
        if found_kw:
            file_step_stats[fname]['with_structure'] += 1
            all_with_structure += 1

        # 良い例の候補（構造化キーワードあり + 番号付き + 30字以上）
        has_number = bool(re.search(r'(?:^|\n)\s*[1１①⑴\(1\)][\.\．\s、）\)]', steps))
        if found_kw and has_number and len(steps) >= 30:
            all_good_examples.append({
                'file': fname,
                'sheet': case['sheet'],
                'row': case['row'],
                'steps': steps,
                'expected': case.get('expected', ''),
                'score': len(steps),  # 長いほうがより詳細
            })

    # 1. 構造化キーワードの使用率
    lines.append("### 1. 構造化キーワードの使用率\n")
    lines.append("| ファイル | 総ケース数 | 手順あり | 構造化キーワードあり | 使用率 |")
    lines.append("|---------|-----------|---------|-------------------|--------|")
    for name in file_names:
        s = file_step_stats[name]
        rate = s['with_structure'] / s['total'] * 100 if s['total'] > 0 else 0
        lines.append(f"| {name} | {s['total']} | {s['with_steps']} | {s['with_structure']} | {rate:.1f}% |")
    rate = all_with_structure / all_total * 100 if all_total > 0 else 0
    lines.append(f"| **全体** | {all_total} | {all_with_steps} | {all_with_structure} | {rate:.1f}% |")
    lines.append("")

    # 構造化キーワード内訳
    lines.append("#### 構造化キーワードの出現回数\n")
    lines.append("| キーワード | 出現回数 |")
    lines.append("|-----------|---------|")
    for kw, cnt in all_structure_counts.most_common():
        lines.append(f"| {kw} | {cnt} |")
    lines.append("")

    # 2. 番号付き手順の使用率
    lines.append("### 2. 番号付き手順の使用率\n")
    lines.append("| ファイル | 総ケース数 | 手順あり | 番号付き手順 | 使用率（手順ありに対する比率） |")
    lines.append("|---------|-----------|---------|-------------|------------------------------|")
    for name in file_names:
        s = file_step_stats[name]
        rate = s['with_numbered'] / s['with_steps'] * 100 if s['with_steps'] > 0 else 0
        lines.append(f"| {name} | {s['total']} | {s['with_steps']} | {s['with_numbered']} | {rate:.1f}% |")
    rate = all_with_numbered / all_with_steps * 100 if all_with_steps > 0 else 0
    lines.append(f"| **全体** | {all_total} | {all_with_steps} | {all_with_numbered} | {rate:.1f}% |")
    lines.append("")

    # 3. 手順の平均文字数（ファイル別）
    lines.append("### 3. 手順の平均文字数（ファイル別）\n")
    lines.append("| ファイル | 手順あり件数 | 平均文字数 | 中央値 | 最小 | 最大 |")
    lines.append("|---------|-----------|----------|--------|------|------|")
    all_step_lengths = []
    for name in file_names:
        lengths = file_step_stats[name]['step_lengths']
        all_step_lengths.extend(lengths)
        if lengths:
            avg = sum(lengths) / len(lengths)
            sorted_l = sorted(lengths)
            mid = len(sorted_l) // 2
            median = sorted_l[mid] if len(sorted_l) % 2 == 1 else (sorted_l[mid-1] + sorted_l[mid]) / 2
            lines.append(f"| {name} | {len(lengths)} | {avg:.0f}字 | {median:.0f}字 | {min(lengths)}字 | {max(lengths)}字 |")
        else:
            lines.append(f"| {name} | 0 | - | - | - | - |")
    if all_step_lengths:
        avg = sum(all_step_lengths) / len(all_step_lengths)
        sorted_l = sorted(all_step_lengths)
        mid = len(sorted_l) // 2
        median = sorted_l[mid] if len(sorted_l) % 2 == 1 else (sorted_l[mid-1] + sorted_l[mid]) / 2
        lines.append(f"| **全体** | {len(all_step_lengths)} | {avg:.0f}字 | {median:.0f}字 | {min(all_step_lengths)}字 | {max(all_step_lengths)}字 |")
    lines.append("")

    # 4. 手順の具体的な良い例を5件抽出
    lines.append("### 4. 手順の具体的な良い例（5件）\n")

    # スコア順にソートし、異なるファイルから選ぶ
    all_good_examples.sort(key=lambda x: x['score'], reverse=True)
    selected = []
    seen_files = set()
    # まずファイルを分散させる
    for ex in all_good_examples:
        if ex['file'] not in seen_files and len(selected) < 5:
            selected.append(ex)
            seen_files.add(ex['file'])
    # 5件に満たない場合は追加
    for ex in all_good_examples:
        if len(selected) >= 5:
            break
        if ex not in selected:
            selected.append(ex)

    for i, ex in enumerate(selected[:5], 1):
        lines.append(f"#### 良い例 {i}: [{ex['file']}] {ex['sheet']} / 行{ex['row']}\n")
        lines.append("**手順:**")
        lines.append("```")
        lines.append(ex['steps'])
        lines.append("```")
        lines.append("")
        lines.append(f"**期待値:** {ex['expected'].replace(chr(10), ' ')}")
        lines.append("")
        # 良い点を分析
        good_points = []
        if any(kw in ex['steps'] for kw in ['【前提条件】', '【前提】', '【セットアップ】']):
            good_points.append("前提条件/セットアップが明示されている")
        if any(kw in ex['steps'] for kw in ['【実行手順】', '【確認手順】', '【手順】']):
            good_points.append("実行手順セクションが構造化されている")
        if re.search(r'(?:^|\n)\s*[1１①⑴\(1\)][\.\．\s、）\)]', ex['steps']):
            good_points.append("番号付きステップで順序が明確")
        if len(ex['steps']) >= 80:
            good_points.append(f"十分な詳細さ（{len(ex['steps'])}字）")
        if good_points:
            lines.append("**良い点:** " + " / ".join(good_points))
        lines.append("")
        lines.append("---\n")

    return '\n'.join(lines)


# =====================================================================
# メイン
# =====================================================================

def main():
    print("=" * 60)
    print("高品質テスト項目書 深掘り分析")
    print("=" * 60)
    print()

    # 全ファイルからデータ読み込み
    all_data = []
    for file_info in FILES:
        print(f"読み込み中: {file_info['short_name']}...")
        cases = read_all_cases(file_info)
        print(f"  -> {len(cases)}件のテストケースを取得")
        all_data.extend(cases)

    print(f"\n合計: {len(all_data)}件のテストケース\n")

    # 各分析を実行
    print("分析1: 取引先マスタの曖昧表現の具体例...")
    report_1 = analysis_1(all_data)

    print("分析2: 期待値の良い書き方パターンTOP20...")
    report_2 = analysis_2(all_data)

    print("分析3: 「こと」が2回以上含まれるケースの全リスト...")
    report_3 = analysis_3(all_data)

    print("分析4: 期待値の文字数分布...")
    report_4 = analysis_4(all_data)

    print("分析5: 手順の書き方パターン分析...")
    report_5 = analysis_5(all_data)

    # Markdownレポート生成
    report = f"""# 高品質テスト項目書 詳細分析レポート

- 生成日: {datetime.now().strftime('%Y-%m-%d %H:%M')}
- 分析対象ファイル数: 4件
- 合計テストケース数: {len(all_data)}件

### 分析対象ファイル
1. クリッピング練習問題の複数ページ対応
2. ログイン方法のフラグでログイン制限
3. 取引先マスタのテーブル構造変更（企業マスタ追加）
4. 改善後_v7（ログイン制限の改善版）

### 除外シートパターン
改善前, 質問, 振り返り, レビュー指摘, memo, 運用ルール, 観点, 進捗, チェック表, バグ一覧, 設定 等

---

{report_1}

---

{report_2}

---

{report_3}

---

{report_4}

---

{report_5}
"""

    # レポート保存
    output_path = r'C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\評価ツール\学習レポート\2026-02-27\詳細分析レポート.md'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(report)

    print(f"\nレポートを保存しました: {output_path}")
    print()

    # レポートも標準出力に出力
    print(report)


if __name__ == '__main__':
    main()
