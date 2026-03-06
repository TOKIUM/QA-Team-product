# -*- coding: utf-8 -*-
"""
DEVSEC-133 オープンリダイレクタ: 観点シート→テストケース生成スクリプト
観点シートの階層構造を解析し、テストケースシートに書き込む
"""
import openpyxl
from common import format_test_steps

XLSX = "テスト項目書_[DEVSEC-133] オープンリダイレクタ.xlsx"


# --- 観点シート解析 ---
def parse_kanten(ws, start_row, end_row, product_label):
    """観点シートの階層構造を解析してテストケースのリストを返す"""
    cases = []
    ctx = {"confirm1": "", "confirm2": "", "screen": "", "target": "", "detail1": "", "detail2": ""}

    for r in range(start_row, end_row + 1):
        c2 = ws.cell(r, 2).value
        c5 = ws.cell(r, 5).value
        c9 = ws.cell(r, 9).value
        c12 = ws.cell(r, 12).value
        c15 = ws.cell(r, 15).value
        c18 = ws.cell(r, 18).value

        if c2:
            ctx["confirm1"] = str(c2).strip()
            ctx["confirm2"] = ""
            ctx["screen"] = ""
            ctx["target"] = ""
        if c5:
            ctx["confirm2"] = str(c5).strip()
            ctx["screen"] = ""
            ctx["target"] = ""
        if c9:
            ctx["screen"] = str(c9).strip()
            ctx["target"] = ""
        if c12:
            ctx["target"] = str(c12).strip()

        detail1 = str(c15).strip() if c15 else ""
        detail2 = str(c18).strip() if c18 else ""

        if ctx["target"].startswith("〇") and not detail1:
            continue
        if not detail1:
            continue

        cases.append({
            "product": product_label,
            "confirm": ctx["confirm2"] or ctx["confirm1"],
            "screen": ctx["screen"],
            "target": ctx["target"],
            "detail": detail1,
            "detail2": detail2,
        })
    return cases


# --- 手順/期待値ルール ---
# (preconditions, steps, expected) のタプルを返す

def _fmt(pre, steps):
    """format_test_stepsのショートカット"""
    return format_test_steps(pre, steps)


def generate_steps_and_expected(case):
    """テストケースから手順と期待値を生成（format_test_steps使用）"""
    product = case["product"]
    confirm = case["confirm"]
    screen = case["screen"]
    target = case["target"]
    detail = case["detail"]
    detail2 = case["detail2"]

    if product == "TH":
        base_url = "dev.keihi.com（THテナント）"
    else:
        base_url = "WDLのログイン画面"

    steps_text = ""
    expected = ""

    # === パスワード再設定フロー ===
    if confirm == "パスワード再設定":
        if screen == "ログイン画面":
            if "忘れた場合" in target:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「パスワードを忘れた場合」リンクをクリックする",
                ])
                expected = "パスワード再設定画面へ遷移すること"
            elif "変更後のログイン" in target:
                if "変更前" in detail:
                    steps_text = _fmt(
                        ["パスワード再設定が完了していること"],
                        [f"{base_url}のログイン画面を開く",
                         "変更前のパスワードでログインを試みる"])
                    expected = "ログインに失敗し、エラーメッセージが表示されること"
                elif "変更後" in detail:
                    steps_text = _fmt(
                        ["パスワード再設定が完了していること"],
                        [f"{base_url}のログイン画面を開く",
                         "変更後のパスワードでログインする"])
                    expected = "正常にログインできること"
        elif screen == "パスワード再設定画面":
            if "メールアドレス入力" in target:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「パスワードを忘れた場合」リンクをクリックする",
                    "メールアドレス入力欄にメールアドレスを入力する",
                ])
                expected = "メールアドレスが入力できること"
            elif "送信" in target:
                if "連打" in detail:
                    steps_text = _fmt([], [
                        f"{base_url}のログイン画面を開く",
                        "「パスワードを忘れた場合」リンクをクリックする",
                        "メールアドレスを入力する",
                        "「送信」ボタンを連続でクリックする",
                    ])
                    expected = "エラーにならず、メールが重複送信されないこと"
                else:
                    steps_text = _fmt([], [
                        f"{base_url}のログイン画面を開く",
                        "「パスワードを忘れた場合」リンクをクリックする",
                        "メールアドレスを入力する",
                        "「送信」ボタンをクリックする",
                    ])
                    expected = "送信が完了し、メール送信完了の旨が表示されること"
            elif "メール受信" in target:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「パスワードを忘れた場合」リンクをクリックする",
                    "メールアドレスを入力し、「送信」ボタンをクリックする",
                    "メール受信を確認する",
                ])
                expected = "入力したメールアドレス宛にパスワード再設定メールが届くこと"
            elif "メール記載のリンク" in target:
                steps_text = _fmt(
                    ["パスワード再設定メールを受信していること"],
                    ["受信したメールを開く",
                     "メール本文に記載されたリンクをクリックする"])
                expected = "パスワード変更画面へ遷移すること"
        elif screen == "パスワード変更画面":
            if "パスワード入力" == target:
                steps_text = _fmt(
                    ["パスワード再設定メールのリンクからパスワード変更画面を開いていること"],
                    ["「パスワード」入力欄に新しいパスワードを入力する"])
                expected = "パスワードが入力できること"
            elif "再入力" in target:
                steps_text = _fmt(
                    ["パスワード変更画面でパスワードを入力済みであること"],
                    ["「パスワード(再入力)」欄に同じパスワードを入力する"])
                expected = "パスワードが入力できること"
            elif "変更する" in target:
                if "連打" in detail:
                    steps_text = _fmt(
                        ["パスワード変更画面でパスワードを入力済みであること"],
                        ["「パスワードを変更する」ボタンを連続でクリックする"])
                    expected = "エラーにならず、パスワード変更が正常に完了すること"
                else:
                    steps_text = _fmt(
                        ["パスワード変更画面でパスワードを入力済みであること"],
                        ["「パスワードを変更する」ボタンをクリックする"])
                    expected = "パスワード変更が完了し、完了メッセージが表示されること"

    # === 新規登録フロー ===
    elif confirm == "新規登録":
        if screen == "ログイン画面":
            if "新規登録はこちら" in target:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「新規登録はこちら」リンクをクリックする",
                ])
                expected = "ユーザー登録画面へ遷移すること"
            elif "登録完了後のログイン" in target:
                if "ログイン後" in detail and "戻る" in detail:
                    steps_text = _fmt(
                        ["新規登録が完了し、ログイン済みであること"],
                        ["ブラウザの戻るボタンを押下する"])
                    expected = "ログイン後の画面にとどまること（ログイン画面に戻らないこと）"
                elif "ログアウト後" in detail and "戻る" in detail:
                    steps_text = _fmt(
                        ["ログイン後、ログアウトしていること"],
                        ["ブラウザの戻るボタンを押下する"])
                    expected = "ログイン画面にとどまること（ログイン後の画面が表示されないこと）"
                else:
                    steps_text = _fmt(
                        ["新規登録が完了していること"],
                        [f"{base_url}のログイン画面を開く",
                         "登録したメールアドレスとパスワードでログインする"])
                    expected = "正常にログインできること"
        elif screen == "ユーザー登録画面":
            if "メールアドレス入力" in target:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「新規登録はこちら」リンクをクリックする",
                    "メールアドレス入力欄にメールアドレスを入力する",
                ])
                expected = "メールアドレスが入力できること"
            elif target == "パスワード入力":
                steps_text = _fmt(
                    ["ユーザー登録画面を開いていること"],
                    ["「パスワード」入力欄にパスワードを入力する"])
                expected = "パスワードが入力できること"
            elif "再入力" in target and "パスワード" in target:
                steps_text = _fmt(
                    ["ユーザー登録画面を開いていること"],
                    ["「パスワード(再入力)」欄に同じパスワードを入力する"])
                expected = "パスワードが入力できること"
            elif "ユーザー登録する" in target:
                if "エラー" in detail:
                    steps_text = _fmt(
                        ["予約のないユーザーのメールアドレスでユーザー登録画面を開いていること"],
                        ["メールアドレス・パスワードを入力する",
                         "「ユーザー登録する」ボタンをクリックする"])
                    expected = "エラーメッセージが表示され、登録できないこと"
                elif "連打" in detail:
                    steps_text = _fmt(
                        ["ユーザー登録画面で必要情報を入力済みであること"],
                        ["「ユーザー登録する」ボタンを連続でクリックする"])
                    expected = "エラーにならず、ユーザー登録が正常に完了すること"
                else:
                    steps_text = _fmt(
                        ["ユーザー登録画面で必要情報を入力済みであること"],
                        ["「ユーザー登録する」ボタンをクリックする"])
                    expected = "ユーザー登録が完了し、確認メールが送信されること"
            elif "メール受信" in target:
                steps_text = _fmt(
                    ["ユーザー登録が完了していること"],
                    ["登録したメールアドレスのメール受信を確認する"])
                expected = "確認メールが届くこと"
            elif "メール記載のリンク" in target:
                steps_text = _fmt(
                    ["確認メールを受信していること"],
                    ["受信した確認メールを開く",
                     "メール本文に記載されたリンクをクリックする"])
                expected = "登録完了画面へ遷移すること"
            elif "氏名" in target:
                steps_text = _fmt(
                    ["ユーザー登録画面（初回ログイン）を開いていること"],
                    ["「氏名」入力欄に氏名を入力する"])
                expected = "氏名が入力できること"
            elif "ユーザー登録を続ける" in target:
                if "連打" in detail:
                    steps_text = _fmt(
                        ["ユーザー登録画面で必要情報を入力済みであること"],
                        ["「ユーザー登録を続ける」ボタンを連続でクリックする"])
                    expected = "エラーにならず、次の画面へ正常に遷移すること"
                else:
                    steps_text = _fmt(
                        ["ユーザー登録画面で氏名・メールアドレス・パスワードを入力済みであること"],
                        ["「ユーザー登録を続ける」ボタンをクリックする"])
                    expected = "次の画面（送付元識別コード入力）へ遷移すること"
            elif "送付元識別コード" in target:
                if "異常系" in target:
                    if "小文字" in detail:
                        steps_text = _fmt(
                            ["送付元識別コード入力画面を開いていること"],
                            ["メール記載の送付元識別コードを全文小文字で入力する",
                             "「ユーザー登録する」ボタンをクリックする"])
                        expected = "小文字で入力しても大文字で記入されること"
                    else:
                        steps_text = _fmt(
                            ["送付元識別コード入力画面を開いていること"],
                            ["メール記載の送付元識別コードから1文字削除した値を入力する",
                             "「ユーザー登録する」ボタンをクリックする"])
                        expected = "エラーが表示されること"
                else:
                    steps_text = _fmt(
                        ["送付元識別コード入力画面を開いていること"],
                        ["メール記載の「送付元識別コード」を入力する"])
                    expected = "送付元識別コードが入力できること"

    # === アカウントロック解除フロー ===
    elif confirm == "アカウントロック解除":
        if "5回以上失敗" in target:
            steps_text = _fmt([], [
                f"{base_url}のログイン画面を開く",
                "正しいメールアドレスと誤ったパスワードでログインを5回以上試行する",
                "メール受信を確認する",
            ])
            expected = "アカウントロックされ、設定したメールアドレスにロック解除メールが届くこと"
        elif "ロック中" in target:
            steps_text = _fmt(
                ["アカウントがロックされていること"],
                [f"{base_url}のログイン画面を開く",
                 "正しいメールアドレスとパスワードでログインを試みる"])
            expected = "ロックされていてログインできない状態であること"
        elif "メール記載のリンク" in target:
            steps_text = _fmt(
                ["ロック解除メールを受信していること"],
                ["受信したメールを開く",
                 "メール本文に記載されたリンクをクリックする"])
            expected = "ロック解除完了画面へ遷移すること"
        elif "ロック解除完了後" in target:
            steps_text = _fmt(
                ["ロック解除が完了していること"],
                [f"{base_url}のログイン画面を開く",
                 "正しいメールアドレスとパスワードでログインする"])
            expected = "正常にログインできること"

    # === TOKIUM IDでログイン (THのみ) ===
    elif "TOKIUM ID" in confirm:
        if "新規登録" in confirm:
            steps_text = _fmt(
                ["TOKIUM本体にアカウントを作成済みであること",
                 "開発に上記アカウントの予約情報の作成/THとの連携を依頼済みであること"],
                [f"{base_url}のログイン画面を開く",
                 "「TOKIUM IDでログイン」ボタンをクリックする",
                 "TOKIUM IDの認証情報でログインする"])
            expected = "TH画面の請求書画面が表示されること"
        elif screen == "ログイン画面":
            if "連打" in detail:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「TOKIUM IDでログイン」ボタンを連続でクリックする",
                ])
                expected = "エラーにならず、正常にTOKIUM ID認証画面へ遷移すること"
            elif "遷移" in detail:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「TOKIUM IDでログイン」ボタンをクリックする",
                ])
                expected = "TOKIUM ID認証画面へ遷移すること"
            else:
                steps_text = _fmt([], [
                    f"{base_url}のログイン画面を開く",
                    "「TOKIUM IDでログイン」ボタンをクリックする",
                ])
                expected = "ボタンが押下でき、反応すること"
        elif screen == "TOKIUM ID認証画面":
            if "メールアドレス入力" in target:
                steps_text = _fmt(
                    ["TOKIUM ID認証画面を開いていること"],
                    ["メールアドレス入力欄にメールアドレスを入力する"])
                expected = "メールアドレスが入力できること"
            elif target == "パスワード入力":
                steps_text = _fmt(
                    ["TOKIUM ID認証画面を開いていること"],
                    ["パスワード入力欄にパスワードを入力する"])
                expected = "パスワードが入力できること"
            elif "ログイン" in target:
                if "連打" in detail:
                    steps_text = _fmt(
                        ["TOKIUM ID認証画面でメールアドレス・パスワードを入力済みであること"],
                        ["「ログイン」ボタンを連続でクリックする"])
                    expected = "エラーにならず、正常にログイン処理が完了すること"
                elif "遷移" in detail:
                    steps_text = _fmt(
                        ["TOKIUM ID認証画面でメールアドレス・パスワードを入力済みであること"],
                        ["「ログイン」ボタンをクリックする"])
                    expected = "THログイン後の画面へ遷移すること"
                else:
                    steps_text = _fmt(
                        ["TOKIUM ID認証画面でメールアドレス・パスワードを入力済みであること"],
                        ["「ログイン」ボタンをクリックする"])
                    expected = "ボタンが押下でき、反応すること"

    # フォールバック
    if not steps_text:
        steps_text = _fmt([], [f"{screen}を開く", f"{target}を操作する"])
        expected = detail2 if detail2 else detail

    return steps_text, expected


def write_testcases(wb, cases):
    """テストケースシートに書き込む"""
    ws = wb["テストケース"]
    data_start = 3

    row = data_start
    for i, case in enumerate(cases, 1):
        ws.cell(row, 2, i)  # No
        ws.cell(row, 3, case["confirm"])  # 確認すること
        ws.cell(row, 4, case["screen"])  # 画面
        ws.cell(row, 5, case["target"])  # 確認対象
        ws.cell(row, 6, case["detail"])  # 詳細
        steps, expected = generate_steps_and_expected(case)
        ws.cell(row, 7, steps)  # テスト実行手順
        ws.cell(row, 8, expected)  # 期待値
        ws.cell(row, 9, case.get("detail2", ""))  # 備考

        # 環境列をクリアしてからマーク
        for col in range(10, 18):
            ws.cell(row, col, None)
        if case["product"] == "TH":
            ws.cell(row, 14, "未")
        else:
            ws.cell(row, 15, "未")

        row += 1

    # 残りのテンプレート行をクリア
    for r in range(row, min(row + 20, ws.max_row + 1)):
        for c in range(2, 21):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c, None)

    return len(cases)


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws_kanten = wb["観点"]

    th_cases = parse_kanten(ws_kanten, 36, 74, "TH")
    wdl_cases = parse_kanten(ws_kanten, 79, 118, "WDL")
    all_cases = th_cases + wdl_cases

    print(f"TH: {len(th_cases)}件 / WDL: {len(wdl_cases)}件 / 合計: {len(all_cases)}件\n")

    # サンプル出力（前提条件ありのケースを3件表示）
    shown = 0
    for i, c in enumerate(all_cases, 1):
        steps, expected = generate_steps_and_expected(c)
        if "【前提条件】" in steps and shown < 3:
            print(f"--- TC{i} [{c['product']}] {c['confirm']} / {c['target']} ---")
            print(steps)
            print(f"→ 期待値: {expected}\n")
            shown += 1

    count = write_testcases(wb, all_cases)
    wb.save(XLSX)
    print(f"=== {count}件をテストケースシートに書き込みました ===")


if __name__ == "__main__":
    main()
