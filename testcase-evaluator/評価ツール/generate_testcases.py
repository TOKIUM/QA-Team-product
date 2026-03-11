#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
テスト項目書 自動生成ツール（JIRA連携 + Excel変換 + 評価）

JIRAチケットの取得、JSON→Excel変換、自動評価を統合したツール。
Claude Codeと組み合わせて仕様→テスト項目書の自動生成パイプラインを構成する。

使い方:
  # JIRAチケットを取得して仕様書Markdownを出力
  python generate_testcases.py --jira TH-1234

  # 複数チケットをまとめて取得
  python generate_testcases.py --jira TH-1234 TH-1235 TH-1236

  # JSON → Excel変換 + 評価（テンプレート準拠）
  python generate_testcases.py input.json -o output.xlsx

  # 標準入力からJSON
  echo '[{...}]' | python generate_testcases.py -

  # 評価なしでExcel変換のみ
  python generate_testcases.py input.json --no-eval

  # 観点の自動選択付き（仕様書テキストから判定）
  python generate_testcases.py input.json --spec spec.md -o output.xlsx
"""

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

import io
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl")
    sys.exit(1)


SCRIPT_DIR = Path(__file__).parent
JIRA_CONFIG_PATH = Path.home() / ".config" / "daily-summary" / "jira_config.json"
TEMPLATE_PATH = SCRIPT_DIR.parent / "テスト項目書テンプレート" / "【テンプレ】　テスト項目書_XXX.xlsx"

# テストケースシートの列マッピング（テンプレート準拠: B列開始, Row2がヘッダー）
TC_COL = {
    "no": 2,          # B: No (数式)
    "confirm": 3,     # C: 確認すること
    "screen": 4,      # D: 画面
    "target": 5,      # E: 確認対象
    "detail": 6,      # F: 詳細
    "steps": 7,       # G: テスト実行手順
    "expected": 8,     # H: 期待値
    "notes": 9,        # I: 備考
    "autify": 10,      # J: Autify
    "exec_k": 11,      # K: TK+TI+TD (Win11/Chrome)
    "exec_l": 12,      # L: TK+TD
    "exec_m": 13,      # M: TI+TD (Win11/Edge)
    "exec_n": 14,      # N: TK単体 (Win11/Firefox)
    "exec_o": 15,      # O: TD単体 (Mac/Safari)
    "exec_p": 16,      # P: TK+TI+TD (iOS)
    "exec_q": 17,      # Q: TK+TI+TD (Android)
    "executor": 18,    # R: 実行者
    "exec_notes": 19,  # S: 実行備考
    "jira_no": 20,     # T: 起票JIRA No
}

TC_HEADER_ROW = 2
TC_DATA_START_ROW = 3

# 観点シートのセルマッピング
KANTEN_CELLS = {
    # 環境（B11:C16）
    "env": {
        "TK+TI+TD": ("C", 11),
        "TK+TD": ("C", 12),
        "TI+TD": ("C", 13),
        "TK単体": ("C", 14),
        "TD単体": ("C", 15),
        "TDセルフプラン": ("C", 16),
    },
    # OS（F11:F12）
    "os": {
        "Windows": ("F", 11),
        "Mac": ("F", 12),
    },
    # ブラウザ（F15:F18）
    "browser": {
        "Chrome": ("F", 15),
        "Edge": ("F", 16),
        "Firefox": ("F", 17),
        "Safari": ("F", 18),
    },
    # アプリ（I11:I22）
    "app": {
        "iOS_iOS26": ("I", 11),
        "iOS_iOS18": ("I", 12),
        "iOS_iOS17": ("I", 13),
        "iOS_iOS16": ("I", 14),
        "iPad": ("I", 15),
        "Android_OS15": ("I", 16),
        "Android_OS14": ("I", 17),
        "Android_OS13": ("I", 18),
        "Android_OS12": ("I", 19),
        "Android_OS11": ("I", 20),
        "Android_OS10": ("I", 21),
        "Android_OS9": ("I", 22),
    },
    # 主要観点（L11:L25）
    "perspective": {
        "負荷": ("L", 11),
        "金額確認": ("L", 12),
        "時間経過": ("L", 13),
        "高頻度": ("L", 14),
        "既存機能への影響": ("L", 15),
        "互換性": ("L", 16),
        "データ移行": ("L", 17),
        "UIUX": ("L", 18),
        "セキュリティ/権限": ("L", 19),
        "お客様提供ドキュメント": ("L", 20),
        "統一性": ("L", 21),
        "英語化": ("L", 22),
        "外部API": ("L", 23),
        "本番環境への影響(チャットボタン等)": ("L", 24),
        "契約変更(プラン・オプション変更)": ("L", 25),
    },
    # 権限（N11:N15）
    "permission": {
        "権限なし": ("N", 11),
        "集計者": ("N", 12),
        "承認者": ("N", 13),
        "管理者": ("N", 14),
        "取引先管理者": ("N", 15),
    },
}


# =====================================================================
# JIRA取得
# =====================================================================

def load_jira_config():
    """JIRA設定を読み込む"""
    if not JIRA_CONFIG_PATH.exists():
        print(f"ERROR: JIRA設定ファイルが見つかりません: {JIRA_CONFIG_PATH}")
        print("  以下の形式で作成してください:")
        print('  {"email": "...", "api_token": "...", "base_url": "https://xxx.atlassian.net"}')
        return None

    with open(JIRA_CONFIG_PATH, encoding="utf-8") as f:
        config = json.load(f)

    required = ["email", "api_token", "base_url"]
    missing = [k for k in required if not config.get(k)]
    if missing:
        print(f"ERROR: JIRA設定に不足があります: {missing}")
        return None

    return config


def fetch_jira_ticket(config, ticket_key):
    """JIRAチケットの詳細を取得する"""
    import urllib.request
    import urllib.error
    import base64

    base_url = config["base_url"].rstrip("/")
    auth_str = f"{config['email']}:{config['api_token']}"
    auth_b64 = base64.b64encode(auth_str.encode()).decode()

    url = (
        f"{base_url}/rest/api/3/issue/{ticket_key}"
        f"?fields=summary,description,status,priority,issuetype,"
        f"assignee,reporter,labels,components,fixVersions,"
        f"customfield_10014,parent,subtasks,issuelinks,attachment,comment"
    )

    req = urllib.request.Request(url, headers={
        "Accept": "application/json",
        "Authorization": f"Basic {auth_b64}",
    })

    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:500]
        print(f"ERROR: JIRA API {e.code}: {ticket_key}")
        print(f"  {body}")
        return None
    except urllib.error.URLError as e:
        print(f"ERROR: JIRA接続失敗: {e}")
        return None


def adf_to_markdown(node, depth=0):
    """Atlassian Document Format (ADF) をMarkdownに変換する"""
    if not node or not isinstance(node, dict):
        return ""

    node_type = node.get("type", "")
    content = node.get("content", [])
    text = node.get("text", "")

    if node_type == "text":
        marks = node.get("marks", [])
        result = text
        for mark in marks:
            mt = mark.get("type", "")
            if mt == "strong":
                result = f"**{result}**"
            elif mt == "em":
                result = f"*{result}*"
            elif mt == "code":
                result = f"`{result}`"
            elif mt == "link":
                href = mark.get("attrs", {}).get("href", "")
                result = f"[{result}]({href})"
        return result

    parts = []
    for child in content:
        parts.append(adf_to_markdown(child, depth))

    joined = "".join(parts)

    if node_type == "paragraph":
        return joined + "\n\n"
    elif node_type == "heading":
        level = node.get("attrs", {}).get("level", 1)
        return "#" * level + " " + joined + "\n\n"
    elif node_type == "bulletList":
        items = []
        for child in content:
            item_text = adf_to_markdown(child, depth + 1).strip()
            items.append(f"{'  ' * depth}- {item_text}")
        return "\n".join(items) + "\n\n"
    elif node_type == "orderedList":
        items = []
        for i, child in enumerate(content, 1):
            item_text = adf_to_markdown(child, depth + 1).strip()
            items.append(f"{'  ' * depth}{i}. {item_text}")
        return "\n".join(items) + "\n\n"
    elif node_type == "listItem":
        return joined
    elif node_type == "codeBlock":
        lang = node.get("attrs", {}).get("language", "")
        return f"```{lang}\n{joined}```\n\n"
    elif node_type == "blockquote":
        lines = joined.strip().split("\n")
        return "\n".join(f"> {line}" for line in lines) + "\n\n"
    elif node_type == "table":
        return format_adf_table(content)
    elif node_type == "tableRow":
        return joined
    elif node_type in ("tableCell", "tableHeader"):
        return joined.strip()
    elif node_type == "panel":
        panel_type = node.get("attrs", {}).get("panelType", "info")
        return f"> [{panel_type}] {joined}\n\n"
    elif node_type == "rule":
        return "---\n\n"
    elif node_type == "doc":
        return joined

    return joined


def format_adf_table(rows):
    """ADF tableノードをMarkdownテーブルに変換"""
    md_rows = []
    for row_node in rows:
        cells = []
        for cell_node in row_node.get("content", []):
            cell_text = adf_to_markdown(cell_node).strip().replace("\n", " ")
            cells.append(cell_text)
        md_rows.append("| " + " | ".join(cells) + " |")

    if len(md_rows) > 0:
        col_count = md_rows[0].count("|") - 1
        separator = "| " + " | ".join(["---"] * max(col_count, 1)) + " |"
        md_rows.insert(1, separator)

    return "\n".join(md_rows) + "\n\n"


def ticket_to_markdown(ticket_data):
    """JIRAチケットデータをMarkdown仕様書形式に変換する"""
    fields = ticket_data.get("fields", {})
    key = ticket_data.get("key", "")

    lines = []
    lines.append(f"# {key}: {fields.get('summary', '(タイトルなし)')}")
    lines.append("")

    # メタ情報
    meta = []
    if fields.get("issuetype"):
        meta.append(f"種類: {fields['issuetype'].get('name', '')}")
    if fields.get("status"):
        meta.append(f"ステータス: {fields['status'].get('name', '')}")
    if fields.get("priority"):
        meta.append(f"優先度: {fields['priority'].get('name', '')}")
    if fields.get("assignee"):
        meta.append(f"担当: {fields['assignee'].get('displayName', '')}")
    if fields.get("reporter"):
        meta.append(f"報告者: {fields['reporter'].get('displayName', '')}")
    if fields.get("labels"):
        meta.append(f"ラベル: {', '.join(fields['labels'])}")
    if fields.get("components"):
        names = [c.get("name", "") for c in fields["components"]]
        meta.append(f"コンポーネント: {', '.join(names)}")
    if fields.get("fixVersions"):
        names = [v.get("name", "") for v in fields["fixVersions"]]
        meta.append(f"修正バージョン: {', '.join(names)}")
    if fields.get("parent"):
        parent = fields["parent"]
        meta.append(f"親チケット: {parent.get('key', '')} - {parent.get('fields', {}).get('summary', '')}")

    if meta:
        for m in meta:
            lines.append(f"- {m}")
        lines.append("")

    # 説明
    desc = fields.get("description")
    if desc:
        lines.append("## 説明")
        lines.append("")
        if isinstance(desc, dict):
            lines.append(adf_to_markdown(desc))
        elif isinstance(desc, str):
            lines.append(desc)
        lines.append("")

    # サブタスク
    subtasks = fields.get("subtasks", [])
    if subtasks:
        lines.append("## サブタスク")
        lines.append("")
        for st in subtasks:
            st_fields = st.get("fields", {})
            status = st_fields.get("status", {}).get("name", "")
            lines.append(f"- [{st.get('key', '')}] {st_fields.get('summary', '')} ({status})")
        lines.append("")

    # リンクされたチケット
    links = fields.get("issuelinks", [])
    if links:
        lines.append("## 関連チケット")
        lines.append("")
        for link in links:
            link_type = link.get("type", {}).get("outward", "関連")
            if link.get("outwardIssue"):
                target = link["outwardIssue"]
                lines.append(f"- {link_type}: [{target.get('key', '')}] {target.get('fields', {}).get('summary', '')}")
            elif link.get("inwardIssue"):
                link_type = link.get("type", {}).get("inward", "関連")
                target = link["inwardIssue"]
                lines.append(f"- {link_type}: [{target.get('key', '')}] {target.get('fields', {}).get('summary', '')}")
        lines.append("")

    # コメント（最新5件）
    comments = fields.get("comment", {})
    if isinstance(comments, dict):
        comment_list = comments.get("comments", [])
    elif isinstance(comments, list):
        comment_list = comments
    else:
        comment_list = []

    if comment_list:
        lines.append("## コメント（最新5件）")
        lines.append("")
        for c in comment_list[-5:]:
            author = c.get("author", {}).get("displayName", "不明")
            created = c.get("created", "")[:10]
            body = c.get("body", "")
            if isinstance(body, dict):
                body_text = adf_to_markdown(body).strip()
            elif isinstance(body, str):
                body_text = body
            else:
                body_text = str(body)
            lines.append(f"### {author} ({created})")
            lines.append(body_text)
            lines.append("")

    return "\n".join(lines)


def fetch_and_format_tickets(ticket_keys):
    """複数JIRAチケットを取得してMarkdownに変換する"""
    config = load_jira_config()
    if not config:
        sys.exit(1)

    all_md = []
    for key in ticket_keys:
        print(f"  JIRA取得中: {key} ...")
        data = fetch_jira_ticket(config, key)
        if data:
            md = ticket_to_markdown(data)
            all_md.append(md)
            print(f"  取得完了: {key} ({len(md)} 文字)")
        else:
            print(f"  取得失敗: {key}")

    return "\n\n---\n\n".join(all_md)


# =====================================================================
# 観点の自動判定
# =====================================================================

def analyze_perspectives(spec_text, test_cases=None):
    """仕様書テキストとテストケースから観点シートの選択項目を自動判定する。

    Returns:
        dict: カテゴリ別の選択結果
        {
            "env": {"TK+TI+TD": True, "TK+TD": False, ...},
            "os": {"Windows": True, "Mac": False},
            "browser": {"Chrome": True, ...},
            "app": {"iOS_iOS26": False, ...},
            "perspective": {"負荷": False, "UIUX": True, ...},
            "permission": {"権限なし": False, ...},
            "ticket_url": "...",
            "requirements_url": "...",
        }
    """
    text = (spec_text or "").lower()
    tc_text = ""
    if test_cases:
        tc_text = json.dumps(test_cases, ensure_ascii=False).lower()
    combined = text + " " + tc_text

    result = {
        "env": {},
        "os": {},
        "browser": {},
        "app": {},
        "perspective": {},
        "permission": {},
        "ticket_url": "",
        "requirements_url": "",
    }

    # --- チケットURL抽出 ---
    jira_match = re.search(r"(https?://\S+atlassian\.net/browse/\S+)", spec_text or "")
    if jira_match:
        result["ticket_url"] = jira_match.group(1)
    ticket_key_match = re.search(r"^#\s+([\w]+-\d+):", spec_text or "", re.MULTILINE)
    if ticket_key_match and not result["ticket_url"]:
        result["ticket_url"] = ticket_key_match.group(1)

    # --- 環境判定 ---
    # デフォルト: Web系の機能なら TK+TI+TD をベースに
    # 注意: "tk","ti","td","doc"等の短い略語は偽陽性を起こすため、より具体的なキーワードを使用
    has_tk = any(w in combined for w in ["tokium経費", "経費精算", "tokium expense", "経費申請"])
    has_ti = any(w in combined for w in ["tokium請求書", "請求書発行", "tokium invoice", "invoicing"])
    has_td = any(w in combined for w in ["tokiumdoc", "電子帳簿", "tokium document", "tokium doc"])

    if has_tk and has_ti and has_td:
        result["env"]["TK+TI+TD"] = True
    elif has_tk and has_td:
        result["env"]["TK+TD"] = True
    elif has_ti and has_td:
        result["env"]["TI+TD"] = True
    elif has_tk:
        result["env"]["TK単体"] = True
    elif has_td:
        result["env"]["TD単体"] = True
    else:
        # デフォルトは全環境
        result["env"]["TK+TI+TD"] = True

    # --- OS ---
    result["os"]["Windows"] = True  # デフォルトでWindows有効
    has_mac = any(w in combined for w in ["mac", "safari", "macos"])
    result["os"]["Mac"] = has_mac

    # --- ブラウザ ---
    result["browser"]["Chrome"] = True  # デフォルト
    result["browser"]["Edge"] = True    # デフォルト
    has_firefox = any(w in combined for w in ["firefox"])
    has_safari = has_mac or any(w in combined for w in ["safari"])
    result["browser"]["Firefox"] = has_firefox
    result["browser"]["Safari"] = has_safari

    # --- アプリ（モバイル） ---
    # 注意: "アプリ"は"アプリケーション"にもマッチするため除外。"モバイル"も汎用的すぎるため除外
    has_ios = any(w in combined for w in ["ios", "iphone", "ipad", "iosアプリ", "スマホ"])
    has_android = any(w in combined for w in ["android", "androidアプリ", "スマホ"])
    has_ipad = any(w in combined for w in ["ipad", "タブレット"])

    if has_ios:
        result["app"]["iOS_iOS26"] = True
        result["app"]["iOS_iOS18"] = True
    if has_ipad:
        result["app"]["iPad"] = True
    if has_android:
        result["app"]["Android_OS15"] = True
        result["app"]["Android_OS14"] = True

    # --- 主要観点 ---
    perspective_keywords = {
        "負荷": ["負荷テスト", "パフォーマンス", "高負荷", "速度改善", "レスポンス改善", "大量データ処理"],
        "金額確認": ["金額計算", "税計算", "消費税率", "端数", "按分", "丸め", "小数点",
                    "税率変更", "単価変更", "金額変更", "合計額", "小計額"],
        "時間経過": ["時間経過", "タイムアウト", "有効期限", "期限", "日付", "期間"],
        "高頻度": ["高頻度", "連打", "連続", "繰り返し"],
        "既存機能への影響": ["既存", "影響", "回帰", "リグレッション", "デグレ"],
        "互換性": ["互換", "移行", "バージョン", "下位互換"],
        "データ移行": ["データ移行", "マイグレーション", "移行"],
        "UIUX": ["ui", "ux", "画面", "表示", "デザイン", "レイアウト", "ボタン", "フォーム"],
        "セキュリティ/権限": ["セキュリティ", "権限", "認証", "ログイン", "パスワード", "アクセス制限",
                          "ip制限", "saml", "sso"],
        "お客様提供ドキュメント": ["ヘルプページ", "マニュアル更新", "利用ガイド", "ヘルプセンター"],
        "統一性": ["統一", "一貫性", "命名規則"],
        "英語化": ["英語", "多言語", "i18n", "翻訳", "ローカライズ"],
        "外部API": ["外部api", "外部連携", "webhook", "api連携", "apiキー", "api仕様"],
        "本番環境への影響(チャットボタン等)": ["本番", "チャット", "ウィジェット"],
        "契約変更(プラン・オプション変更)": ["プラン", "契約", "オプション", "ライセンス"],
    }

    for name, keywords in perspective_keywords.items():
        matched = any(kw in combined for kw in keywords)
        result["perspective"][name] = matched

    # UIUX は画面系のテストならほぼ常にTrue
    if test_cases and len(test_cases) > 0:
        result["perspective"]["UIUX"] = True
        result["perspective"]["既存機能への影響"] = True

    # --- 権限 ---
    # 注意: "承認"や"取引先"は機能名としても頻出するため、権限ロール名に限定
    permission_keywords = {
        "権限なし": ["権限なし", "一般ユーザー"],
        "集計者": ["集計者権限", "集計者ロール"],
        "承認者": ["承認者権限", "承認者ロール"],
        "管理者": ["管理者権限", "管理者ロール", "admin権限"],
        "取引先管理者": ["取引先管理者権限", "取引先管理者ロール"],
    }

    has_any_permission = False
    for name, keywords in permission_keywords.items():
        matched = any(kw in combined for kw in keywords)
        result["permission"][name] = matched
        if matched:
            has_any_permission = True

    # 権限の言及がなければ管理者のみ（一般的なテストパターン）
    if not has_any_permission:
        result["permission"]["管理者"] = True

    return result


# =====================================================================
# 観点シート — テスト設計の観点を自動生成
# =====================================================================

# 観点シートレイアウト定数
KANTEN_PERSPECTIVE_START = 27  # 「■観点の目次」ヘッダー行


def _generate_decision_table(test_cases, spec_text, environments):
    """テストケースの構造から意味のあるデシジョンテーブルを生成する。

    条件はテストケースのconfirm/targetから分岐要因を抽出する:
    - 環境（TH/WDL等）
    - フロー（機能名）
    - テスト種別（正常系/異常系）
    - 入力状態（正常/未登録/未入力/期限切れ等）

    Returns:
        dict or None: DTセクション。テストケースが少なすぎる場合はNone。
    """
    if not test_cases or len(test_cases) < 3:
        return None

    # --- Step 1: テストケースからパターンをグループ化 ---
    # 「確認すること」のベース名でグループ化
    groups = {}
    for tc in test_cases:
        confirm = tc.get("confirm", "")
        # 異常系プレフィクスを除去してベース名取得
        is_abnormal = any(kw in confirm for kw in ["異常", "エラー"])
        base = re.sub(r"^(異常系[：:]?\s*)", "", confirm)
        # 環境を抽出
        env_match = re.search(r"[（(]([\w/]+?)(?:のみ)?[）)]", base)
        env = env_match.group(1) if env_match else ""
        # フロー名（環境部分を除去）
        flow = re.sub(r"[（(][\w/]+?(?:のみ)?[）)]", "", base).strip()

        key = (flow, env, is_abnormal)
        if key not in groups:
            groups[key] = []
        groups[key].append(tc)

    if len(groups) < 2:
        return None

    # --- Step 2: 条件を抽出 ---
    conditions = []
    condition_values_map = {}  # {pattern_idx: {cond_name: Y/N}}

    # 条件1: 環境
    all_envs = sorted(set(env for _, env, _ in groups.keys() if env))
    if len(all_envs) >= 2:
        for env in all_envs:
            conditions.append(f"対象環境が{env}")

    # 条件2: テスト種別
    has_normal = any(not abn for _, _, abn in groups.keys())
    has_abnormal = any(abn for _, _, abn in groups.keys())
    if has_normal and has_abnormal:
        conditions.append("正常系の操作")

    # 条件3: 入力状態（異常系テストケースのtargetから抽出）
    input_conditions = set()
    for tc in test_cases:
        confirm = tc.get("confirm", "")
        target = tc.get("target", "")
        if any(kw in confirm for kw in ["異常", "エラー"]):
            # targetから具体的な入力状態を抽出
            for pattern, label in [
                (r"未登録|存在しない", "未登録の入力値"),
                (r"未入力|空", "必須項目が未入力"),
                (r"期限切れ|有効期限", "URLまたはトークンが期限切れ"),
                (r"登録済み|重複", "登録済みデータの入力"),
                (r"ロック|ブロック", "アカウントがロック状態"),
            ]:
                if re.search(pattern, target):
                    input_conditions.add(label)

    for ic in sorted(input_conditions):
        conditions.append(ic)

    if not conditions:
        return None

    # --- Step 3: パターンを生成（グループごとに1パターン、最大7） ---
    pattern_keys = sorted(groups.keys(), key=lambda k: (k[2], k[0], k[1]))[:7]
    num_patterns = len(pattern_keys)

    # パターンラベル（簡潔に）
    pattern_labels = []
    for flow, env, is_abnormal in pattern_keys:
        label = flow
        if env:
            label += f"({env})"
        if is_abnormal:
            # 異常系は代表的なtargetを使う
            tcs = groups[(flow, env, is_abnormal)]
            targets = [tc.get("target", "") for tc in tcs]
            label = targets[0][:15] if targets else "異常系"
        pattern_labels.append(label)

    # --- Step 4: 各パターンの条件値を計算 ---
    dt_pattern_values = []
    for flow, env, is_abnormal in pattern_keys:
        col = {}
        tcs = groups[(flow, env, is_abnormal)]
        targets_text = " ".join(tc.get("target", "") for tc in tcs).lower()

        # 環境条件
        for e in all_envs:
            if len(all_envs) >= 2:
                col[f"cond_対象環境が{e}"] = "Y" if env == e else "N"

        # テスト種別条件
        if has_normal and has_abnormal:
            col["cond_正常系の操作"] = "N" if is_abnormal else "Y"

        # 入力状態条件
        for ic in sorted(input_conditions):
            matched = False
            for pattern, label in [
                (r"未登録|存在しない", "未登録の入力値"),
                (r"未入力|空", "必須項目が未入力"),
                (r"期限切れ|有効期限", "URLまたはトークンが期限切れ"),
                (r"登録済み|重複", "登録済みデータの入力"),
                (r"ロック|ブロック", "アカウントがロック状態"),
            ]:
                if label == ic and re.search(pattern, targets_text):
                    matched = True
                    break
            col[f"cond_{ic}"] = "Y" if matched else "N"

        dt_pattern_values.append(col)

    # --- Step 5: 動作を生成 ---
    # グループごとの代表的な期待結果を動作にする
    actions = []
    action_map = {}  # {action_text: set of pattern_indices}
    for p_idx, (flow, env, is_abnormal) in enumerate(pattern_keys):
        tcs = groups[(flow, env, is_abnormal)]
        # 代表的な期待値（最初のテストケース）
        expected = tcs[0].get("expected", "")
        # 長すぎる場合は省略
        display = expected if len(expected) <= 40 else expected[:37] + "..."
        if display and display not in actions:
            actions.append(display)
        if display:
            if display not in action_map:
                action_map[display] = set()
            action_map[display].add(p_idx)

    # 動作値を設定
    for p_idx in range(num_patterns):
        col = dt_pattern_values[p_idx]
        for act in actions:
            col[f"act_{act}"] = "○" if p_idx in action_map.get(act, set()) else ""

    # --- Step 6: 全パターンNの条件を除外 ---
    conditions_to_remove = []
    for cond in conditions:
        all_n = all(
            dt_pattern_values[p].get(f"cond_{cond}", "N") == "N"
            for p in range(num_patterns)
        )
        all_y = all(
            dt_pattern_values[p].get(f"cond_{cond}", "N") == "Y"
            for p in range(num_patterns)
        )
        if all_n or all_y:
            conditions_to_remove.append(cond)

    for cond in conditions_to_remove:
        conditions.remove(cond)

    if not conditions:
        return None

    return {
        "type": "decision_table",
        "title": "条件分岐テーブル",
        "conditions": conditions,
        "actions": actions,
        "num_patterns": num_patterns,
        "pattern_labels": pattern_labels,
        "pattern_values": dt_pattern_values,
    }


def generate_perspective_sections(spec_text, test_cases):
    """仕様書とテストケースから観点設計セクションを自動生成する。

    Returns:
        list[dict]: セクションのリスト。各セクションは形式に応じた構造。
    """
    sections = []

    # --- テストケースから構造を抽出 ---
    confirms = []       # ユニークな「確認すること」グループ
    screens = []        # ユニークな画面
    targets = []        # ユニークな確認対象
    normal_cases = []
    abnormal_cases = []

    seen_confirms = set()
    seen_screens = set()
    seen_targets = set()

    for tc in (test_cases or []):
        c = tc.get("confirm", "")
        s = tc.get("screen", "")
        t = tc.get("target", "")

        if c and c not in seen_confirms:
            confirms.append(c)
            seen_confirms.add(c)
        if s and s not in seen_screens:
            screens.append(s)
            seen_screens.add(s)
        if t and t not in seen_targets:
            targets.append(t)
            seen_targets.add(t)

        is_abnormal = any(kw in c for kw in ["異常", "エラー", "バリデーション", "不正"])
        if is_abnormal:
            abnormal_cases.append(tc)
        else:
            normal_cases.append(tc)

    # --- 仕様書から環境・フローを抽出 ---
    spec = spec_text or ""
    environments = []
    env_patterns = [
        ("TH", r"\bTH\b"),
        ("WDL", r"\bWDL\b"),
        ("TD", r"\bTD\b"),
        ("TK", r"\bTK\b"),
    ]
    for env_name, pattern in env_patterns:
        if re.search(pattern, spec):
            environments.append(env_name)
    if not environments:
        environments = ["全環境"]

    # --- セクション1: リスト形式（機能別テスト観点一覧） ---
    list_rows = []
    for tc in (test_cases or []):
        c = tc.get("confirm", "")
        s = tc.get("screen", "")
        t = tc.get("target", "")
        d = tc.get("detail", "")
        if c or t:
            list_rows.append({"confirm": c, "screen": s, "target": t, "detail": d})

    # 重複を除去しつつ順序保持（confirm+target の組み合わせでユニーク化）
    seen = set()
    unique_list_rows = []
    for row in list_rows:
        key = (row["confirm"], row["target"])
        if key not in seen:
            seen.add(key)
            unique_list_rows.append(row)

    sections.append({
        "type": "list",
        "title": "機能別テスト観点一覧",
        "headers": ["確認すること", "画面", "確認対象", "詳細"],
        "rows": unique_list_rows,
    })

    # --- セクション2: テスト種別マトリクス（正常系/異常系/境界値） ---
    # 確認すること × テスト種別 の組合せ
    # 正常系/異常系/境界値のどれがカバーされているかを可視化
    base_confirms = []
    seen_base = set()
    for c in confirms:
        # 「異常系：」プレフィクスを除去してベース名を取得
        base = re.sub(r"^(異常系[：:]?\s*)", "", c)
        if base not in seen_base:
            seen_base.add(base)
            base_confirms.append(base)

    type_matrix_rows = []
    for base in base_confirms:
        has_normal = any(
            not any(kw in tc.get("confirm", "") for kw in ["異常", "エラー"])
            and base in tc.get("confirm", "")
            for tc in (test_cases or [])
        )
        has_abnormal = any(
            any(kw in tc.get("confirm", "") for kw in ["異常", "エラー"])
            and base in tc.get("confirm", "")
            for tc in (test_cases or [])
        )
        has_boundary = any(
            any(kw in tc.get("confirm", "") + tc.get("target", "") for kw in ["境界", "上限", "下限", "最大", "最小"])
            and base in tc.get("confirm", "")
            for tc in (test_cases or [])
        )
        type_matrix_rows.append({
            "target": base,
            "values": {
                "正常系": has_normal,
                "異常系": has_abnormal,
                "境界値": has_boundary,
            }
        })

    sections.append({
        "type": "matrix",
        "title": "テスト種別カバレッジ",
        "col_headers": ["正常系", "異常系", "境界値"],
        "rows": type_matrix_rows,
    })

    # --- セクション3: 環境×機能マトリクス（組合せ形式） ---
    if len(environments) > 1:
        env_matrix_rows = []
        for base in base_confirms:
            values = {}
            # 「（THのみ）」「（WDLのみ）」等の制限を検出
            only_match = re.search(r"[（(]([\w/]+)のみ[）)]", base)
            only_env = only_match.group(1) if only_match else None

            for env in environments:
                if only_env:
                    # 「〜のみ」の場合、その環境のみ○
                    has_env = env in only_env
                else:
                    # テストケースの confirm に環境名が含まれているかで判定
                    has_env = any(
                        env in tc.get("confirm", "")
                        for tc in (test_cases or [])
                        if base in tc.get("confirm", "")
                    )
                    # 仕様書に環境×機能の対応記述があるか
                    if not has_env:
                        has_env = env.lower() in spec.lower() and base.lower() in spec.lower()
                values[env] = has_env
            env_matrix_rows.append({"target": base, "values": values})

        sections.append({
            "type": "matrix",
            "title": "環境別テスト対象",
            "col_headers": environments,
            "rows": env_matrix_rows,
        })

    # --- セクション4: デシジョンテーブル ---
    # テストケースのconfirm/targetから分岐要因を自動抽出し、
    # 意味のある条件×パターンのDTを生成する。
    dt_section = _generate_decision_table(test_cases, spec, environments)
    if dt_section:
        sections.append(dt_section)

    # --- セクション5: リスク観点チェックリスト ---
    # 各観点に対して「仕様の主題に直結するキーワード(primary)」で判定。
    # 汎用的なキーワード（「保存」「入力」「エラー」「認証」等）は
    # 大半の仕様に出現するため判定に使わない。
    risk_items = []
    risk_perspectives = [
        ("データ整合性", "入力→保存→再表示で値が保持されるか",
         ["データ保存", "データ更新", "データベース", "整合", "不整合",
          "保持される", "反映される", "再表示"]),
        ("セッション管理", "操作中のタイムアウト・セッション切れ時の挙動",
         ["セッション", "タイムアウト", "有効期限", "期限切れ",
          "ログアウト", "session", "expire"]),
        ("並行操作", "複数ユーザー/タブから同時操作した場合の競合",
         ["並行", "同時操作", "排他制御", "排他ロック", "競合",
          "デッドロック", "race condition", "concurrent"]),
        ("入力バリデーション", "必須チェック・形式チェック・文字数制限",
         ["バリデーション", "入力チェック", "必須チェック", "形式チェック",
          "文字数制限", "桁数", "validation"]),
        ("エラーハンドリング", "サーバーエラー・ネットワーク断時のUI表示",
         ["エラーハンドリング", "エラー処理", "例外処理", "サーバーエラー",
          "ネットワーク断", "500エラー", "タイムアウトエラー"]),
        ("権限制御", "操作権限のないユーザーでのアクセス",
         ["権限", "アクセス制御", "ロール", "管理者権限",
          "操作権限", "認可", "authorization", "permission"]),
        ("セキュリティ", "脆弱性対策・不正アクセス防止",
         ["脆弱性", "セキュリティ", "xss", "csrf", "インジェクション",
          "リダイレクト", "改ざん", "不正アクセス", "セキュリティ診断",
          "devsec", "オープンリダイレクタ"]),
        ("ブラウザ互換", "対象ブラウザでの表示・動作差異",
         ["ブラウザ互換", "クロスブラウザ", "ブラウザ差異",
          "chrome", "edge", "firefox", "safari", "ie"]),
        ("パフォーマンス", "大量データ時の応答速度・表示崩れ",
         ["パフォーマンス", "レスポンス時間", "負荷", "大量データ",
          "スロークエリ", "速度改善", "performance", "latency"]),
    ]

    # 仕様書のみを対象に判定（テストケースは仕様の派生なので含めない）
    spec_lower = spec.lower()
    for name, description, keywords in risk_perspectives:
        relevant = any(kw in spec_lower for kw in keywords)
        risk_items.append({
            "perspective": name,
            "description": description,
            "relevant": relevant,
        })

    sections.append({
        "type": "risk_checklist",
        "title": "リスク観点チェックリスト",
        "items": risk_items,
    })

    return sections


def _fill_perspective_sections(ws, sections):
    """観点シートの目次セクション（Row 27以降）に設計内容を書き込む。"""
    from openpyxl.utils import get_column_letter

    # スタイル定義
    section_title_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=11)
    header_font = Font(bold=True, size=10)
    data_font = Font(size=10)
    toc_font = Font(size=10, color="1F4E79", underline="single")

    # セクション別カラーテーマ（ヘッダー背景 / タイトルバー背景）
    list_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    list_title_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    matrix_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    matrix_title_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    dt_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    dt_title_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    risk_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    risk_title_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    check_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    toc_bg = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    thick_bottom = Border(bottom=Side(style="medium"))
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 色テーママッピング
    theme_map = {
        "list": (list_title_fill, list_fill),
        "matrix": (matrix_title_fill, matrix_fill),
        "decision_table": (dt_title_fill, dt_fill),
        "risk_checklist": (risk_title_fill, risk_fill),
    }

    # ヘルパー: セル結合 + 罫線 + 値セット
    def merged_cell(row, col_start, col_end, value, font=None, fill=None, alignment=None, border=None):
        """結合セルに値・書式を設定する。"""
        if col_start < col_end:
            ws.merge_cells(
                start_row=row, start_column=col_start,
                end_row=row, end_column=col_end
            )
        cell = ws.cell(row=row, column=col_start, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
            # 結合範囲の右端にも罫線
            if col_start < col_end:
                for c in range(col_start, col_end + 1):
                    ws.cell(row=row, column=c).border = border
        return cell

    def set_row_height(row, height):
        ws.row_dimensions[row].height = height

    # --- テンプレート既存データのクリア（Row 27〜120） ---
    for row in range(KANTEN_PERSPECTIVE_START, 121):
        for col in range(1, 20):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.number_format = "General"
            cell.font = Font(size=10)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment()
    # 結合セルも解除
    merges_to_remove = [
        m for m in ws.merged_cells.ranges
        if m.min_row >= KANTEN_PERSPECTIVE_START and m.min_row <= 120
    ]
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))

    # --- 列幅の調整（観点セクション用） ---
    col_widths = {
        "B": 30, "C": 14, "D": 14, "E": 14, "F": 14,
        "G": 14, "H": 14, "I": 14, "J": 14, "K": 14,
        "L": 18, "M": 14, "N": 14, "O": 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- 目次ヘッダー ---
    current_row = KANTEN_PERSPECTIVE_START
    merged_cell(current_row, 2, 15, "■ 観点の目次",
                font=Font(bold=True, size=13), alignment=Alignment(vertical="center"))
    set_row_height(current_row, 28)
    current_row += 1

    # 目次リンク用の行を確保（後で埋める）
    toc_start_row = current_row
    current_row += len(sections) + 1  # 目次分 + 空行

    section_start_rows = []

    for sec_idx, section in enumerate(sections):
        sec_type = section["type"]
        title = section.get("title", "")
        section_start_rows.append(current_row)
        title_fill, header_fill = theme_map.get(sec_type, (list_title_fill, list_fill))

        sec_type_label = {
            "list": "リスト形式",
            "matrix": "組合せ形式",
            "decision_table": "デシジョンテーブル形式",
            "risk_checklist": "チェックリスト形式",
        }.get(sec_type, "")

        # ===== セクションタイトルバー（結合+色付き帯） =====
        merged_cell(current_row, 2, 15,
                    f"  目次{sec_idx + 1}  |  {sec_type_label}  |  {title}",
                    font=section_title_font, fill=title_fill,
                    alignment=Alignment(vertical="center"))
        set_row_height(current_row, 26)
        current_row += 1

        if sec_type == "list":
            # ========= リスト形式 =========
            # ヘッダー（結合で列幅を確保）
            headers = [
                (2, 4, "確認すること"),
                (5, 7, "画面"),
                (8, 11, "確認対象"),
                (12, 15, "詳細"),
            ]
            for cs, ce, hdr in headers:
                merged_cell(current_row, cs, ce, hdr,
                            font=header_font, fill=header_fill,
                            border=thin_border, alignment=center)
            set_row_height(current_row, 22)
            current_row += 1

            # データ行
            for row_data in section.get("rows", []):
                data = [
                    (2, 4, row_data.get("confirm", "")),
                    (5, 7, row_data.get("screen", "")),
                    (8, 11, row_data.get("target", "")),
                    (12, 15, row_data.get("detail", "")),
                ]
                for cs, ce, val in data:
                    merged_cell(current_row, cs, ce, val,
                                font=data_font, border=thin_border,
                                alignment=wrap_top)
                set_row_height(current_row, 30)
                current_row += 1

            current_row += 1  # 空行

        elif sec_type == "matrix":
            # ========= 組合せ形式（マトリクス） =========
            col_headers = section.get("col_headers", [])
            rows = section.get("rows", [])

            # ヘッダー行: 確認対象(結合B-E) | col1 | col2 | ...
            merged_cell(current_row, 2, 5, "確認対象",
                        font=header_font, fill=header_fill,
                        border=thin_border, alignment=center)
            for j, hdr in enumerate(col_headers):
                col = 6 + j
                cell = ws.cell(row=current_row, column=col, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = wrap_center
            set_row_height(current_row, 22)
            current_row += 1

            # データ行
            for row_data in rows:
                merged_cell(current_row, 2, 5, row_data["target"],
                            font=data_font, border=thin_border,
                            alignment=wrap_top)
                values = row_data.get("values", {})
                for j, hdr in enumerate(col_headers):
                    col = 6 + j
                    is_checked = values.get(hdr, False)
                    cell = ws.cell(row=current_row, column=col,
                                   value="○" if is_checked else "×")
                    cell.border = thin_border
                    cell.alignment = center
                    cell.font = Font(size=12 if is_checked else 10,
                                     bold=is_checked,
                                     color="2E75B6" if is_checked else "AAAAAA")
                    if is_checked:
                        cell.fill = check_fill
                set_row_height(current_row, 24)
                current_row += 1

            current_row += 1  # 空行

        elif sec_type == "decision_table":
            # ========= デシジョンテーブル形式 =========
            conditions = section.get("conditions", [])
            actions = section.get("actions", [])
            num_patterns = section.get("num_patterns", 0)
            pattern_values = section.get("pattern_values", [])
            # DT列配置: B-E=条件/動作ラベル, F〜=パターン列
            dt_data_start = 6

            # パターンヘッダー（ラベルがあれば使用）
            pattern_labels = section.get("pattern_labels", [])
            merged_cell(current_row, 2, 5, "",
                        fill=dt_fill, border=thin_border)
            for p in range(num_patterns):
                col = dt_data_start + p
                label = pattern_labels[p] if p < len(pattern_labels) else f"パターン{p + 1}"
                cell = ws.cell(row=current_row, column=col, value=label)
                cell.font = header_font
                cell.fill = dt_fill
                cell.border = thin_border
                cell.alignment = wrap_center
            set_row_height(current_row, 30)
            current_row += 1

            # 条件セクションヘッダー
            merged_cell(current_row, 2, 5, "【条件】",
                        font=Font(bold=True, size=10, color="FFFFFF"),
                        fill=dt_title_fill, border=thin_border,
                        alignment=Alignment(vertical="center"))
            for p in range(num_patterns):
                cell = ws.cell(row=current_row, column=dt_data_start + p)
                cell.fill = dt_title_fill
                cell.border = thin_border
            set_row_height(current_row, 22)
            current_row += 1

            for cond in conditions:
                merged_cell(current_row, 2, 5, cond,
                            font=data_font, border=thin_border,
                            alignment=wrap_top)
                for p in range(num_patterns):
                    col = dt_data_start + p
                    val = pattern_values[p].get(f"cond_{cond}", "-") if p < len(pattern_values) else "-"
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = center
                    cell.font = Font(size=10, bold=(val == "Y"),
                                     color="2E75B6" if val == "Y" else "808080")
                set_row_height(current_row, 24)
                current_row += 1

            # 動作セクションヘッダー
            merged_cell(current_row, 2, 5, "【動作】",
                        font=Font(bold=True, size=10, color="FFFFFF"),
                        fill=dt_title_fill, border=thin_border,
                        alignment=Alignment(vertical="center"))
            for p in range(num_patterns):
                cell = ws.cell(row=current_row, column=dt_data_start + p)
                cell.fill = dt_title_fill
                cell.border = thin_border
            set_row_height(current_row, 22)
            current_row += 1

            for act in actions:
                merged_cell(current_row, 2, 5, act,
                            font=data_font, border=thin_border,
                            alignment=wrap_top)
                for p in range(num_patterns):
                    col = dt_data_start + p
                    val = pattern_values[p].get(f"act_{act}", "") if p < len(pattern_values) else ""
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = center
                    if val == "○":
                        cell.font = Font(size=12, bold=True, color="2E75B6")
                        cell.fill = check_fill
                    else:
                        cell.font = Font(size=10, color="808080")
                set_row_height(current_row, 30)
                current_row += 1

            current_row += 1  # 空行

        elif sec_type == "risk_checklist":
            # ========= リスク観点チェックリスト =========
            # ヘッダー（結合で列幅確保）
            headers = [
                (2, 4, "観点"),
                (5, 11, "説明"),
                (12, 12, "該当"),
                (13, 15, "対応状況"),
            ]
            for cs, ce, hdr in headers:
                merged_cell(current_row, cs, ce, hdr,
                            font=header_font, fill=risk_fill,
                            border=thin_border, alignment=center)
            set_row_height(current_row, 22)
            current_row += 1

            for item in section.get("items", []):
                merged_cell(current_row, 2, 4, item["perspective"],
                            font=Font(bold=True, size=10), border=thin_border,
                            alignment=wrap_top)
                merged_cell(current_row, 5, 11, item["description"],
                            font=data_font, border=thin_border,
                            alignment=wrap_top)
                relevant_val = "○" if item["relevant"] else "-"
                cell = ws.cell(row=current_row, column=12, value=relevant_val)
                cell.border = thin_border
                cell.alignment = center
                cell.font = Font(size=12, bold=True, color="2E75B6") if item["relevant"] else Font(size=10, color="AAAAAA")
                if item["relevant"]:
                    cell.fill = check_fill
                merged_cell(current_row, 13, 15, "",
                            border=thin_border, alignment=wrap_top)
                set_row_height(current_row, 36)
                current_row += 1

            current_row += 1  # 空行

    # --- 目次リンクを埋める ---
    for i, section in enumerate(sections):
        row = toc_start_row + i
        title = section.get("title", "")
        sec_type_label = {
            "list": "リスト形式",
            "matrix": "組合せ形式",
            "decision_table": "デシジョンテーブル形式",
            "risk_checklist": "チェックリスト形式",
        }.get(section["type"], "")
        merged_cell(row, 2, 15,
                    f"  {i + 1}.  {title}（{sec_type_label}）  ──  行 {section_start_rows[i]}",
                    font=toc_font, fill=toc_bg,
                    alignment=Alignment(vertical="center"))
        set_row_height(row, 20)


# =====================================================================
# Excel出力（テンプレート準拠）
# =====================================================================

def write_excel(test_cases, output_path, spec_text=None, ticket_keys=None):
    """テストケースをテンプレート準拠のExcelに出力する"""

    # テンプレートの存在確認
    if TEMPLATE_PATH.exists():
        # テンプレートをコピーして使用
        shutil.copy2(TEMPLATE_PATH, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb["テストケース"]
        use_template = True
        print(f"  テンプレート使用: {TEMPLATE_PATH.name}")
    else:
        # テンプレートがない場合はスクラッチで作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "テストケース"
        use_template = False
        _create_testcase_sheet_from_scratch(ws)
        print("  テンプレートなし: スクラッチ作成")

    # Row 3（テンプレートのサンプル行）をクリアしてデータ投入
    # テンプレートの Row 3 はサンプル行なので上書き
    wrap = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, tc in enumerate(test_cases):
        row = TC_DATA_START_ROW + i

        # No列: 数値（data_only=True で読み込まれても値が残るように）
        ws.cell(row=row, column=TC_COL["no"], value=i + 1)

        # テストケースデータ
        field_map = {
            "confirm": tc.get("confirm", ""),
            "screen": tc.get("screen", ""),
            "target": tc.get("target", ""),
            "detail": tc.get("detail", ""),
            "steps": tc.get("steps", ""),
            "expected": tc.get("expected", ""),
            "notes": tc.get("notes", ""),
        }

        for field, value in field_map.items():
            cell = ws.cell(row=row, column=TC_COL[field], value=value)
            cell.alignment = wrap
            cell.border = thin_border

        # 実行列 K-Q に「未」をセット
        for exec_col in ["exec_k", "exec_l", "exec_m", "exec_n", "exec_o", "exec_p", "exec_q"]:
            cell = ws.cell(row=row, column=TC_COL[exec_col], value="未")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # No列のスタイル
        no_cell = ws.cell(row=row, column=TC_COL["no"])
        no_cell.alignment = Alignment(horizontal="center", vertical="center")
        no_cell.border = thin_border

        # 残りの列（Autify, 実行者, 実行備考, JIRA No）にボーダーのみ
        for extra_col in ["autify", "executor", "exec_notes", "jira_no"]:
            cell = ws.cell(row=row, column=TC_COL[extra_col])
            cell.border = thin_border

    # --- 観点シートの自動入力 ---
    perspectives = analyze_perspectives(spec_text, test_cases)

    if "観点" in wb.sheetnames:
        ws_kanten = wb["観点"]
        _fill_kanten_sheet(ws_kanten, perspectives, ticket_keys)

        # 観点設計セクション（リスト/組合せ/デシジョンテーブル/チェックリスト）
        sections = generate_perspective_sections(spec_text, test_cases)
        _fill_perspective_sections(ws_kanten, sections)

    wb.save(output_path)
    return len(test_cases)


def _fill_kanten_sheet(ws, perspectives, ticket_keys=None):
    """観点シートに自動判定結果を入力する"""

    # チケットURL等の情報
    if ticket_keys:
        base_url = "https://beartail.atlassian.net/browse"
        ticket_url = ", ".join(f"{base_url}/{k}" for k in ticket_keys)
        ws.cell(row=4, column=9, value=ticket_url)  # I4: 開発チケット

    if perspectives.get("ticket_url"):
        if not ticket_keys:
            ws.cell(row=4, column=9, value=perspectives["ticket_url"])

    # テンプレートの既存○マークをクリア（KANTEN_CELLSに定義のない列も含む）
    MARK_COLUMNS = [3, 6, 9, 12, 14, 15]  # C, F, I, L, N, O
    for row in range(11, 26):
        for col in MARK_COLUMNS:
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip() in ('○', '◯', '〇', 'O'):
                ws.cell(row=row, column=col, value="-")

    # 各カテゴリの選択を書き込む
    for category, items in KANTEN_CELLS.items():
        selections = perspectives.get(category, {})
        for item_name, (col_letter, row_num) in items.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            is_selected = selections.get(item_name, False)
            ws.cell(row=row_num, column=col_idx, value="〇" if is_selected else "-")


def _create_testcase_sheet_from_scratch(ws):
    """テンプレートがない場合にテストケースシートをスクラッチ作成"""
    header_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    env_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1: 環境ヘッダー
    env_headers = {
        11: "Win11/\nChrome",
        13: "Win11/\nEdge",
        14: "Win11/\nFirefox",
        15: "Mac/\nSafari",
        16: "iOS/\nOSver●●",
        17: "Android/\nOSver●●",
    }
    for col, label in env_headers.items():
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = env_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # Row 2: 列ヘッダー
    headers = {
        2: "No", 3: "確認すること", 4: "画面", 5: "確認対象",
        6: "詳細", 7: "テスト実行手順", 8: "期待値", 9: "備考",
        10: "Autify", 11: "TK+TI+TD", 12: "TK+TD", 13: "TI+TD",
        14: "TK単体", 15: "TD単体", 16: "TK+TI+TD", 17: "TK+TI+TD",
        18: "実行者", 19: "実行備考", 20: "起票JIRA No",
    }

    for col, label in headers.items():
        cell = ws.cell(row=2, column=col, value=label)
        if col <= 9:
            cell.fill = header_fill
        elif col <= 17:
            cell.fill = env_fill
        else:
            cell.fill = env_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # 列幅
    widths = {"A": 0.75, "B": 7.63, "C": 22.38, "D": 13, "E": 23.5,
              "F": 22.38, "G": 19.88, "H": 26, "I": 7.88, "J": 6.38,
              "K": 11.13, "L": 13, "M": 13, "N": 13, "O": 13,
              "P": 13, "Q": 13, "R": 8.75, "S": 10.88}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "C3"


# =====================================================================
# 評価
# =====================================================================

def run_evaluation(excel_path):
    """evaluate_testcases.py でスコアを取得する"""
    eval_script = SCRIPT_DIR / "evaluate_testcases.py"
    benchmark = SCRIPT_DIR / "benchmark.json"

    if not eval_script.exists():
        return {"score": -1, "feedback": "evaluate_testcases.py が見つかりません"}

    cmd = [sys.executable, str(eval_script), str(excel_path),
           "--benchmark", str(benchmark)]

    result = subprocess.run(
        cmd, capture_output=True, text=True,
        encoding="utf-8", errors="replace", cwd=str(SCRIPT_DIR),
    )

    output = ((result.stdout or "") + "\n" + (result.stderr or "")).strip()
    if not output:
        return {"score": 0, "feedback": "評価スクリプトの出力が空です"}

    score_match = re.search(r"総合スコア[：:]\s*(\d+)", output)
    score = int(score_match.group(1)) if score_match else 0
    return {"score": score, "feedback": output[:3000]}


# =====================================================================
# メイン
# =====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="テスト項目書 自動生成ツール（JIRA連携 + Excel変換 + 評価）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
例:
  python generate_testcases.py --jira TH-1234          # JIRAチケット取得→仕様書出力
  python generate_testcases.py --jira TH-1234 TH-1235  # 複数チケット取得
  python generate_testcases.py testcases.json           # JSON→Excel変換+評価
  python generate_testcases.py testcases.json --spec spec.md  # 観点自動判定付き
  python generate_testcases.py testcases.json -o out.xlsx --no-eval
        """,
    )

    # JIRA取得モード
    parser.add_argument(
        "--jira", nargs="+", metavar="TICKET_KEY",
        help="JIRAチケットキー（例: TH-1234）。複数指定可",
    )
    parser.add_argument(
        "--spec-output", metavar="PATH",
        help="JIRA取得時の仕様書出力先（デフォルト: spec_<チケットキー>.md）",
    )

    # Excel変換モード
    parser.add_argument(
        "input", nargs="?",
        help="テストケースJSONファイル（- で標準入力）。--jira と排他",
    )
    parser.add_argument("-o", "--output", help="出力Excelパス")
    parser.add_argument("--no-eval", action="store_true", help="評価をスキップ")
    parser.add_argument(
        "--spec", metavar="PATH",
        help="仕様書ファイル（観点の自動判定に使用）",
    )
    parser.add_argument(
        "--ticket-keys", nargs="*", metavar="KEY",
        help="関連JIRAチケットキー（観点シートのチケットURLに使用）",
    )

    args = parser.parse_args()

    # === JIRAモード ===
    if args.jira:
        print(f"JIRA取得モード: {', '.join(args.jira)}")
        spec_md = fetch_and_format_tickets(args.jira)

        if not spec_md:
            print("ERROR: JIRAチケットを取得できませんでした")
            sys.exit(1)

        # 仕様書ファイルとして保存
        if args.spec_output:
            spec_path = Path(args.spec_output)
        else:
            spec_path = Path(f"spec_{'_'.join(args.jira)}.md")

        with open(spec_path, "w", encoding="utf-8") as f:
            f.write(spec_md)

        print(f"\n仕様書保存: {spec_path} ({len(spec_md)} 文字)")
        print(f"\n次のステップ:")
        print(f"  Claude Codeで「{spec_path} からテスト項目書を作って」と指示してください")
        return 0

    # === Excel変換モード ===
    if not args.input:
        parser.print_help()
        sys.exit(1)

    # JSON読み込み
    if args.input == "-":
        raw = sys.stdin.read()
    else:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"ERROR: {input_path} が見つかりません")
            sys.exit(1)
        with open(input_path, encoding="utf-8") as f:
            raw = f.read()

    test_cases = json.loads(raw)
    print(f"テストケース: {len(test_cases)} 件")

    # 仕様書テキスト読み込み（観点自動判定用）
    spec_text = None
    if args.spec:
        spec_path = Path(args.spec)
        if spec_path.exists():
            with open(spec_path, encoding="utf-8") as f:
                spec_text = f.read()
            print(f"  仕様書読み込み: {spec_path}")

    # 出力先
    if args.output:
        output_path = Path(args.output)
    elif args.input != "-":
        output_path = Path(args.input).with_suffix(".xlsx")
    else:
        output_path = Path("testcases_output.xlsx")

    # Excel出力
    count = write_excel(test_cases, output_path, spec_text=spec_text,
                        ticket_keys=args.ticket_keys or args.jira)
    print(f"Excel保存: {output_path} ({count} 件)")

    # 評価
    if not args.no_eval:
        print("\n評価中...")
        eval_result = run_evaluation(output_path)
        score = eval_result.get("score", eval_result.get("total_score", 0))
        print(f"スコア: {score}/100")

        feedback = eval_result.get("feedback", "")
        if feedback:
            print(f"\nフィードバック:\n{feedback[:2000]}")

        # JSON結果も保存
        result_path = output_path.with_suffix(".eval.json")
        with open(result_path, "w", encoding="utf-8") as f:
            json.dump(eval_result, f, ensure_ascii=False, indent=2)
        print(f"評価結果: {result_path}")

        return score
    return -1


if __name__ == "__main__":
    score = main()
    sys.exit(0 if (isinstance(score, int) and (score >= 80 or score <= 0)) else 1)
