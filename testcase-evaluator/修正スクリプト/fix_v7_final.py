#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""v7最終修正:
1. テストケース: リカバリー→パスワード再設定 (3件)
2. テストケース: 期待値の複数「こと」を1つに統合 (7件)
3. 観点シート: ID/PW→メールアドレス・パスワード, PW→パスワード, リカバリー→パスワード再設定
"""
import os, sys, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'memo', 'ログイン情報', '修正対象', '最新版')
v7_path = os.path.join(path, 'テスト項目書_改善後_v7.xlsx')
wb = openpyxl.load_workbook(v7_path)

# =====================================================================
# 修正1: テストケースのリカバリー→パスワード再設定
# =====================================================================
print("=" * 70)
print("【修正1】テストケース: リカバリー→パスワード再設定")
print("=" * 70)

ws_tc = wb['テストケース_改善後']
fix1_count = 0
for row in range(3, ws_tc.max_row + 1):
    no = ws_tc.cell(row=row, column=2).value
    if no is None:
        continue
    for col in range(3, 10):
        cell = ws_tc.cell(row=row, column=col)
        val = cell.value
        if val and isinstance(val, str) and 'リカバリー' in val:
            original = val
            val = val.replace('パスワードリカバリー', 'パスワード再設定')
            val = val.replace('リカバリーフロー', 'パスワード再設定フロー')
            val = val.replace('リカバリー後', 'パスワード再設定後')
            val = val.replace('リカバリーまたは', 'パスワード再設定または')
            val = val.replace('リカバリー', 'パスワード再設定')
            if val != original:
                cell.value = val
                fix1_count += 1
                print(f"  No.{no} Col{col}: {repr(original[:60])} → {repr(val[:60])}")

print(f"\nリカバリー修正: {fix1_count}件")

# =====================================================================
# 修正2: テストケースの期待値 複数「こと」→1つに統合
# =====================================================================
print()
print("=" * 70)
print("【修正2】テストケース: 期待値の複数「こと」を統合")
print("=" * 70)

fix2_count = 0
for row in range(3, ws_tc.max_row + 1):
    no = ws_tc.cell(row=row, column=2).value
    if no is None:
        continue
    cell = ws_tc.cell(row=row, column=8)
    exp = cell.value
    if not exp:
        continue

    new_exp = exp

    # No.77: 重複する「こと」を統合
    if no == 77:
        new_exp = '表示文言が以下と完全一致すること：\n「この事業所ではTOKIUM ID でのログインのみ許可されています。TOKIUM ID でログインしてください。」'

    # No.136: 括弧内の方が具体的なのでそちらを採用
    if no == 136:
        new_exp = '元のパスワードでTHにログインできること'

    # No.149: 括弧内は別ケース(No.150)で確認済みなので削除
    if no == 149:
        new_exp = 'エラーメッセージが表示されること'

    # No.153: 括弧内を本文に組み込み
    if no == 153:
        new_exp = 'THログイン画面（メールアドレス欄・パスワード欄・ログインボタン）が表示されること'

    # No.176: 1文に統合
    if no == 176:
        new_exp = '5回連続でエラーが発生せずにログイン・ログアウトが完了すること'

    # No.191: 括弧内が本質なので統合
    if no == 191:
        new_exp = '認証エラーではなくSSO限定エラーが優先表示されること'

    # No.192: SSO限定エラー文言を前に出して1文に
    if no == 192:
        new_exp = 'IP制限エラーではなくSSO限定エラー「この事業所ではTOKIUM ID でのログインのみ許可されています。TOKIUM ID でログインしてください。」が表示されること'

    if new_exp != exp:
        cell.value = new_exp
        fix2_count += 1
        print(f"  No.{no}: {repr(exp[:50])} → {repr(new_exp[:50])}")

print(f"\n期待値統合: {fix2_count}件")

# =====================================================================
# 修正3: 観点シートの用語統一
# =====================================================================
print()
print("=" * 70)
print("【修正3】観点シート: 用語統一")
print("=" * 70)

ws_k = wb['観点']

# DT条件行（Row 39, 51）は短縮ラベルなので「メール・PW」→「メール+PW」に
# それ以外の文中は「メールアドレス・パスワード」に統一

fix3_count = 0
dt_condition_rows = {39, 51}  # DTの条件行（ログイン方式）

for row in range(1, ws_k.max_row + 1):
    for col in range(1, ws_k.max_column + 1):
        cell = ws_k.cell(row=row, column=col)
        val = cell.value
        if not val or not isinstance(val, str):
            continue

        original = val

        # リカバリー → パスワード再設定
        if 'リカバリー' in val:
            val = val.replace('リカバリーフロー正常動作', 'パスワード再設定フローが正常動作すること')
            val = val.replace('リカバリー', 'パスワード再設定')

        # DT条件行のID/PW → メール+PW（セル幅の制約）
        if row in dt_condition_rows:
            val = val.replace('ID/PW', 'メール+PW')
        else:
            # 文中のID/PW → メールアドレス・パスワード
            # ただし短い文脈では略称を使う
            val = val.replace('ID/PWログイン', 'メールアドレス・パスワードでのログイン')
            val = val.replace('ID/PWは拒否', 'メールアドレス・パスワードは拒否')
            val = val.replace('ID/PW許可復活', 'メールアドレス・パスワードでのログイン許可復活')
            val = val.replace('ID/PW拒否', 'メールアドレス・パスワードでのログイン拒否')
            val = val.replace('ID/PWセッション', 'メールアドレス・パスワードセッション')
            val = val.replace('ID/PWフォーム', 'メールアドレス・パスワードフォーム')
            val = val.replace('ID/PWもTOKIUM ID', 'メールアドレス・パスワードもTOKIUM ID')
            val = val.replace('ID/PWとTOKIUM ID', 'メールアドレス・パスワードとTOKIUM ID')
            val = val.replace('でもID/PW', 'でもメールアドレス・パスワード')
            val = val.replace('ID/PWで', 'メールアドレス・パスワードで')
            val = val.replace('正しいID/PW', '正しいメールアドレス・パスワード')
            val = val.replace('ID/PW', 'メールアドレス・パスワード')

        # PW → パスワード（ID/PW処理後の残存PW、パスPWは除外）
        # 「PW」単体（パスワードの略）を置換
        if re.search(r'(?<!パス)PW', val):
            # 個別パターン
            val = val.replace('誤PW', '誤パスワード')
            val = val.replace('誤ったPW', '誤ったパスワード')
            val = val.replace('PW連続誤入力', 'パスワード連続誤入力')
            val = val.replace('PW誤入力', 'パスワード誤入力')
            val = val.replace('PWがマスク表示', 'パスワードがマスク表示')
            val = val.replace('PWマネージャー', 'パスワードマネージャー')
            val = val.replace('PW再設定', 'パスワード再設定')
            val = val.replace('PW変更後', 'パスワード変更後')
            val = val.replace('PW入力', 'パスワード入力')
            val = val.replace('PWを忘れた場合', 'パスワードを忘れた場合')
            val = val.replace('PWを再設定', 'パスワードを再設定')
            # 残存PW
            val = re.sub(r'(?<!パス)PW', 'パスワード', val)

        if val != original:
            cell.value = val
            fix3_count += 1
            if fix3_count <= 30:
                print(f"  Row {row}, Col {col}: {repr(original[:50])} → {repr(val[:50])}")

print(f"\n観点シート用語統一: {fix3_count}件")

# =====================================================================
# 保存
# =====================================================================
wb.save(v7_path)
print(f"\n保存完了: {v7_path}")

# =====================================================================
# 最終検証
# =====================================================================
print()
print("=" * 70)
print("【最終検証】")
print("=" * 70)

# テストケース検証
wb2 = openpyxl.load_workbook(v7_path, data_only=True)
ws2 = wb2['テストケース_改善後']

cases = []
for row in range(3, ws2.max_row + 1):
    no = ws2.cell(row=row, column=2).value
    if no is None:
        continue
    cases.append({
        'no': no,
        'confirm': ws2.cell(row=row, column=3).value or '',
        'target': ws2.cell(row=row, column=5).value or '',
        'detail': ws2.cell(row=row, column=6).value or '',
        'steps': ws2.cell(row=row, column=7).value or '',
        'expected': ws2.cell(row=row, column=8).value or '',
        'notes': ws2.cell(row=row, column=9).value or '',
    })

# リカバリー残存
tc_recovery = 0
for c in cases:
    for f in ['confirm','target','detail','steps','expected','notes']:
        if 'リカバリー' in (c.get(f, '') or ''):
            tc_recovery += 1
            print(f"  TC リカバリー残存: No.{c['no']} {f}")
print(f"TC リカバリー残存: {tc_recovery}件 {'✓' if tc_recovery == 0 else '✗'}")

# ID/PW残存
tc_idpw = 0
for c in cases:
    for f in ['confirm','target','detail','steps','expected','notes']:
        if 'ID/PW' in (c.get(f, '') or ''):
            tc_idpw += 1
print(f"TC ID/PW残存: {tc_idpw}件 {'✓' if tc_idpw == 0 else '✗'}")

# PW残存
tc_pw = 0
for c in cases:
    for f in ['confirm','target','detail','steps','expected','notes']:
        val = c.get(f, '') or ''
        if val != '-' and re.search(r'(?<!パス)PW', val):
            tc_pw += 1
print(f"TC PW残存: {tc_pw}件 {'✓' if tc_pw == 0 else '✗'}")

# 複数こと
multi_koto = 0
for c in cases:
    exp = c['expected']
    if not exp or exp == '-':
        continue
    matches = re.findall(r'こと', exp)
    if len(matches) >= 2:
        multi_koto += 1
        print(f"  TC 複数こと: No.{c['no']}: {exp[:60]}")
print(f"TC 複数「こと」: {multi_koto}件 {'✓' if multi_koto == 0 else '✗'}")

# 観点シート検証
ws_k2 = wb2['観点']
k_idpw = 0
k_pw = 0
k_recovery = 0
for row in range(1, ws_k2.max_row + 1):
    for col in range(1, ws_k2.max_column + 1):
        v = ws_k2.cell(row=row, column=col).value
        if not v or not isinstance(v, str):
            continue
        if 'ID/PW' in v:
            k_idpw += 1
        if re.search(r'(?<!パス)PW', v):
            k_pw += 1
        if 'リカバリー' in v:
            k_recovery += 1

print(f"観点 ID/PW残存: {k_idpw}件 {'✓' if k_idpw == 0 else '✗'}")
print(f"観点 PW残存: {k_pw}件 {'✓' if k_pw == 0 else '✗'}")
print(f"観点 リカバリー残存: {k_recovery}件 {'✓' if k_recovery == 0 else '✗'}")

print(f"\n総ケース数: {len(cases)}件")
