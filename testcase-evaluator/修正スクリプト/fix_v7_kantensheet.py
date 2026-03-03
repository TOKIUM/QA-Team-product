#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""v7パッチ: 観点シートにNEW-1, NEW-4, NEW-5, NEW-6を追加
insert_rowsは挿入点より上の行に影響しないため、
下のセクションから順に挿入すればオフセット不要。
"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment

path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'memo', 'ログイン情報', '修正対象', '最新版')
v7_path = os.path.join(path, 'テスト項目書_改善後_v7.xlsx')
wb = openpyxl.load_workbook(v7_path)
ws = wb['観点']

base_font = Font(name='Arial', size=10)
top_align = Alignment(vertical='top', wrap_text=True)

def write_cell(row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = base_font
    cell.alignment = top_align

print("観点シートにNEW行を追加します")
print()

# 元の行配置（挿入前）:
#   Row 72: テナント切替の詳細行 → Row 73: 既存セッション
#   Row 95: エラーメッセージが同一 → Row 96: PW/メールアドレスセクション
#   Row 119: PW再設定メール送信の詳細行 → Row 120: 高頻度セクション
#   Row 127: ローディング表示の詳細行 → Row 128: 制限あり時のログイン画面表示

# =====================================================================
# Step 1: NEW-5 → UI/UXセクション（Row 128の前に1行挿入）
# Row 127 = ローディング表示の詳細行
# Row 128 = 制限あり時のログイン画面表示 ← この前に挿入
# =====================================================================
ws.insert_rows(128, 1)
write_cell(128, 4, 'パスワード入力フィールド')
write_cell(128, 6, 'ログイン画面')
write_cell(128, 7, 'プレースホルダ表示')
write_cell(128, 10, '・パスワード欄に「8文字以上のパスワードを入力」がプレースホルダとして表示されること')
print("  NEW-5: Row 128 に挿入（UI/UX → パスワード入力フィールド）")
# Row 128以下が+1シフト。Row 127以上は変化なし。

# =====================================================================
# Step 2: NEW-4 → 既存機能セクション（Row 120の前に2行挿入）
# Row 119 = PW再設定メール送信の詳細行（「TOKIUM IDでログインしてください」の事前案内の有無）
# Row 120 = 高頻度セクション ← この前に挿入
# ※ Row 120 < 128 なのでStep1の影響なし
# =====================================================================
ws.insert_rows(120, 2)
write_cell(120, 4, 'パスワード再設定画面の注意書き文言')
write_cell(120, 6, 'パスワード再設定画面')
write_cell(120, 7, '注意書き文言')
write_cell(120, 10, '・「「TOKIUM ID でログイン」経由でログインしたユーザーのパスワードはこちらから再設定できません。」が表示されること')
write_cell(121, 10, '・TOKIUM IDでログインするユーザーへの案内として正しい文言であること')
print("  NEW-4: Row 120-121 に挿入（既存機能 → パスワード再設定画面の注意書き）")
# Row 120以下が+2シフト。Row 119以上は変化なし。

# =====================================================================
# Step 3: NEW-6 → セキュリティセクション（Row 96の前に1行挿入）
# Row 95 = エラーメッセージが同一
# Row 96 = PW/メールアドレスセクション開始 ← この前に挿入
# ※ Row 96 < 120 なのでStep2の影響なし
# =====================================================================
ws.insert_rows(96, 1)
write_cell(96, 10, '・正しいメールアドレス＋誤ったパスワード →「メールアドレスまたはパスワードが正しくありません」が表示されること')
print("  NEW-6: Row 96 に挿入（セキュリティ → 認証動作のエラーメッセージ）")
# Row 96以下が+1シフト。Row 95以上は変化なし。

# =====================================================================
# Step 4: NEW-1 → 目次1続き テナント切り替えの後（Row 73の前に2行挿入）
# Row 72 = テナント切替詳細行（・制限あり→制限なしの事業所へ切替　→可）
# Row 73 = 既存セッション ← この前に挿入
# ※ Row 73 < 96 なのでStep3の影響なし
# =====================================================================
ws.insert_rows(73, 2)
write_cell(73, 4, 'マルチテナント（3テナント以上）')
write_cell(73, 6, 'ログイン画面')
write_cell(73, 7, '3テナント以上所属ユーザーの\nログイン判定')
write_cell(73, 10, '・3テナント以上に所属するユーザーで、最古の事業所の制限設定が適用されること')
write_cell(74, 10, '・既存テストは2テナントのみのため、3テナント以上での動作確認を追加')
print("  NEW-1: Row 73-74 に挿入（目次1続き → マルチテナント3テナント以上）")

# =====================================================================
# 検証
# =====================================================================
print("\n=== 挿入結果の検証 ===")

# NEW-1: Row 73-74 の前後を確認
print("\n--- NEW-1 (テナント切替の後、既存セッションの前) ---")
for row in range(71, 78):
    vals = []
    for col in [2, 4, 6, 7, 10]:
        v = ws.cell(row=row, column=col).value
        if v:
            vals.append(f'Col{col}:{str(v)[:60]}')
    if vals:
        print(f"  Row {row}: {' | '.join(vals)}")
    else:
        print(f"  Row {row}: (empty)")

# NEW-6: Row 98 (73+2=元96, 96+2=98) の前後を確認
# 元Row 95 → +2(Step4) = Row 97
# 元Row 96(Step3で挿入) → +2(Step4) = Row 98
print("\n--- NEW-6 (エラーメッセージが同一の後) ---")
for row in range(97, 102):
    vals = []
    for col in [2, 4, 6, 7, 10]:
        v = ws.cell(row=row, column=col).value
        if v:
            vals.append(f'Col{col}:{str(v)[:60]}')
    if vals:
        print(f"  Row {row}: {' | '.join(vals)}")

# NEW-4: 元Row 120-121(Step2で挿入) → +1(Step3) +2(Step4) = Row 123-124
print("\n--- NEW-4 (PW再設定メール送信の後) ---")
for row in range(121, 128):
    vals = []
    for col in [2, 4, 6, 7, 10]:
        v = ws.cell(row=row, column=col).value
        if v:
            vals.append(f'Col{col}:{str(v)[:60]}')
    if vals:
        print(f"  Row {row}: {' | '.join(vals)}")

# NEW-5: 元Row 128(Step1で挿入) → +2(Step2) +1(Step3) +2(Step4) = Row 133
print("\n--- NEW-5 (ローディング表示の後) ---")
for row in range(131, 137):
    vals = []
    for col in [2, 4, 6, 7, 10]:
        v = ws.cell(row=row, column=col).value
        if v:
            vals.append(f'Col{col}:{str(v)[:60]}')
    if vals:
        print(f"  Row {row}: {' | '.join(vals)}")

# 保存
wb.save(v7_path)
print(f"\n保存完了: {v7_path}")
