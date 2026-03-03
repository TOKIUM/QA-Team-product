#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""観点シートに2項目を追加:
1. SSO限定エラー vs IP制限エラーの優先順位
2. 通常IP制限のID/PWログインテスト
"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment

path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'memo', 'ログイン情報', '修正対象', '最新版')
v7_path = os.path.join(path, 'テスト項目書_改善後_v7.xlsx')
wb = openpyxl.load_workbook(v7_path)
ws = wb['観点']

base_font = Font(name='Arial', size=10)
top_align = Alignment(vertical='top', wrap_text=True)

def write_cell(row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = base_font
    cell.alignment = top_align

print("観点シートに2項目を追加します\n")

# =====================================================================
# 挿入計画（下から順に実施）
#
# 項目2: 通常IP制限 → SAML IP制限（Row 87）の後、Row 88（管理ユーザー判定）の前に挿入
# 項目1: SSO vs IP優先順位 → エラーメッセージ差別化（Row 80）の後、Row 81（設定有効化）の前に挿入
# =====================================================================

# Step 1: 項目2「通常IP制限のID/PWログイン」→ Row 88の前に2行挿入
ws.insert_rows(88, 2)
write_cell(88, 4, '通常IP制限のID/PWログイン')
write_cell(88, 6, 'ログイン画面')
write_cell(88, 7, 'IP制限あり環境での\nID/PWログイン拒否')
write_cell(88, 10, '・IP制限あり＋制限フラグなし＋許可外IP →ID/PWログイン不可（IP制限エラー）')
write_cell(89, 10, '・IP制限あり＋制限フラグなし＋許可内IP →ID/PWログイン可')
print("  項目2: Row 88-89 に挿入（SAML IP制限の後 → 通常IP制限のID/PWログイン）")

# Step 2: 項目1「SSO限定エラー vs IP制限エラーの優先順位」→ Row 81の前に2行挿入
# ※Step 1で+2されているが、Row 81 < 88なので影響なし
ws.insert_rows(81, 2)
write_cell(81, 4, 'SSO限定エラーとIP制限エラーの優先順位')
write_cell(81, 6, 'ログイン画面')
write_cell(81, 7, '制限あり＋IP制限あり環境での\nエラー優先順位')
write_cell(81, 10, '・制限あり＋IP制限あり＋許可外IP →ID/PWログイン →SSO限定エラーが表示されること（IP制限エラーではない）')
write_cell(82, 10, '・SSO制限がIP制限より先に評価されるため、SSO限定エラーが優先表示される')
print("  項目1: Row 81-82 に挿入（エラーメッセージ差別化の後 → SSO vs IP優先順位）")

# =====================================================================
# 検証
# =====================================================================
print("\n=== 挿入結果の検証 ===")

print("\n--- 項目1: SSO vs IP優先順位 (Row 79-84) ---")
for row in range(79, 85):
    vals = []
    for col in [2, 4, 6, 7, 10]:
        v = ws.cell(row=row, column=col).value
        if v:
            vals.append(f'Col{col}:{str(v)[:60]}')
    if vals:
        print(f"  Row {row}: {' | '.join(vals)}")

print("\n--- 項目2: 通常IP制限 (Row 89-93) ---")
for row in range(89, 94):
    vals = []
    for col in [2, 4, 6, 7, 10]:
        v = ws.cell(row=row, column=col).value
        if v:
            vals.append(f'Col{col}:{str(v)[:60]}')
    if vals:
        print(f"  Row {row}: {' | '.join(vals)}")

# 保存
wb.save(v7_path)
print(f"\n保存完了: {v7_path}")
