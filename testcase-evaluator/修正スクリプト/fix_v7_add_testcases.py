#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""v7パッチ: テストケース2件追加
1. SSO限定エラー vs IP制限エラーの優先順位
2. 通常IP制限のID/PWログイン（許可外IP/許可内IP）
"""
import os, sys, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'memo', 'ログイン情報', '修正対象', '最新版')
v7_path = os.path.join(path, 'テスト項目書_改善後_v7.xlsx')
wb = openpyxl.load_workbook(v7_path, data_only=True)
ws = wb['テストケース_改善後']

# 現行データ読み込み
cases = []
for row in range(3, ws.max_row + 1):
    no = ws.cell(row=row, column=2).value
    if no is None:
        continue
    cases.append({
        'no': no,
        'confirm': ws.cell(row=row, column=3).value or '',
        'screen': ws.cell(row=row, column=4).value or '',
        'target': ws.cell(row=row, column=5).value or '',
        'detail': ws.cell(row=row, column=6).value or '',
        'steps': ws.cell(row=row, column=7).value or '',
        'expected': ws.cell(row=row, column=8).value or '',
        'notes': ws.cell(row=row, column=9).value or '',
        'orig_no': str(ws.cell(row=row, column=10).value or ''),
    })

print(f"現行: {len(cases)}件")

# =====================================================================
# テストケース1: SSO限定エラー vs IP制限エラーの優先順位
# =====================================================================
# No.190-191の「エラー優先順位」カテゴリに追加
# No.191の後に挿入（同じカテゴリ内に配置）

tc1_cases = [
    {
        'confirm': 'エラー優先順位',
        'screen': 'ログイン画面',
        'target': 'SSO限定エラーがIP制限エラーより優先されること',
        'detail': '制限あり＋IP制限あり＋許可外IP\n制限エラーとIP制限エラーのどちらが優先されるか\n※SSO制限がIP制限より先に評価されるため',
        'steps': ('【前提条件】\n'
                  '・制限フラグ: 当日以前（制限発動中）\n'
                  '・IP制限: あり（テスト実行環境のIPが許可外）\n'
                  '・有効なテストアカウントが存在すること\n'
                  '\n'
                  '【実行手順】\n'
                  '1. 許可外IPからログイン画面を開く\n'
                  '2. 正しいメールアドレス＋パスワードを入力\n'
                  '3. 「ログイン」ボタンを押下'),
        'expected': '「この事業所ではTOKIUM ID でのログインのみ許可されています。TOKIUM ID でログインしてください。」が表示されること（IP制限エラーではなくSSO限定エラーが優先されること）',
        'notes': '仕様書4.2: SSO制限がIP制限より先に評価される',
        'orig_no': 'NEW-7',
    },
]

# =====================================================================
# テストケース2-3: 通常IP制限のID/PWログイン
# =====================================================================
tc2_cases = [
    {
        'confirm': 'IP制限（既存機能確認）',
        'screen': 'ログイン画面',
        'target': 'IP制限あり＋許可外IPでメールアドレス・パスワードでのログイン不可',
        'detail': 'IP制限あり＋制限フラグなし＋許可外IP\n通常のIP制限が機能していることを確認',
        'steps': ('【前提条件】\n'
                  '・制限フラグ: null（制限なし）\n'
                  '・IP制限: あり（テスト実行環境のIPが許可外）\n'
                  '・有効なテストアカウントが存在すること\n'
                  '\n'
                  '【実行手順】\n'
                  '1. 許可外IPからログイン画面を開く\n'
                  '2. 正しいメールアドレス＋パスワードを入力\n'
                  '3. 「ログイン」ボタンを押下'),
        'expected': 'IP制限によりログインが拒否されること',
        'notes': '',
        'orig_no': 'NEW-8',
    },
    {
        'confirm': 'IP制限（既存機能確認）',
        'screen': 'ログイン画面',
        'target': 'IP制限あり＋許可内IPでメールアドレス・パスワードでのログイン可',
        'detail': 'IP制限あり＋制限フラグなし＋許可内IP\n許可されたIPからは通常通りログインできることを確認',
        'steps': ('【前提条件】\n'
                  '・制限フラグ: null（制限なし）\n'
                  '・IP制限: あり（テスト実行環境のIPが許可内）\n'
                  '・有効なテストアカウントが存在すること\n'
                  '\n'
                  '【実行手順】\n'
                  '1. 許可内IPからログイン画面を開く\n'
                  '2. 正しいメールアドレス＋パスワードを入力\n'
                  '3. 「ログイン」ボタンを押下'),
        'expected': 'ログインが成功し、ログイン後画面に遷移すること',
        'notes': '',
        'orig_no': 'NEW-9',
    },
]

# =====================================================================
# ケース結合
# =====================================================================
# No.191(エラー優先順位の最後)の後にtc1を挿入
# 末尾にtc2を追加

# No.191の位置を特定
insert_idx = None
for idx, case in enumerate(cases):
    if case['orig_no'] == '77':
        # エラー優先順位セクションの最終ケースを探す
        # No.191 = orig_no 77の最後のケース
        pass
    # orig_no 77のケースを全て確認して最後のインデックスを取得
    if str(case['orig_no']) == '77':
        insert_idx = idx

if insert_idx is not None:
    # No.191の後に挿入
    insert_pos = insert_idx + 1
    for i, tc in enumerate(tc1_cases):
        cases.insert(insert_pos + i, tc)
    print(f"  SSO vs IP優先順位: index {insert_pos} に{len(tc1_cases)}件挿入")
else:
    # fallback: 末尾に追加
    cases.extend(tc1_cases)
    print("  SSO vs IP優先順位: 末尾に追加（位置特定失敗）")

# IP制限テストケースは末尾に追加
cases.extend(tc2_cases)
print(f"  IP制限テスト: 末尾に{len(tc2_cases)}件追加")

# No振り直し
for idx, case in enumerate(cases):
    case['no'] = idx + 1

print(f"\n追加後: {len(cases)}件（+{len(tc1_cases) + len(tc2_cases)}件）")

# =====================================================================
# Output
# =====================================================================
wb_out = openpyxl.load_workbook(v7_path)
if 'テストケース_改善後' in wb_out.sheetnames:
    del wb_out['テストケース_改善後']

ws_new = wb_out.create_sheet('テストケース_改善後', 0)

header_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
cat_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
data_font = Font(size=10)
wrap_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

for col, val in {12:'Win11/\nChrome',14:'Win11/\nEdge',15:'Win11/\nFirefox',16:'Mac/\nSafari',17:'iOS/\nOSver●●',18:'Android/\nOSver●●'}.items():
    cell = ws_new.cell(row=1, column=col, value=val)
    cell.font = Font(bold=True, size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col, val in {2:'No',3:'確認すること',4:'画面',5:'確認対象',6:'詳細',7:'テスト実行手順',8:'期待値',9:'備考',10:'元No.',11:'Autify',12:'TK+TI+TD',13:'TK+TD',14:'TI+TD',15:'TK単体',16:'TD単体',17:'TK+TI+TD',18:'TK+TI+TD',19:'実行者',20:'実行備考',21:'起票JIRA No'}.items():
    cell = ws_new.cell(row=2, column=col, value=val)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

prev_category = None
for idx, case in enumerate(cases):
    row_num = idx + 3
    ws_new.cell(row=row_num, column=2, value=case['no']).font = data_font
    ws_new.cell(row=row_num, column=2).alignment = center_align
    ws_new.cell(row=row_num, column=2).border = thin_border
    for col, key in {3:'confirm',4:'screen',5:'target',6:'detail',7:'steps',8:'expected',9:'notes'}.items():
        cell = ws_new.cell(row=row_num, column=col, value=case.get(key, ''))
        cell.font = data_font; cell.alignment = wrap_align; cell.border = thin_border
    category = case.get('confirm', '')
    if category != prev_category:
        ws_new.cell(row=row_num, column=3).fill = cat_fill
        prev_category = category
    cell_orig = ws_new.cell(row=row_num, column=10, value=case.get('orig_no', ''))
    cell_orig.font = data_font; cell_orig.alignment = center_align; cell_orig.border = thin_border
    for col in range(12, 19):
        cell = ws_new.cell(row=row_num, column=col, value='未')
        cell.font = data_font; cell.alignment = center_align; cell.border = thin_border
    ws_new.cell(row=row_num, column=11).border = thin_border
    for col in [19, 20, 21]:
        ws_new.cell(row=row_num, column=col).border = thin_border

for col, width in {2:6,3:28,4:18,5:28,6:40,7:55,8:55,9:30,10:8,11:8,12:10,13:8,14:8,15:8,16:8,17:10,18:10,19:10,20:15,21:15}.items():
    ws_new.column_dimensions[get_column_letter(col)].width = width
for row_num in range(3, len(cases) + 3):
    ws_new.row_dimensions[row_num].height = 72

wb_out.save(v7_path)

print(f"\n保存完了: {v7_path}")

# =====================================================================
# 追加ケースの確認
# =====================================================================
print("\n=== 追加テストケースの確認 ===")
for c in cases:
    if c['orig_no'].startswith('NEW-7') or c['orig_no'].startswith('NEW-8') or c['orig_no'].startswith('NEW-9'):
        print(f"\nNo.{c['no']} (元No.{c['orig_no']}):")
        print(f"  確認すること: {c['confirm']}")
        print(f"  画面: {c['screen']}")
        print(f"  確認対象: {c['target']}")
        print(f"  詳細: {c['detail']}")
        print(f"  手順: {c['steps'][:120]}")
        print(f"  期待値: {c['expected'][:100]}")
        print(f"  備考: {c['notes']}")

# 最終チェック
print("\n=== 最終チェック ===")
pw = sum(1 for c in cases for f in ['confirm','target','detail','steps','expected','notes'] if re.search(r'(?<!パス)PW', c.get(f, '') or ''))
print(f"PW残存: {pw}件 {'✓' if pw == 0 else '✗'}")
idpw = sum(1 for c in cases for f in ['confirm','target','detail','steps','expected','notes'] if 'ID/PW' in (c.get(f, '') or ''))
print(f"ID/PW残存: {idpw}件 {'✓' if idpw == 0 else '✗'}")
print(f"総ケース数: {len(cases)}件")
