#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
v7修正:
1. スクリーンショットに基づくUI文言修正
2. 日本語の不自然な表現を修正
3. 残りの曖昧表現チェック
"""
import os, sys, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

path = os.path.join(os.environ['USERPROFILE'], 'Desktop', 'memo', 'ログイン情報', '修正対象', '最新版')
v6_path = os.path.join(path, 'テスト項目書_改善後_v6.xlsx')
wb = openpyxl.load_workbook(v6_path, data_only=True)
ws = wb['テストケース_改善後']

cases = []
for row in range(3, ws.max_row + 1):
    no = ws.cell(row=row, column=2).value
    if no is None:
        continue
    cases.append({
        'no': no,
        'confirm': ws.cell(row=row, column=3).value or '',
        'screen': ws.cell(row=row, column=4).value or '',
        'target': ws.cell(row=row, column=5).value or '',
        'detail': ws.cell(row=row, column=6).value or '',
        'steps': ws.cell(row=row, column=7).value or '',
        'expected': ws.cell(row=row, column=8).value or '',
        'notes': ws.cell(row=row, column=9).value or '',
        'orig_no': str(ws.cell(row=row, column=10).value or ''),
    })

print(f"読み込み: {len(cases)}件\n")
text_fields = ['confirm', 'screen', 'target', 'detail', 'steps', 'expected', 'notes']

# =====================================================================
# 修正1: スクリーンショットに基づくUI文言修正
# =====================================================================
print("=" * 70)
print("【修正1】スクリーンショットに基づくUI文言修正")
print("=" * 70)

fix1_count = 0

for case in cases:
    for field in text_fields:
        val = case[field]
        if not val or val == '-':
            continue
        original = val

        # (A) 「ユーザー登録はこちら」→「新規登録はこちら」
        # スクリーンショットのTHログイン画面のリンクテキストは「新規登録はこちら」
        val = val.replace('「ユーザー登録はこちら」リンク', '「新規登録はこちら」リンク')
        val = val.replace('「ユーザー登録はこちら」', '「新規登録はこちら」')
        val = val.replace('ユーザー登録はこちら', '新規登録はこちら')

        # (B) 「/registrationに遷移」→「ユーザー登録画面（/registration）に遷移」
        val = val.replace('/registrationに遷移すること', 'ユーザー登録画面に遷移すること')
        val = val.replace('/registration', 'ユーザー登録画面（/registration）')

        # (C) 「ユーザー登録画面（リグレッション）」→確認すること列のカテゴリ名は維持
        # ただしリンクテキストとしては「新規登録はこちら」が正しい

        # (D) パスワード再設定画面の注意書き
        # 実画面: 「「TOKIUM ID でログイン」経由でログインしたユーザーのパスワードはこちらから再設定できません。」
        # v6: 「「TOKIUM ID でログイン」経由でログインしたユーザーのパスワードはこちらから再設定できません。」
        # → 一致確認OK

        # (E) パスワード再設定画面のボタン
        # 実画面: 「送信する」ボタン
        val = val.replace('「送信」ボタン', '「送信する」ボタン')

        # (F) パスワード再設定画面のタイトル
        # 実画面: 「パスワード再設定」
        val = val.replace('パスワードリカバリー画面', 'パスワード再設定画面')

        # (G) 画面名の統一（スクリーンショット確認済み）
        # THログイン画面 = invoicing-staging.keihi.com/login
        # パスワード再設定画面 = invoicing-staging.keihi.com/recovery
        # ユーザー登録画面 = invoicing-staging.keihi.com/registration

        if val != original:
            case[field] = val
            fix1_count += 1
            if fix1_count <= 20:
                print(f"  No.{case['no']} (元No.{case['orig_no']}) {field}")

print(f"\nUI文言修正: {fix1_count}件\n")

# =====================================================================
# 修正2: 日本語の不自然な表現を修正
# =====================================================================
print("=" * 70)
print("【修正2】日本語の不自然な表現修正")
print("=" * 70)

fix2_count = 0

# 不自然な表現のパターンと修正（長い複合パターンを先に配置）
awkward_patterns = [
    # 複合パターン（先にマッチさせて重複を防ぐ）
    ('押下後を注意深く観察', '押下後の挙動を確認'),
    ('外観を目視確認', '外観を確認する'),
    # 個別パターン
    ('注意深く観察', '挙動を確認'),
    ('を注意深く', 'の挙動を'),
    ('押下後を', '押下後の挙動を'),
    ('目視確認', '確認する'),
    # 「比較」（文末で動詞が省略されている）
    # 「エラーを記録」→「エラーメッセージを確認する」系はOK（テスト手順として自然）
    # 「リカバリーリンク」→「パスワード再設定リンク」（画面と一致させる）
    ('リカバリーリンク', 'パスワード再設定リンク'),
    ('リカバリーメール', 'パスワード再設定メール'),
    ('リカバリー正常完了', 'パスワード再設定が完了'),
    # 「フォーム表示されること」→「フォームが表示されること」（助詞抜け）
    ('フォーム表示されること', 'フォームが表示されること'),
    # 「ログイン後画面に遷移」→ OK (自然)
    # 「切替先の事業所画面」→ OK
    # 「ダッシュボード等の」→ 削除（画面名として曖昧）
    ('ダッシュボード等の', ''),
]

for case in cases:
    for field in text_fields:
        val = case[field]
        if not val or val == '-':
            continue
        original = val
        for old, new in awkward_patterns:
            if old in val:
                val = val.replace(old, new)
        if val != original:
            case[field] = val
            fix2_count += 1
            print(f"  No.{case['no']} (元No.{case['orig_no']}) {field}: {repr(original[:60])} → {repr(val[:60])}")

print(f"\n日本語表現修正: {fix2_count}件\n")

# =====================================================================
# 修正3: 全文チェック - 残りの曖昧・不自然な表現
# =====================================================================
print("=" * 70)
print("【チェック3】全期待値の品質チェック")
print("=" * 70)

# 曖昧表現パターン
ambiguous_check = [
    '正常', '問題ない', '仕様通り', '適切な', '正しく', '中途半端',
    '副作用', '影響がない', '注意深く', '目視',
]

amb_count = 0
for case in cases:
    exp = case['expected']
    if not exp or exp == '-':
        continue
    for word in ambiguous_check:
        if word in exp:
            amb_count += 1
            print(f"  No.{case['no']} 期待値に「{word}」: {exp[:80]}")
            break

print(f"\n期待値の曖昧表現: {amb_count}件\n")

# 手順の不自然チェック
print("--- 手順の不自然表現チェック ---")
step_issues = 0
step_awkward = ['注意深く', '目視', 'を観察', '押下後を']
for case in cases:
    steps = case['steps']
    if not steps or steps == '-':
        continue
    for word in step_awkward:
        if word in steps:
            step_issues += 1
            # 該当行を抽出
            for line in steps.split('\n'):
                if word in line:
                    print(f"  No.{case['no']} 手順に「{word}」: {line.strip()[:80]}")
                    break
            break

print(f"手順の不自然表現: {step_issues}件\n")

# =====================================================================
# 修正4: 全文のURL/画面名チェック
# =====================================================================
print("=" * 70)
print("【修正4】URL/画面名チェック")
print("=" * 70)

fix4_count = 0

# 画面名のマッピング（スクリーンショット確認済み）
# THログイン画面のUI要素:
#   ラベル: メールアドレス, パスワード
#   プレースホルダ: (メールアドレス欄はなし), 8文字以上のパスワードを入力
#   ボタン: ログイン, TOKIUM ID でログイン
#   リンク: パスワードを忘れた場合, 新規登録はこちら

# パスワード再設定画面のUI要素:
#   テキスト: 登録されたメールアドレスにパスワード再設定用のメールを送信いたします。
#   ラベル: メールアドレス
#   ボタン: 送信する
#   注意書き: 「TOKIUM ID でログイン」経由でログインしたユーザーのパスワードはこちらから再設定できません。
#   リンク: 戻る

# ユーザー登録画面のUI要素:
#   ラベル: メールアドレス, パスワード, パスワード（再入力）
#   プレースホルダ: (パスワード再入力)同じパスワードを入力
#   ボタン: ユーザー登録する
#   リンク: ログインする方はこちら

for case in cases:
    for field in text_fields:
        val = case[field]
        if not val or val == '-':
            continue
        original = val

        # 「THログイン画面」→ そのままでOK（正式名称としてテスト項目書で使用）
        # ただし画面列で「ログイン画面」とだけ書いている場合は「THログイン画面」と統一すべきか？
        # → 画面列は「ログイン画面」で統一されているのでそのまま

        # 「パスワードを忘れた場合」→ 画面上もこのテキスト。OK
        # 「新規登録はこちら」→ 画面上もこのテキスト。OK（修正1で対応済み）

        if val != original:
            case[field] = val
            fix4_count += 1

print(f"URL/画面名修正: {fix4_count}件\n")

# =====================================================================
# 修正5: 期待値の細かい品質修正
# =====================================================================
print("=" * 70)
print("【修正5】期待値の細かい品質修正")
print("=" * 70)

fix5_count = 0
for case in cases:
    exp = case['expected']
    orig_no = case['orig_no']
    if not exp or exp == '-':
        continue
    original = exp

    # 「フォーム表示されること」→「フォームが表示されること」
    exp = exp.replace('フォーム表示されること', 'フォームが表示されること')

    # 「全インタラクティブ要素」→「すべての入力欄・ボタン」
    if case['target'] == '全インタラクティブ要素':
        case['target'] = 'すべての入力欄・ボタン'
        fix5_count += 1
        print(f"  No.{case['no']} 確認対象: 全インタラクティブ要素→すべての入力欄・ボタン")

    # 「多重リダイレクトが発生せず、1回だけ遷移すること」はOK
    # 「1回だけ遷移すること」はOK

    # 「エラーなく送信できること」→「エラーが表示されずに送信できること」
    exp = exp.replace('エラーなく送信できること', 'エラーが表示されずに送信できること')

    # 「ログインボタン押下時と同様にフォーム送信されること」→「ログインボタン押下時と同様にログイン処理が実行されること」
    exp = exp.replace('ログインボタン押下時と同様にフォーム送信されること', 'ログインボタン押下時と同様にログイン処理が実行されること')

    # 「制限が再設定メール送信をブロックしないこと」→「制限ありの状態でもパスワード再設定メールが送信されること」
    exp = exp.replace('制限が再設定メール送信をブロックしないこと', '制限ありの状態でもパスワード再設定メールが送信されること')

    # 「ログイン画面に戻らないこと」は具体的でOK
    # 「認証済みページが表示されないこと」は具体的でOK

    # 「ユーザーが次に何をすべきか（管理者に連絡等）が理解できる案内であること」→曖昧
    if '理解できる案内であること' in exp:
        exp = exp.replace(
            'ユーザーが次に何をすべきか（管理者に連絡等）が理解できる案内であること',
            '「管理者にお問い合わせください」等の次の行動を示す案内が表示されていること'
        )
        fix5_count += 1
        print(f"  No.{case['no']}: 理解できる案内→具体的な案内文言に")

    # 「用途がラベルで理解できること」→曖昧
    if '用途がラベルで理解できること' in exp:
        exp = exp.replace(
            '用途がラベルで理解できること',
            '各ボタンのラベルから用途が判別できること（「ログイン」「TOKIUM ID でログイン」）'
        )
        fix5_count += 1
        print(f"  No.{case['no']}: ラベルで理解できる→具体的なラベル名に")

    # 「フォーカスが視覚的に識別できること」→やや曖昧だが許容範囲
    if '視覚的に識別できること' in exp:
        exp = exp.replace(
            'フォーカスが視覚的に識別できること',
            'フォーカスされている要素が枠線やハイライト等で識別できること'
        )
        fix5_count += 1
        print(f"  No.{case['no']}: 視覚的に識別→枠線やハイライト等で識別")

    if exp != original:
        case['expected'] = exp
        if exp != original and '→' not in f"{original}→{exp}":
            fix5_count += 1

print(f"\n期待値品質修正: {fix5_count}件\n")

# =====================================================================
# 最終全文チェック
# =====================================================================
print("=" * 70)
print("【最終チェック】")
print("=" * 70)

# PW残存
pw = sum(1 for c in cases for f in text_fields if re.search(r'(?<!パス)PW', c.get(f, '') or ''))
print(f"PW残存: {pw}件 {'✓' if pw == 0 else '✗'}")

# 曖昧表現
amb = 0
for c in cases:
    exp = c['expected']
    for w in ['正常', '問題ない', '仕様通り', '適切な', '中途半端', '副作用', '注意深く']:
        if w in exp:
            amb += 1
            break
print(f"曖昧表現(期待値): {amb}件 {'✓' if amb == 0 else '✗'}")

# ユーザー登録はこちら残存（正しくは新規登録はこちら）
reg = sum(1 for c in cases for f in text_fields if 'ユーザー登録はこちら' in (c.get(f, '') or ''))
print(f"「ユーザー登録はこちら」残存: {reg}件 {'✓' if reg == 0 else '✗'}")

# /registration 直接記載残存
regurl = sum(1 for c in cases for f in text_fields if '/registration' in (c.get(f, '') or '') and 'ユーザー登録画面' not in (c.get(f, '') or ''))
print(f"/registration直接記載残存: {regurl}件 {'✓' if regurl == 0 else '✗'}")

# ID/PW残存
idpw = sum(1 for c in cases for f in text_fields if 'ID/PW' in (c.get(f, '') or ''))
print(f"ID/PW残存: {idpw}件 {'✓' if idpw == 0 else '✗'}")

# こと末尾
koto = 0
for c in cases:
    exp = c['expected']
    if not exp or exp == '-':
        continue
    s = exp.rstrip()
    if s.endswith('）') or s.endswith(')'):
        inner = re.sub(r'[（(][^）)]*[）)]$', '', s).strip()
        if inner.endswith('こと'):
            continue
    if not s.endswith('こと'):
        koto += 1
        print(f"  こと末尾でない: No.{c['no']}: {repr(s[:80])}")
print(f"「こと」末尾でない: {koto}件 {'✓' if koto == 0 else '✗'}")

# =====================================================================
# Output
# =====================================================================
wb_out = openpyxl.load_workbook(v6_path)
if 'テストケース_改善後' in wb_out.sheetnames:
    del wb_out['テストケース_改善後']

ws_new = wb_out.create_sheet('テストケース_改善後', 0)

header_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
cat_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
data_font = Font(size=10)
wrap_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

for col, val in {12:'Win11/\nChrome',14:'Win11/\nEdge',15:'Win11/\nFirefox',16:'Mac/\nSafari',17:'iOS/\nOSver●●',18:'Android/\nOSver●●'}.items():
    cell = ws_new.cell(row=1, column=col, value=val)
    cell.font = Font(bold=True, size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col, val in {2:'No',3:'確認すること',4:'画面',5:'確認対象',6:'詳細',7:'テスト実行手順',8:'期待値',9:'備考',10:'元No.',11:'Autify',12:'TK+TI+TD',13:'TK+TD',14:'TI+TD',15:'TK単体',16:'TD単体',17:'TK+TI+TD',18:'TK+TI+TD',19:'実行者',20:'実行備考',21:'起票JIRA No'}.items():
    cell = ws_new.cell(row=2, column=col, value=val)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

prev_category = None
for idx, case in enumerate(cases):
    row_num = idx + 3
    ws_new.cell(row=row_num, column=2, value=case['no']).font = data_font
    ws_new.cell(row=row_num, column=2).alignment = center_align
    ws_new.cell(row=row_num, column=2).border = thin_border
    for col, key in {3:'confirm',4:'screen',5:'target',6:'detail',7:'steps',8:'expected',9:'notes'}.items():
        cell = ws_new.cell(row=row_num, column=col, value=case.get(key, ''))
        cell.font = data_font; cell.alignment = wrap_align; cell.border = thin_border
    category = case.get('confirm', '')
    if category != prev_category:
        ws_new.cell(row=row_num, column=3).fill = cat_fill
        prev_category = category
    cell_orig = ws_new.cell(row=row_num, column=10, value=case.get('orig_no', ''))
    cell_orig.font = data_font; cell_orig.alignment = center_align; cell_orig.border = thin_border
    for col in range(12, 19):
        cell = ws_new.cell(row=row_num, column=col, value='未')
        cell.font = data_font; cell.alignment = center_align; cell.border = thin_border
    ws_new.cell(row=row_num, column=11).border = thin_border
    for col in [19, 20, 21]:
        ws_new.cell(row=row_num, column=col).border = thin_border

for col, width in {2:6,3:28,4:18,5:28,6:40,7:55,8:55,9:30,10:8,11:8,12:10,13:8,14:8,15:8,16:8,17:10,18:10,19:10,20:15,21:15}.items():
    ws_new.column_dimensions[get_column_letter(col)].width = width
for row_num in range(3, len(cases) + 3):
    ws_new.row_dimensions[row_num].height = 72

output_path = os.path.join(path, 'テスト項目書_改善後_v7.xlsx')
wb_out.save(output_path)

print(f"\n===== v7修正完了 =====")
print(f"ケース数: {len(cases)}件")
print(f"出力先: {output_path}")
print(f"\n--- 修正サマリー ---")
print(f"  1. UI文言修正: {fix1_count}件")
print(f"  2. 日本語表現修正: {fix2_count}件")
print(f"  3. 期待値品質修正: {fix5_count}件")
