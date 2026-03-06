#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
観点シート 内容妥当性 自動評価ツール

テスト項目書（Excel）の観点シートに記載された設計内容が
仕様書・テストケースと整合しているかを自動評価する。

使い方:
  # 基本（Excel単体で評価）
  python evaluate_perspectives.py テスト項目書.xlsx

  # 仕様書を指定して整合性も評価
  python evaluate_perspectives.py テスト項目書.xlsx --spec spec.md

  # テストケースJSONも指定（フル評価）
  python evaluate_perspectives.py テスト項目書.xlsx --spec spec.md --testcases tc.json

  # JSON出力
  python evaluate_perspectives.py テスト項目書.xlsx -o result.json

評価基準（100点満点）:
  1. 構造妥当性（20点）: 目次・セクション構成が正しいか
  2. リスト網羅性（20点）: テストケースの観点がリストに反映されているか
  3. マトリクス整合性（20点）: 組合せ表の○×が実態と合っているか
  4. デシジョンテーブル論理性（20点）: 条件分岐が仕様と整合しているか
  5. リスク観点妥当性（20点）: リスク該当判定が仕様に適切か
"""

import argparse
import json
import os
import re
import sys

import io
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl")
    sys.exit(1)


# =====================================================================
# 定数
# =====================================================================

KANTEN_PERSPECTIVE_START = 27  # 観点設計セクションの開始行

# セクションタイトルのパターン（タイトルバー行 = 「目次N |」で始まる行のみ対象）
# 順序重要: チェックリスト形式 を リスト形式 より先に判定（部分マッチ防止）
SECTION_PATTERNS = [
    ("risk_checklist", re.compile(r"チェックリスト形式")),
    ("decision_table", re.compile(r"デシジョンテーブル形式")),
    ("matrix", re.compile(r"組合せ形式")),
    ("list", re.compile(r"(?<!チェック)リスト形式")),
]

# タイトルバー行の識別パターン（「目次N  |」で始まる行）
TITLE_BAR_PATTERN = re.compile(r"^\s*目次\d+\s*\|")


# =====================================================================
# 観点シート読み取り
# =====================================================================

def read_perspective_sheet(xlsx_path):
    """観点シートから設計セクションを読み取る。

    Returns:
        dict: {
            "toc_items": [...],
            "sections": [{"type": ..., "start_row": ..., "data": ...}, ...],
            "raw_rows": {row_num: [cell_values]},
        }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "観点" not in wb.sheetnames:
        return {"error": "観点シートが存在しません", "toc_items": [], "sections": [], "raw_rows": {}}

    ws = wb["観点"]
    result = {"toc_items": [], "sections": [], "raw_rows": {}}

    # Row 27以降を走査
    max_row = ws.max_row or 150
    for row in range(KANTEN_PERSPECTIVE_START, min(max_row + 1, 200)):
        row_vals = []
        for col in range(1, 20):
            row_vals.append(ws.cell(row=row, column=col).value)
        result["raw_rows"][row] = row_vals

    # 目次を検出
    toc_items = []
    sections = []
    current_section = None

    for row_num, vals in sorted(result["raw_rows"].items()):
        b_val = str(vals[1] or "").strip()  # B列

        # 目次タイトル行
        if "観点の目次" in b_val:
            continue

        # 目次エントリ（「目次N:」または「N. 」で始まる行で「行 NN」を含む）
        toc_match = re.match(r"(?:目次\d+[:：]|^\s*\d+\.\s+)", b_val)
        row_ref = re.search(r"行\s*(\d+)", b_val)
        if toc_match and row_ref:
            toc_items.append({
                "text": b_val,
                "target_row": int(row_ref.group(1)),
                "source_row": row_num,
            })
            continue

        # セクションタイトルバー検出（「目次N |」で始まる行のみ）
        if TITLE_BAR_PATTERN.match(b_val):
            for sec_type, pattern in SECTION_PATTERNS:
                if pattern.search(b_val):
                    if current_section:
                        current_section["end_row"] = row_num - 1
                        sections.append(current_section)
                    current_section = {
                        "type": sec_type,
                        "title": b_val,
                        "start_row": row_num,
                        "end_row": None,
                        "data_rows": [],
                    }
                    break
        else:
            # セクション内のデータ行
            if current_section and any(v is not None for v in vals):
                current_section["data_rows"].append({
                    "row": row_num,
                    "values": vals,
                })

    if current_section:
        current_section["end_row"] = max(result["raw_rows"].keys())
        sections.append(current_section)

    result["toc_items"] = toc_items
    result["sections"] = sections

    wb.close()
    return result


def read_testcase_sheet(xlsx_path):
    """テストケースシートからテストケースを読み取る。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    test_cases = []

    if "テストケース" not in wb.sheetnames:
        wb.close()
        return test_cases

    ws = wb["テストケース"]

    # ヘッダー行を特定（「確認すること」を含む行）
    header_row = None
    for row in range(1, 10):
        for col in range(1, 20):
            val = ws.cell(row=row, column=col).value
            if val and "確認すること" in str(val):
                header_row = row
                break
        if header_row:
            break

    if not header_row:
        wb.close()
        return test_cases

    # ヘッダーマッピング
    headers = {}
    for col in range(1, 20):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col

    # データ行読み取り
    confirm_col = headers.get("確認すること")
    screen_col = headers.get("画面")
    target_col = headers.get("確認対象")
    expected_col = headers.get("期待値")
    steps_col = headers.get("手順")

    for row in range(header_row + 1, ws.max_row + 1):
        confirm = ws.cell(row=row, column=confirm_col).value if confirm_col else None
        if not confirm:
            continue
        tc = {
            "confirm": str(confirm).strip(),
            "screen": str(ws.cell(row=row, column=screen_col).value or "").strip() if screen_col else "",
            "target": str(ws.cell(row=row, column=target_col).value or "").strip() if target_col else "",
            "expected": str(ws.cell(row=row, column=expected_col).value or "").strip() if expected_col else "",
            "steps": str(ws.cell(row=row, column=steps_col).value or "").strip() if steps_col else "",
        }
        test_cases.append(tc)

    wb.close()
    return test_cases


# =====================================================================
# 評価ロジック
# =====================================================================

def evaluate_structure(perspective_data):
    """1. 構造妥当性（20点）"""
    score = 20
    issues = []
    sections = perspective_data["sections"]
    toc_items = perspective_data["toc_items"]

    # 目次が存在するか
    if not toc_items:
        score -= 5
        issues.append("目次が存在しない")

    # セクション数チェック（最低3セクション必要）
    if len(sections) < 3:
        score -= 5
        issues.append(f"セクション数が不足（{len(sections)}個。3個以上推奨）")

    # 必須セクションタイプ
    section_types = {s["type"] for s in sections}
    required = {"list", "matrix"}
    missing = required - section_types
    if missing:
        score -= 3 * len(missing)
        for m in missing:
            issues.append(f"必須セクション「{m}」が欠落")

    # 目次とセクションの対応
    if toc_items and sections:
        toc_target_rows = {t["target_row"] for t in toc_items}
        section_start_rows = {s["start_row"] for s in sections}
        mismatched = toc_target_rows - section_start_rows
        if mismatched:
            score -= 2
            issues.append(f"目次の参照行とセクション開始行が不一致（行: {mismatched}）")

    # 各セクションにデータ行があるか
    empty_sections = [s for s in sections if not s["data_rows"]]
    if empty_sections:
        score -= 2 * len(empty_sections)
        for s in empty_sections:
            issues.append(f"セクション「{s['title'][:30]}」にデータ行がない")

    return max(0, score), issues


def evaluate_list_coverage(perspective_data, test_cases):
    """2. リスト網羅性（20点）: テストケースの観点がリストに反映されているか"""
    score = 20
    issues = []

    list_sections = [s for s in perspective_data["sections"] if s["type"] == "list"]
    if not list_sections:
        return 0, ["リスト形式セクションが存在しない"]

    if not test_cases:
        return score, ["テストケースが指定されていないためリスト網羅性は未評価"]

    # リストセクションのテキストを結合
    list_texts = []
    for sec in list_sections:
        for dr in sec["data_rows"]:
            row_text = " ".join(str(v) for v in dr["values"] if v is not None)
            list_texts.append(row_text.lower())
    all_list_text = " ".join(list_texts)

    # テストケースの「確認すること」がリストに含まれるか
    tc_confirms = set()
    for tc in test_cases:
        c = tc.get("confirm", "")
        # 異常系プレフィクスを除去
        base = re.sub(r"^(異常系[：:]?\s*)", "", c)
        if base:
            tc_confirms.add(base)

    missing = []
    for confirm in tc_confirms:
        if confirm.lower() not in all_list_text:
            missing.append(confirm)

    if missing:
        coverage_ratio = 1 - len(missing) / max(len(tc_confirms), 1)
        deduction = int((1 - coverage_ratio) * 15)
        score -= deduction
        if len(missing) <= 5:
            for m in missing:
                issues.append(f"リストに未記載: 「{m}」")
        else:
            issues.append(f"リストに未記載の観点が{len(missing)}件（{len(tc_confirms)}件中）")

    # リストにテストケースに無い余分な行がないか（情報提供のみ、減点なし）
    list_row_count = sum(len(s["data_rows"]) for s in list_sections)
    header_rows = len(list_sections)  # ヘッダー行分
    data_rows = list_row_count - header_rows
    if data_rows > len(tc_confirms) * 1.5:
        issues.append(f"（参考）リスト行数({data_rows})がテスト観点数({len(tc_confirms)})の1.5倍超")

    return max(0, score), issues


def evaluate_matrix_consistency(perspective_data, test_cases, spec_text):
    """3. マトリクス整合性（20点）: 組合せ表の○×が実態と合っているか"""
    score = 20
    issues = []

    matrix_sections = [s for s in perspective_data["sections"] if s["type"] == "matrix"]
    if not matrix_sections:
        return 0, ["マトリクス形式セクションが存在しない"]

    if not test_cases:
        return score, ["テストケースが指定されていないためマトリクス整合性は未評価"]

    for sec in matrix_sections:
        title = sec.get("title", "")
        data_rows = sec["data_rows"]

        if not data_rows:
            score -= 5
            issues.append(f"マトリクス「{title[:30]}」にデータがない")
            continue

        # ヘッダー行を特定（最初のデータ行）
        header_row = data_rows[0]
        header_vals = [str(v).strip() for v in header_row["values"] if v is not None]

        # テスト種別マトリクスの検証
        if "正常系" in " ".join(header_vals) or "異常系" in " ".join(header_vals):
            _score, _issues = _check_type_matrix(data_rows[1:], test_cases)
            score += _score  # _score is negative or zero
            issues.extend(_issues)

        # 環境マトリクスの検証
        env_names = ["TH", "WDL", "TD", "TK"]
        if any(env in " ".join(header_vals) for env in env_names):
            _score, _issues = _check_env_matrix(data_rows[1:], test_cases, spec_text)
            score += _score
            issues.extend(_issues)

    return max(0, score), issues


def _check_type_matrix(data_rows, test_cases):
    """テスト種別マトリクスの○×をテストケースと照合"""
    deduction = 0
    issues = []
    mismatches = 0

    for dr in data_rows:
        vals = dr["values"]
        # B列=確認対象名
        target = str(vals[1] or "").strip()
        if not target:
            continue

        # テストケースから実際のカバー状況を算出
        has_normal = any(
            not any(kw in tc.get("confirm", "") for kw in ["異常", "エラー"])
            and target in tc.get("confirm", "")
            for tc in test_cases
        )
        has_abnormal = any(
            any(kw in tc.get("confirm", "") for kw in ["異常", "エラー"])
            and target in tc.get("confirm", "")
            for tc in test_cases
        )

        # マトリクスの○×を取得（列位置は可変なので○×を探す）
        matrix_vals = []
        for v in vals[2:]:
            if v is not None:
                matrix_vals.append(str(v).strip())

        # ○が正常系に対応するかチェック
        if len(matrix_vals) >= 2:
            normal_marked = matrix_vals[0] == "○"
            abnormal_marked = matrix_vals[1] == "○"

            if has_normal != normal_marked:
                mismatches += 1
                issues.append(f"種別マトリクス不整合: 「{target[:20]}」正常系=実態:{has_normal}/記載:{normal_marked}")
            if has_abnormal != abnormal_marked:
                mismatches += 1
                issues.append(f"種別マトリクス不整合: 「{target[:20]}」異常系=実態:{has_abnormal}/記載:{abnormal_marked}")

    if mismatches > 0:
        deduction = -min(10, mismatches * 2)

    return deduction, issues


def _check_env_matrix(data_rows, test_cases, spec_text):
    """環境マトリクスの○×を仕様・テストケースと照合"""
    deduction = 0
    issues = []
    mismatches = 0

    for dr in data_rows:
        vals = dr["values"]
        target = str(vals[1] or "").strip()
        if not target:
            continue

        # 「（THのみ）」パターンの検出
        only_match = re.search(r"[（(]([\w/]+)のみ[）)]", target)
        if only_match:
            only_env = only_match.group(1)
            # ○×の中に、only_env以外で○があれば不整合
            for v in vals[2:]:
                v_str = str(v or "").strip()
                if v_str == "○":
                    # この列の環境名を特定する必要があるが、
                    # ここでは構造的に検証が難しいため、パターン検出のみ
                    pass

    if mismatches > 0:
        deduction = -min(10, mismatches * 2)

    return deduction, issues


def evaluate_decision_table(perspective_data, test_cases, spec_text):
    """4. デシジョンテーブル論理性（20点）"""
    score = 20
    issues = []

    dt_sections = [s for s in perspective_data["sections"] if s["type"] == "decision_table"]
    if not dt_sections:
        # DTが無い場合、仕様に条件分岐があるなら減点
        if spec_text and any(kw in spec_text for kw in ["場合", "条件", "とき", "if", "IF"]):
            score -= 10
            issues.append("仕様に条件分岐があるがデシジョンテーブルが存在しない")
        else:
            issues.append("デシジョンテーブルなし（仕様に条件分岐が少ないため許容）")
        return max(0, score), issues

    for sec in dt_sections:
        data_rows = sec["data_rows"]
        if not data_rows:
            score -= 10
            issues.append("デシジョンテーブルにデータがない")
            continue

        # 条件セクションと動作セクションの検出
        condition_rows = []
        action_rows = []
        in_conditions = False
        in_actions = False

        for dr in data_rows:
            vals = dr["values"]
            b_val = str(vals[1] or "").strip()

            if "条件" in b_val and len(b_val) < 10:
                in_conditions = True
                in_actions = False
                continue
            elif "動作" in b_val and len(b_val) < 10:
                in_conditions = False
                in_actions = True
                continue

            if in_conditions and b_val:
                condition_rows.append(dr)
            elif in_actions and b_val:
                action_rows.append(dr)

        # 条件数チェック
        if not condition_rows:
            score -= 5
            issues.append("デシジョンテーブルに条件行がない")

        # 動作数チェック
        if not action_rows:
            score -= 5
            issues.append("デシジョンテーブルに動作行がない")

        # パターン列のY/N/X値の妥当性
        pattern_count = 0
        inconsistent_patterns = 0
        for dr in condition_rows:
            vals = dr["values"]
            yn_vals = [str(v or "").strip() for v in vals[4:] if v is not None]
            yn_valid = [v for v in yn_vals if v in ("Y", "N", "-", "")]
            if yn_vals:
                pattern_count = max(pattern_count, len(yn_vals))
                if len(yn_valid) < len(yn_vals) * 0.8:
                    inconsistent_patterns += 1

        if inconsistent_patterns > 0:
            score -= 3
            issues.append(f"デシジョンテーブルの条件値にY/N以外の不正な値が{inconsistent_patterns}行")

        # 動作行のカバレッジ
        # DTでは「各パターンに1動作」の1対1対応が一般的。
        # 動作マーク（X/○等）がない列は「その動作は発生しない」を意味し正常。
        # 問題になるのは「動作行が全て空 = どのパターンにも動作が紐づかない」場合のみ。
        if action_rows and pattern_count > 0:
            any_action_marked = False
            for dr in action_rows:
                vals = dr["values"]
                action_vals = [str(v or "").strip() for v in vals[4:] if v is not None]
                if any(v for v in action_vals):
                    any_action_marked = True
                    break

            if not any_action_marked:
                score -= 5
                issues.append("デシジョンテーブルの動作行に一つもマーク（X/○）がない")

        # テストケースとの整合（DTのパターンがテストケースの主要機能をカバーしているか）
        if test_cases and action_rows:
            # テストケースの「確認すること」からベース機能名を抽出
            tc_flows = set()
            for tc in test_cases:
                confirm = tc.get("confirm", "")
                base = re.sub(r"^(異常系[：:]?\s*)", "", confirm)
                base = re.sub(r"[（(][\w/]+?(?:のみ)?[）)]", "", base).strip()
                if base:
                    tc_flows.add(base)

            # DTの動作行テキストに機能名が反映されているか
            dt_all_text = " ".join(
                str(dr["values"][1] or "") for dr in data_rows
            ).lower()
            unmatched = 0
            for flow in tc_flows:
                # フロー名の主要部分（2文字以上の単語）で部分一致
                flow_words = [w for w in flow.split() if len(w) >= 2]
                if flow_words and not any(w.lower() in dt_all_text for w in flow_words):
                    unmatched += 1
            if tc_flows and unmatched > len(tc_flows) * 0.5:
                score -= 3
                issues.append(f"テストケースの主要機能{unmatched}件がDTに反映されていない")

    return max(0, score), issues


def evaluate_risk_checklist(perspective_data, test_cases, spec_text):
    """5. リスク観点妥当性（20点）

    評価観点:
    - 観点数の十分さ
    - ○/-の比率バランス（多すぎ/少なすぎ）
    - 仕様→リスク: 仕様の主題に関連する観点が該当になっているか
    - リスク→仕様: 該当○とした観点が仕様に実質的な根拠があるか（偽陽性検出）
    """
    score = 20
    issues = []

    risk_sections = [s for s in perspective_data["sections"] if s["type"] == "risk_checklist"]
    if not risk_sections:
        score -= 10
        issues.append("リスク観点チェックリストが存在しない")
        return max(0, score), issues

    for sec in risk_sections:
        data_rows = sec["data_rows"]

        # ヘッダー行をスキップ
        content_rows = [
            dr for dr in data_rows
            if str(dr["values"][1] or "").strip() not in ("観点", "", "None")
        ]

        if not content_rows:
            score -= 10
            issues.append("リスク観点チェックリストにデータがない")
            continue

        # 観点数チェック（最低5観点）
        if len(content_rows) < 5:
            score -= 3
            issues.append(f"リスク観点が{len(content_rows)}個（5個以上推奨）")

        # --- 各行の観点名と該当判定を抽出 ---
        risk_entries = []
        for dr in content_rows:
            vals = dr["values"]
            name = str(vals[1] or "").strip()
            is_relevant = any(str(v or "").strip() == "○" for v in vals)
            risk_entries.append({"name": name, "relevant": is_relevant, "row": dr["row"]})

        relevant_count = sum(1 for e in risk_entries if e["relevant"])
        total_count = len(risk_entries)

        # --- ○/-の比率バランス ---
        if relevant_count == 0:
            score -= 5
            issues.append("すべてのリスク観点が「該当なし」— 仕様に対して過少")
        elif relevant_count == total_count:
            score -= 5
            issues.append("すべてのリスク観点が「該当あり」— 判定が甘い（全該当はありえない）")
        elif total_count >= 5 and relevant_count / total_count > 0.7:
            score -= 3
            issues.append(
                f"該当率が高すぎる（{relevant_count}/{total_count}={relevant_count*100//total_count}%）"
                f"— 通常は30-60%が適正範囲"
            )

        # --- 仕様テキストとの双方向整合チェック ---
        if spec_text:
            # 各リスク観点に対する「仕様の主題との関連度」を判定
            # キーワードの出現回数と文脈で判定（単なる存在ではなく密度）
            _risk_score, _risk_issues = _check_risk_spec_alignment(
                risk_entries, spec_text, test_cases
            )
            score += _risk_score  # negative or zero
            issues.extend(_risk_issues)

    return max(0, score), issues


def _check_risk_spec_alignment(risk_entries, spec_text, test_cases):
    """リスク観点と仕様書の双方向整合チェック。

    - 仕様→リスク: 仕様の主題に必須の観点が「該当なし」→ 見落とし
    - リスク→仕様: 「該当あり」だが仕様に実質的根拠なし → 偽陽性
    """
    deduction = 0
    issues = []

    spec_lower = spec_text.lower()
    tc_text = ""
    if test_cases:
        tc_text = json.dumps(test_cases, ensure_ascii=False).lower()

    # 各観点の「実質的関連度」判定ルール
    # primary_keywords: 仕様の主題に直結するキーワード（これが多く出れば関連あり）
    # context_keywords: 文脈上出るだけで関連とは言えないキーワード（「保存」「入力」等）
    # min_hits: 関連ありと判定するためのprimary_keywords最低ヒット数
    risk_relevance_rules = {
        "データ整合性": {
            "primary": ["データ保存", "データ更新", "データベース", "整合", "不整合",
                        "保持される", "反映される", "再表示"],
            "context_only": ["保存", "登録", "更新", "データ"],
            "min_hits": 1,
            "description": "データの永続化・整合性が仕様の主題に含まれる場合",
        },
        "セッション管理": {
            "primary": ["セッション管理", "セッション切れ", "セッションタイムアウト",
                        "ログアウト", "session timeout", "session expire"],
            "context_only": ["セッション", "タイムアウト", "有効期限", "期限切れ",
                             "ログイン", "認証"],
            "min_hits": 1,
            "description": "セッション管理自体が仕様の主題に含まれる場合",
        },
        "並行操作": {
            "primary": ["並行", "同時操作", "排他制御", "排他ロック", "競合",
                        "デッドロック", "race condition", "concurrent"],
            "context_only": [],
            "min_hits": 1,
            "description": "複数ユーザー/スレッドの競合が仕様に含まれる場合",
        },
        "入力バリデーション": {
            "primary": ["バリデーション", "入力チェック", "必須チェック", "形式チェック",
                        "文字数制限", "桁数", "validation", "入力規則"],
            "context_only": ["必須", "入力", "チェック", "エラー"],
            "min_hits": 1,
            "description": "入力値の検証ルールが仕様の主題に含まれる場合",
        },
        "エラーハンドリング": {
            "primary": ["エラーハンドリング", "エラー処理", "例外処理", "サーバーエラー",
                        "ネットワーク断", "500エラー", "タイムアウトエラー"],
            "context_only": ["エラー", "失敗", "異常"],
            "min_hits": 1,
            "description": "エラー時の挙動が仕様の主題に含まれる場合",
        },
        "権限制御": {
            "primary": ["権限", "アクセス制御", "ロール", "管理者権限",
                        "操作権限", "認可", "authorization", "permission"],
            "context_only": ["認証", "ログイン", "管理者"],
            "min_hits": 1,
            "description": "権限管理が仕様の主題に含まれる場合",
        },
        "セキュリティ": {
            "primary": ["脆弱性", "セキュリティ", "xss", "csrf", "インジェクション",
                        "リダイレクト", "改ざん", "不正アクセス", "セキュリティ診断",
                        "devsec", "オープンリダイレクタ"],
            "context_only": [],
            "min_hits": 1,
            "description": "セキュリティが仕様の主題に含まれる場合",
        },
        "ブラウザ互換": {
            "primary": ["ブラウザ互換", "クロスブラウザ", "ブラウザ差異",
                        "chrome", "edge", "firefox", "safari", "ie"],
            "context_only": ["ブラウザ"],
            "min_hits": 1,
            "description": "ブラウザ互換性が仕様の主題に含まれる場合",
        },
        "パフォーマンス": {
            "primary": ["パフォーマンス", "レスポンス時間", "負荷", "大量データ",
                        "スロークエリ", "速度改善", "performance", "latency"],
            "context_only": ["速度", "遅い"],
            "min_hits": 1,
            "description": "性能要件が仕様の主題に含まれる場合",
        },
    }

    # 仕様書のみで判定（テストケースは仕様の派生なので含めない。
    # テストケースの期待値に「バリデーション」等があっても、
    # それは仕様の主題がバリデーションであることを意味しない）
    combined_text = spec_lower
    false_positives = []
    false_negatives = []

    for entry in risk_entries:
        name = entry["name"]
        is_relevant = entry["relevant"]
        rule = risk_relevance_rules.get(name)

        if not rule:
            continue

        # primaryキーワードのヒット数で実質的関連度を判定
        primary_hits = sum(1 for kw in rule["primary"] if kw in combined_text)
        actually_relevant = primary_hits >= rule["min_hits"]

        if is_relevant and not actually_relevant:
            # 偽陽性: ○だが仕様に実質的な根拠がない
            false_positives.append(name)
        elif not is_relevant and actually_relevant:
            # 偽陰性: -だが仕様に実質的な関連がある
            false_negatives.append(name)

    # 偽陽性の減点（該当と判定しているが根拠なし）
    if false_positives:
        deduction -= min(8, len(false_positives) * 2)
        for fp in false_positives:
            issues.append(f"偽陽性: 「{fp}」が該当○だが、仕様に実質的な根拠がない")

    # 偽陰性の減点（該当なしだが仕様に関連あり）
    if false_negatives:
        deduction -= min(5, len(false_negatives) * 2)
        for fn in false_negatives:
            issues.append(f"見落とし: 「{fn}」が該当-だが、仕様に関連する記述がある")

    return deduction, issues


# =====================================================================
# 統合評価
# =====================================================================

def evaluate(xlsx_path, spec_path=None, testcases_path=None):
    """観点シートの内容妥当性を総合評価する。

    Returns:
        dict: 評価結果
    """
    # --- データ読み込み ---
    perspective_data = read_perspective_sheet(xlsx_path)
    if perspective_data.get("error"):
        return {
            "file": xlsx_path,
            "score": 0,
            "max_score": 100,
            "error": perspective_data["error"],
            "categories": [],
        }

    # テストケース（Excelから読み取り、またはJSON指定）
    test_cases = []
    if testcases_path:
        with open(testcases_path, "r", encoding="utf-8") as f:
            test_cases = json.load(f)
    else:
        test_cases = read_testcase_sheet(xlsx_path)

    # 仕様書
    spec_text = ""
    if spec_path and os.path.exists(spec_path):
        with open(spec_path, "r", encoding="utf-8") as f:
            spec_text = f.read()

    # --- 5カテゴリ評価 ---
    categories = []

    s1, i1 = evaluate_structure(perspective_data)
    categories.append({"name": "構造妥当性", "score": s1, "max": 20, "issues": i1})

    s2, i2 = evaluate_list_coverage(perspective_data, test_cases)
    categories.append({"name": "リスト網羅性", "score": s2, "max": 20, "issues": i2})

    s3, i3 = evaluate_matrix_consistency(perspective_data, test_cases, spec_text)
    categories.append({"name": "マトリクス整合性", "score": s3, "max": 20, "issues": i3})

    s4, i4 = evaluate_decision_table(perspective_data, test_cases, spec_text)
    categories.append({"name": "デシジョンテーブル論理性", "score": s4, "max": 20, "issues": i4})

    s5, i5 = evaluate_risk_checklist(perspective_data, test_cases, spec_text)
    categories.append({"name": "リスク観点妥当性", "score": s5, "max": 20, "issues": i5})

    total = sum(c["score"] for c in categories)

    return {
        "file": os.path.basename(xlsx_path),
        "score": total,
        "max_score": 100,
        "sections_found": len(perspective_data["sections"]),
        "toc_items": len(perspective_data["toc_items"]),
        "testcases_used": len(test_cases),
        "spec_provided": bool(spec_text),
        "categories": categories,
    }


# =====================================================================
# レポート出力
# =====================================================================

def print_report(result):
    """評価結果をコンソール出力する。"""
    print("=" * 70)
    print("観点シート 内容妥当性 評価レポート")
    print("=" * 70)
    print()
    print(f"ファイル: {result['file']}")
    print(f"検出セクション数: {result.get('sections_found', 0)}")
    print(f"目次項目数: {result.get('toc_items', 0)}")
    print(f"テストケース数: {result.get('testcases_used', 0)}")
    print(f"仕様書指定: {'あり' if result.get('spec_provided') else 'なし'}")
    print()

    if result.get("error"):
        print(f"ERROR: {result['error']}")
        return

    total = result["score"]
    print(f"★ 総合スコア: {total}/100点")
    print()
    print("-" * 70)
    print("カテゴリ別スコア")
    print("-" * 70)

    for cat in result["categories"]:
        status = "OK" if cat["score"] == cat["max"] else "要改善"
        print(f"  {cat['name']}: {cat['score']}/{cat['max']}点 [{status}]")
        for issue in cat["issues"]:
            print(f"    - {issue}")

    print()
    print("-" * 70)

    # 総合コメント
    if total >= 90:
        print("総評: 観点設計の内容は妥当です。")
    elif total >= 70:
        print("総評: 概ね妥当ですが、一部の観点で改善が必要です。")
    elif total >= 50:
        print("総評: 複数の観点で整合性や網羅性に問題があります。見直しを推奨します。")
    else:
        print("総評: 観点設計に大きな問題があります。再設計を推奨します。")


# =====================================================================
# CLI
# =====================================================================

def main():
    parser = argparse.ArgumentParser(description="観点シート 内容妥当性 自動評価ツール")
    parser.add_argument("xlsx", help="評価対象のテスト項目書（Excel）")
    parser.add_argument("--spec", help="仕様書（Markdown等）のパス")
    parser.add_argument("--testcases", help="テストケースJSON（省略時はExcelから読み取り）")
    parser.add_argument("-o", "--output", help="評価結果JSON出力先")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: ファイルが見つかりません: {args.xlsx}")
        sys.exit(1)

    result = evaluate(args.xlsx, args.spec, args.testcases)
    print_report(result)

    # JSON出力
    output_path = args.output
    if not output_path:
        base = os.path.splitext(args.xlsx)[0]
        output_path = f"{base}.perspective_eval.json"

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"\n評価結果: {output_path}")


if __name__ == "__main__":
    main()
