# -*- coding: utf-8 -*-
"""
2つのExcelファイルを比較分析するスクリプト
"""
import openpyxl
import re
import sys
import io
from collections import Counter

# Windows環境でのUTF-8出力対応
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ファイルパス
FILE_A = r"C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\テスト項目書_ログイン方法でID・PWとTOKIUM IDが存在し、TOKIUM IDだけ利用したい場合にフラグでログイン制限できるように (2).xlsx"
FILE_B = r"C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\成果物\テスト項目書_改善後_v7.xlsx"

# 曖昧表現パターン
AMBIGUOUS_PATTERNS = [
    r'正常であること',
    r'問題ないこと',
    r'仕様通り',
    r'適切に',
    r'正しく',
    r'期待通り',
    r'想定通り',
]

# 良好末尾表現
GOOD_ENDINGS = [
    r'されること$',
    r'されていること$',
    r'であること$',
    r'できること$',
    r'ないこと$',
    r'なること$',
    r'いること$',
    r'れること$',
    r'すること$',
    r'あること$',
]


def load_workbook_info(filepath, label):
    """Excelファイルの基本情報を取得"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\n{'='*80}")
    print(f"【{label}】")
    print(f"ファイル: {filepath.split(chr(92))[-1]}")
    print(f"{'='*80}")

    print(f"\n--- シート名一覧 ---")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"  {i}. 「{name}」 - {ws.max_row}行 x {ws.max_column}列")

    return wb


def find_header_and_data(ws):
    """ヘッダー行を検出し、データ開始行を特定"""
    header_row = None
    headers = {}

    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_vals = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_vals.append(str(val).strip())

        row_text = ' '.join(row_vals).lower()
        # ヘッダー行の検出: 期待値、確認内容、テスト項目などのキーワード
        if any(kw in row_text for kw in ['期待値', '確認内容', 'テスト項目', '手順', 'no', 'id']):
            header_row = row_idx
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    headers[col_idx] = str(val).strip()
            break

    return header_row, headers


def get_column_index(headers, keywords):
    """キーワードに一致する列インデックスを返す"""
    for col_idx, header_name in headers.items():
        for kw in keywords:
            if kw in header_name:
                return col_idx
    return None


def extract_testcases(ws, header_row, headers):
    """テストケースデータを抽出"""
    # 列のマッピング
    no_col = get_column_index(headers, ['No', 'NO', 'no', '#', 'ID', 'id'])
    expected_col = get_column_index(headers, ['期待値', '確認内容', '期待結果', '想定結果'])
    procedure_col = get_column_index(headers, ['手順', '操作手順', 'テスト手順', '確認手順', '操作'])
    group_col = get_column_index(headers, ['グループ', '大項目', 'カテゴリ', '機能', '分類'])
    item_col = get_column_index(headers, ['テスト項目', '項目', '中項目', '確認項目', 'テスト観点'])
    condition_col = get_column_index(headers, ['条件', '前提条件', '前提', 'テスト条件'])

    testcases = []
    data_start = header_row + 1

    current_group = ""
    current_item = ""
    current_procedure = ""

    for row_idx in range(data_start, ws.max_row + 1):
        # 期待値がある行をテストケースとみなす
        expected_val = ws.cell(row=row_idx, column=expected_col).value if expected_col else None

        # グループ情報の更新（結合セル対応）
        if group_col:
            g = ws.cell(row=row_idx, column=group_col).value
            if g is not None and str(g).strip():
                current_group = str(g).strip()

        if item_col:
            it = ws.cell(row=row_idx, column=item_col).value
            if it is not None and str(it).strip():
                current_item = str(it).strip()

        # 手順の取得
        proc_val = None
        if procedure_col:
            p = ws.cell(row=row_idx, column=procedure_col).value
            if p is not None and str(p).strip():
                current_procedure = str(p).strip()
                proc_val = current_procedure
            else:
                proc_val = current_procedure if current_procedure else None

        if expected_val is not None and str(expected_val).strip():
            expected_str = str(expected_val).strip()

            # No取得
            no_val = ""
            if no_col:
                n = ws.cell(row=row_idx, column=no_col).value
                if n is not None:
                    no_val = str(n).strip()

            # 条件取得
            cond_val = ""
            if condition_col:
                c = ws.cell(row=row_idx, column=condition_col).value
                if c is not None:
                    cond_val = str(c).strip()

            tc = {
                'row': row_idx,
                'no': no_val,
                'group': current_group,
                'item': current_item,
                'procedure': proc_val if proc_val else "",
                'expected': expected_str,
                'condition': cond_val,
            }
            testcases.append(tc)

    return testcases, {
        'no_col': no_col,
        'expected_col': expected_col,
        'procedure_col': procedure_col,
        'group_col': group_col,
        'item_col': item_col,
        'condition_col': condition_col,
    }


def analyze_testcases(testcases, label):
    """テストケースの統計情報を分析"""
    total = len(testcases)
    if total == 0:
        print(f"\n  テストケース数: 0件")
        return {}

    # 期待値の文字数
    expected_lengths = [len(tc['expected']) for tc in testcases]
    avg_len = sum(expected_lengths) / total
    min_len = min(expected_lengths)
    max_len = max(expected_lengths)

    # 手順記載率
    has_procedure = sum(1 for tc in testcases if tc['procedure'] and tc['procedure'].strip())
    procedure_rate = has_procedure / total * 100

    # 「こと」の回数分析
    koto_counts = []
    for tc in testcases:
        count = tc['expected'].count('こと')
        koto_counts.append(count)

    koto_1 = sum(1 for c in koto_counts if c == 1)
    koto_2plus = sum(1 for c in koto_counts if c >= 2)
    koto_0 = sum(1 for c in koto_counts if c == 0)

    # 曖昧表現チェック
    ambiguous_count = 0
    ambiguous_examples = []
    for tc in testcases:
        for pat in AMBIGUOUS_PATTERNS:
            if re.search(pat, tc['expected']):
                ambiguous_count += 1
                if len(ambiguous_examples) < 3:
                    ambiguous_examples.append((tc['no'], tc['expected'][:60], pat))
                break

    # 良好末尾表現チェック
    good_ending_count = 0
    for tc in testcases:
        for pat in GOOD_ENDINGS:
            if re.search(pat, tc['expected']):
                good_ending_count += 1
                break

    # 番号付き手順の割合
    numbered_procedure = 0
    for tc in testcases:
        if tc['procedure']:
            proc = tc['procedure'].strip()
            if re.search(r'(^|\n)\s*(1[\.\)）]|①|手順\s*1|Step\s*1)', proc):
                numbered_procedure += 1

    numbered_rate = numbered_procedure / has_procedure * 100 if has_procedure > 0 else 0

    # 期待値の長さ分布
    short_count = sum(1 for l in expected_lengths if l < 10)
    long_count = sum(1 for l in expected_lengths if l > 80)
    normal_count = sum(1 for l in expected_lengths if 10 <= l <= 80)

    stats = {
        'total': total,
        'avg_len': avg_len,
        'min_len': min_len,
        'max_len': max_len,
        'procedure_rate': procedure_rate,
        'has_procedure': has_procedure,
        'koto_1_rate': koto_1 / total * 100,
        'koto_2plus_rate': koto_2plus / total * 100,
        'koto_0_rate': koto_0 / total * 100,
        'ambiguous_rate': ambiguous_count / total * 100,
        'ambiguous_count': ambiguous_count,
        'good_ending_rate': good_ending_count / total * 100,
        'numbered_rate': numbered_rate,
        'short_count': short_count,
        'long_count': long_count,
        'normal_count': normal_count,
    }

    print(f"\n--- テストケース統計 ---")
    print(f"  総テストケース数: {total}件")
    print(f"")
    print(f"  【期待値の文字数】")
    print(f"    平均: {avg_len:.1f}字")
    print(f"    最小: {min_len}字 / 最大: {max_len}字")
    print(f"    適正範囲(10-80字): {normal_count}件 ({normal_count/total*100:.1f}%)")
    print(f"    短すぎ(<10字): {short_count}件 ({short_count/total*100:.1f}%)")
    print(f"    長すぎ(>80字): {long_count}件 ({long_count/total*100:.1f}%)")
    print(f"")
    print(f"  【手順記載率】")
    print(f"    手順あり: {has_procedure}件 / {total}件 ({procedure_rate:.1f}%)")
    print(f"    番号付き手順: {numbered_procedure}件 / {has_procedure}件 ({numbered_rate:.1f}%)")
    print(f"")
    print(f"  【「こと」の出現回数】")
    print(f"    0回: {koto_0}件 ({koto_0/total*100:.1f}%)")
    print(f"    1回（適正）: {koto_1}件 ({koto_1/total*100:.1f}%)")
    print(f"    2回以上（複数混在）: {koto_2plus}件 ({koto_2plus/total*100:.1f}%)")
    print(f"")
    print(f"  【曖昧表現】")
    print(f"    曖昧表現を含むケース: {ambiguous_count}件 ({ambiguous_count/total*100:.1f}%)")
    if ambiguous_examples:
        for no, exp, pat in ambiguous_examples:
            print(f"      例: No.{no} 「{exp}...」 (検出: {pat})")
    print(f"")
    print(f"  【良好末尾表現】")
    print(f"    良好な末尾表現: {good_ending_count}件 ({good_ending_count/total*100:.1f}%)")

    return stats


def compare_testcases(tc_a, tc_b, label_a, label_b):
    """2つのテストケース群を比較"""
    print(f"\n{'='*80}")
    print(f"【差分比較】{label_a} vs {label_b}")
    print(f"{'='*80}")

    print(f"\n--- ケース数の増減 ---")
    diff = len(tc_a) - len(tc_b)
    print(f"  {label_a}: {len(tc_a)}件")
    print(f"  {label_b}: {len(tc_b)}件")
    if diff > 0:
        print(f"  差分: {label_a}が {diff}件 多い")
    elif diff < 0:
        print(f"  差分: {label_b}が {-diff}件 多い")
    else:
        print(f"  差分: 同数")

    # 期待値をキーにした比較
    expected_a = {tc['expected']: tc for tc in tc_a}
    expected_b = {tc['expected']: tc for tc in tc_b}

    only_in_a = set(expected_a.keys()) - set(expected_b.keys())
    only_in_b = set(expected_b.keys()) - set(expected_a.keys())
    common = set(expected_a.keys()) & set(expected_b.keys())

    print(f"\n--- 期待値ベースの比較 ---")
    print(f"  完全一致する期待値: {len(common)}件")
    print(f"  {label_a}にのみ存在: {len(only_in_a)}件")
    print(f"  {label_b}にのみ存在: {len(only_in_b)}件")

    # Noベースの比較（より正確な対応付け）
    no_map_a = {}
    no_map_b = {}
    for tc in tc_a:
        if tc['no']:
            no_map_a[tc['no']] = tc
    for tc in tc_b:
        if tc['no']:
            no_map_b[tc['no']] = tc

    # 同じNoで期待値が異なるケース
    changed_expected = []
    changed_procedure = []

    common_nos = set(no_map_a.keys()) & set(no_map_b.keys())
    for no in sorted(common_nos, key=lambda x: int(x) if x.isdigit() else float('inf')):
        tc_from_a = no_map_a[no]
        tc_from_b = no_map_b[no]

        if tc_from_a['expected'] != tc_from_b['expected']:
            changed_expected.append((no, tc_from_a['expected'], tc_from_b['expected']))

        if tc_from_a['procedure'] != tc_from_b['procedure']:
            if tc_from_a['procedure'] or tc_from_b['procedure']:
                changed_procedure.append((no, tc_from_a['procedure'], tc_from_b['procedure']))

    print(f"\n--- No.ベースの比較 ---")
    print(f"  {label_a}のNo付きケース: {len(no_map_a)}件")
    print(f"  {label_b}のNo付きケース: {len(no_map_b)}件")
    print(f"  共通No: {len(common_nos)}件")
    only_no_a = set(no_map_a.keys()) - set(no_map_b.keys())
    only_no_b = set(no_map_b.keys()) - set(no_map_a.keys())
    if only_no_a:
        print(f"  {label_a}にのみ存在するNo: {sorted(only_no_a, key=lambda x: int(x) if x.isdigit() else float('inf'))[:20]}")
    if only_no_b:
        print(f"  {label_b}にのみ存在するNo: {sorted(only_no_b, key=lambda x: int(x) if x.isdigit() else float('inf'))[:20]}")

    # 期待値の変更例
    print(f"\n--- 期待値が変更されたケース: {len(changed_expected)}件 ---")
    if changed_expected:
        show_count = min(5, len(changed_expected))
        print(f"  (代表例 {show_count}件)")
        for i, (no, exp_a, exp_b) in enumerate(changed_expected[:show_count]):
            print(f"\n  【No.{no}】")
            print(f"    {label_a}: {exp_a[:100]}")
            print(f"    {label_b}: {exp_b[:100]}")
    else:
        print(f"  変更なし")

    # 手順の変更例
    print(f"\n--- 手順が変更されたケース: {len(changed_procedure)}件 ---")
    if changed_procedure:
        show_count = min(5, len(changed_procedure))
        print(f"  (代表例 {show_count}件)")
        for i, (no, proc_a, proc_b) in enumerate(changed_procedure[:show_count]):
            print(f"\n  【No.{no}】")
            print(f"    {label_a}: {proc_a[:100] if proc_a else '(空)'}")
            print(f"    {label_b}: {proc_b[:100] if proc_b else '(空)'}")
    else:
        print(f"  変更なし")

    # 新規追加ケースの例
    if only_in_a:
        print(f"\n--- {label_a}にのみ存在するケース（代表例 最大5件）---")
        for i, exp in enumerate(list(only_in_a)[:5]):
            tc = expected_a[exp]
            print(f"  {i+1}. No.{tc['no']} | {exp[:80]}")

    if only_in_b:
        print(f"\n--- {label_b}にのみ存在するケース（代表例 最大5件）---")
        for i, exp in enumerate(list(only_in_b)[:5]):
            tc = expected_b[exp]
            print(f"  {i+1}. No.{tc['no']} | {exp[:80]}")


def print_headers(ws, sheet_name):
    """ヘッダー行の内容を表示"""
    header_row, headers = find_header_and_data(ws)
    if header_row:
        print(f"\n  シート「{sheet_name}」のヘッダー行（{header_row}行目）:")
        for col_idx in sorted(headers.keys()):
            print(f"    列{col_idx}: {headers[col_idx]}")
    else:
        print(f"\n  シート「{sheet_name}」: ヘッダー行が検出できませんでした")
        # 最初の数行を表示
        print(f"    先頭5行の内容:")
        for r in range(1, min(6, ws.max_row + 1)):
            vals = []
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    vals.append(f"[{c}]{str(v)[:30]}")
            if vals:
                print(f"      {r}行目: {', '.join(vals)}")
    return header_row, headers


def main():
    print("=" * 80)
    print("テスト項目書 比較分析レポート")
    print("=" * 80)

    label_a = "参考例（ユーザー改善版）"
    label_b = "成果物v7（AI改善版）"

    # ファイルA: 参考例
    wb_a = load_workbook_info(FILE_A, label_a)
    all_tc_a = []
    target_ws_a = None
    target_headers_a = None

    for sheet_name in wb_a.sheetnames:
        ws = wb_a[sheet_name]
        header_row, headers = print_headers(ws, sheet_name)
        if header_row and headers:
            expected_col = get_column_index(headers, ['期待値', '確認内容', '期待結果'])
            if expected_col:
                testcases, col_map = extract_testcases(ws, header_row, headers)
                if testcases:
                    target_ws_a = ws
                    target_headers_a = headers
                    all_tc_a = testcases
                    print(f"\n  -> テストケース検出: {len(testcases)}件（シート「{sheet_name}」）")

    if all_tc_a:
        analyze_testcases(all_tc_a, label_a)

    # ファイルB: 成果物v7
    wb_b = load_workbook_info(FILE_B, label_b)
    all_tc_b = []

    for sheet_name in wb_b.sheetnames:
        ws = wb_b[sheet_name]
        header_row, headers = print_headers(ws, sheet_name)
        if header_row and headers:
            expected_col = get_column_index(headers, ['期待値', '確認内容', '期待結果'])
            if expected_col:
                testcases, col_map = extract_testcases(ws, header_row, headers)
                if testcases:
                    all_tc_b = testcases
                    print(f"\n  -> テストケース検出: {len(testcases)}件（シート「{sheet_name}」）")

    if all_tc_b:
        analyze_testcases(all_tc_b, label_b)

    # 比較
    if all_tc_a and all_tc_b:
        compare_testcases(all_tc_a, all_tc_b, label_a, label_b)

    # 統計比較サマリー
    if all_tc_a and all_tc_b:
        stats_a = analyze_testcases.__code__ and None  # 再度計算用に
        print(f"\n{'='*80}")
        print(f"【統計比較サマリー】")
        print(f"{'='*80}")

        def calc_stats(tcs):
            total = len(tcs)
            avg_len = sum(len(tc['expected']) for tc in tcs) / total
            has_proc = sum(1 for tc in tcs if tc['procedure'] and tc['procedure'].strip())
            koto_1 = sum(1 for tc in tcs if tc['expected'].count('こと') == 1)
            koto_2plus = sum(1 for tc in tcs if tc['expected'].count('こと') >= 2)
            ambig = sum(1 for tc in tcs if any(re.search(p, tc['expected']) for p in AMBIGUOUS_PATTERNS))
            good_end = sum(1 for tc in tcs if any(re.search(p, tc['expected']) for p in GOOD_ENDINGS))
            numbered = sum(1 for tc in tcs if tc['procedure'] and re.search(r'(^|\n)\s*(1[\.\)）]|①|手順\s*1|Step\s*1)', tc['procedure'].strip()))
            return {
                'total': total,
                'avg_len': avg_len,
                'proc_rate': has_proc/total*100,
                'koto_1_rate': koto_1/total*100,
                'koto_2plus_rate': koto_2plus/total*100,
                'ambig_rate': ambig/total*100,
                'good_end_rate': good_end/total*100,
                'numbered_rate': numbered/has_proc*100 if has_proc else 0,
            }

        sa = calc_stats(all_tc_a)
        sb = calc_stats(all_tc_b)

        fmt = "{:<30s} {:>15s} {:>15s} {:>10s}"
        print(fmt.format("指標", label_a[:12], label_b[:12], "差"))
        print("-" * 75)
        print(fmt.format("テストケース数", f"{sa['total']}件", f"{sb['total']}件", f"{sa['total']-sb['total']:+d}"))
        print(fmt.format("期待値平均文字数", f"{sa['avg_len']:.1f}字", f"{sb['avg_len']:.1f}字", f"{sa['avg_len']-sb['avg_len']:+.1f}"))
        print(fmt.format("手順記載率", f"{sa['proc_rate']:.1f}%", f"{sb['proc_rate']:.1f}%", f"{sa['proc_rate']-sb['proc_rate']:+.1f}"))
        print(fmt.format("「こと」1回率", f"{sa['koto_1_rate']:.1f}%", f"{sb['koto_1_rate']:.1f}%", f"{sa['koto_1_rate']-sb['koto_1_rate']:+.1f}"))
        print(fmt.format("「こと」2回以上率", f"{sa['koto_2plus_rate']:.1f}%", f"{sb['koto_2plus_rate']:.1f}%", f"{sa['koto_2plus_rate']-sb['koto_2plus_rate']:+.1f}"))
        print(fmt.format("曖昧表現率", f"{sa['ambig_rate']:.1f}%", f"{sb['ambig_rate']:.1f}%", f"{sa['ambig_rate']-sb['ambig_rate']:+.1f}"))
        print(fmt.format("良好末尾表現率", f"{sa['good_end_rate']:.1f}%", f"{sb['good_end_rate']:.1f}%", f"{sa['good_end_rate']-sb['good_end_rate']:+.1f}"))
        print(fmt.format("番号付き手順率", f"{sa['numbered_rate']:.1f}%", f"{sb['numbered_rate']:.1f}%", f"{sa['numbered_rate']-sb['numbered_rate']:+.1f}"))

    print(f"\n\n分析完了。")


if __name__ == '__main__':
    main()
