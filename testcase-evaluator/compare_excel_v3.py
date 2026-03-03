# -*- coding: utf-8 -*-
"""
テスト項目書 比較分析スクリプト v3
参考例（ユーザー改善版） vs 成果物v7（AI改善版）
元Noベースの正確な対応付けによる比較
"""
import openpyxl
import re
import sys
import io
import warnings
from collections import defaultdict

warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

FILE_A = r"C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\参考例\テスト項目書_ログイン方法でID・PWとTOKIUM IDが存在し、TOKIUM IDだけ利用したい場合にフラグでログイン制限できるように (2).xlsx"
FILE_B = r"C:\Users\池田尚人\ClaudeCode用\テストケース評価自動化\成果物\テスト項目書_改善後_v7.xlsx"

AMBIGUOUS_PATTERNS = [
    r'正常であること', r'問題ないこと', r'仕様通り', r'適切に',
    r'正しく', r'期待通り', r'想定通り',
]
GOOD_ENDINGS = [
    r'されること\s*$', r'されていること\s*$', r'であること\s*$',
    r'できること\s*$', r'ないこと\s*$', r'なること\s*$',
    r'いること\s*$', r'れること\s*$', r'すること\s*$', r'あること\s*$',
]


def cell_str(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def extract_testcases(ws, sheet_name, orig_no_col=None):
    """テストケースを抽出。orig_no_col指定時は元No列も取得"""
    testcases = []
    current_group = ""
    current_procedure = ""

    for row_idx in range(3, ws.max_row + 1):
        expected = cell_str(ws, row_idx, 8)
        if not expected:
            continue

        no = cell_str(ws, row_idx, 2)
        group = cell_str(ws, row_idx, 3)
        target = cell_str(ws, row_idx, 5)
        detail = cell_str(ws, row_idx, 6)
        procedure = cell_str(ws, row_idx, 7)

        if group:
            current_group = group
        if procedure and procedure != "-":
            current_procedure = procedure

        orig_no = ""
        if orig_no_col:
            orig_no = cell_str(ws, row_idx, orig_no_col)

        tc = {
            'row': row_idx,
            'sheet': sheet_name,
            'no': no,
            'group': current_group,
            'target': target,
            'detail': detail,
            'procedure': procedure if (procedure and procedure != "-") else "",
            'procedure_inherited': current_procedure if not (procedure and procedure != "-") else procedure,
            'expected': expected,
            'orig_no': orig_no,
        }
        testcases.append(tc)
    return testcases


def calc_stats(testcases):
    total = len(testcases)
    if total == 0:
        return {}

    expected_lengths = [len(tc['expected']) for tc in testcases]
    avg_len = sum(expected_lengths) / total

    has_proc_direct = sum(1 for tc in testcases if tc['procedure'])
    has_proc_inherited = sum(1 for tc in testcases if tc['procedure_inherited'])

    koto_0 = sum(1 for tc in testcases if tc['expected'].count('こと') == 0)
    koto_1 = sum(1 for tc in testcases if tc['expected'].count('こと') == 1)
    koto_2plus = sum(1 for tc in testcases if tc['expected'].count('こと') >= 2)

    ambig = sum(1 for tc in testcases if any(re.search(p, tc['expected']) for p in AMBIGUOUS_PATTERNS))
    good_end = sum(1 for tc in testcases if any(re.search(p, tc['expected']) for p in GOOD_ENDINGS))

    numbered = 0
    for tc in testcases:
        proc = tc['procedure_inherited']
        if proc and re.search(r'(^|\n)\s*([\d]+[\.\)）]|①|【手順|手順\s*[\d]|Step\s*[\d]|【セットアップ|【前提)', proc):
            numbered += 1

    short = sum(1 for l in expected_lengths if l < 10)
    long_ = sum(1 for l in expected_lengths if l > 80)
    normal = sum(1 for l in expected_lengths if 10 <= l <= 80)

    return {
        'total': total,
        'avg_len': avg_len,
        'min_len': min(expected_lengths),
        'max_len': max(expected_lengths),
        'short': short, 'long': long_, 'normal': normal,
        'proc_direct': has_proc_direct,
        'proc_inherited': has_proc_inherited,
        'proc_direct_rate': has_proc_direct / total * 100,
        'proc_inherited_rate': has_proc_inherited / total * 100,
        'numbered': numbered,
        'numbered_rate': numbered / has_proc_inherited * 100 if has_proc_inherited else 0,
        'koto_0': koto_0, 'koto_1': koto_1, 'koto_2plus': koto_2plus,
        'koto_1_rate': koto_1 / total * 100,
        'koto_2plus_rate': koto_2plus / total * 100,
        'ambig': ambig, 'ambig_rate': ambig / total * 100,
        'good_end': good_end, 'good_end_rate': good_end / total * 100,
    }


def print_stats(stats, label):
    s = stats
    print(f"\n--- テストケース統計 [{label}] ---")
    print(f"  総テストケース数: {s['total']}件")
    print(f"")
    print(f"  [期待値の文字数]")
    print(f"    平均: {s['avg_len']:.1f}字 / 最小: {s['min_len']}字 / 最大: {s['max_len']}字")
    print(f"    適正範囲(10-80字): {s['normal']}件 ({s['normal']/s['total']*100:.1f}%)")
    print(f"    短すぎ(<10字): {s['short']}件 / 長すぎ(>80字): {s['long']}件")
    print(f"")
    print(f"  [手順記載率]")
    print(f"    直接記載: {s['proc_direct']}件/{s['total']}件 ({s['proc_direct_rate']:.1f}%)")
    print(f"    継承含む: {s['proc_inherited']}件/{s['total']}件 ({s['proc_inherited_rate']:.1f}%)")
    print(f"    番号付き/構造化手順: {s['numbered']}件 ({s['numbered_rate']:.1f}%)")
    print(f"")
    print(f"  [「こと」の出現回数]")
    print(f"    0回: {s['koto_0']}件 ({s['koto_0']/s['total']*100:.1f}%)")
    print(f"    1回（適正）: {s['koto_1']}件 ({s['koto_1_rate']:.1f}%)")
    print(f"    2回以上（複数混在）: {s['koto_2plus']}件 ({s['koto_2plus_rate']:.1f}%)")
    print(f"")
    print(f"  [曖昧表現]")
    print(f"    曖昧表現を含む: {s['ambig']}件 ({s['ambig_rate']:.1f}%)")
    print(f"")
    print(f"  [良好末尾表現]")
    print(f"    良好な末尾: {s['good_end']}件 ({s['good_end_rate']:.1f}%)")


def main():
    LABEL_A = "参考例（ユーザー改善版）"
    LABEL_B = "成果物v7（AI改善版）"

    print("=" * 80)
    print("テスト項目書 比較分析レポート")
    print("=" * 80)

    # ============================================================
    # ファイルA: 参考例
    # ============================================================
    wb_a = openpyxl.load_workbook(FILE_A, data_only=True)
    print(f"\n{'='*80}")
    print(f"【{LABEL_A}】")
    print(f"ファイル: ...{FILE_A[-80:]}")
    print(f"{'='*80}")

    print(f"\n--- シート名一覧 ---")
    for i, name in enumerate(wb_a.sheetnames, 1):
        ws = wb_a[name]
        print(f"  {i}. 「{name}」 ({ws.max_row}行 x {ws.max_column}列)")

    tc_sheets_a = ['テストケース_目次1', 'テストケース_目次1続き', 'テストケース_目次2', 'テストケース_目次3']
    print(f"\n--- ヘッダー行（2行目）[テストケース_目次1] ---")
    ws_h = wb_a['テストケース_目次1']
    for c in range(2, 21):
        v = cell_str(ws_h, 2, c)
        if v:
            print(f"    列{c}: {v}")

    all_tc_a = []
    for sname in tc_sheets_a:
        ws = wb_a[sname]
        tcs = extract_testcases(ws, sname)
        print(f"  シート「{sname}」: {len(tcs)}件")
        all_tc_a.extend(tcs)

    print(f"\n  合計: {len(all_tc_a)}件")

    # 参考例_目次1のみ（v7の元No対応用）
    tc_a_sheet1 = [tc for tc in all_tc_a if tc['sheet'] == 'テストケース_目次1']
    # 参考例_目次1以外（追加シート）
    tc_a_additional = [tc for tc in all_tc_a if tc['sheet'] != 'テストケース_目次1']

    stats_a = calc_stats(all_tc_a)
    print_stats(stats_a, LABEL_A)

    # 「こと」2回以上の例
    koto2_examples = [tc for tc in all_tc_a if tc['expected'].count('こと') >= 2]
    if koto2_examples:
        print(f"\n    -- 「こと」2回以上の例 --")
        for tc in koto2_examples[:3]:
            print(f"      No.{tc['no']} [{tc['sheet']}]: {tc['expected'][:80]}")

    # ============================================================
    # ファイルB: 成果物v7
    # ============================================================
    wb_b = openpyxl.load_workbook(FILE_B, data_only=True)
    print(f"\n\n{'='*80}")
    print(f"【{LABEL_B}】")
    print(f"ファイル: {FILE_B.split(chr(92))[-1]}")
    print(f"{'='*80}")

    print(f"\n--- シート名一覧 ---")
    for i, name in enumerate(wb_b.sheetnames, 1):
        ws = wb_b[name]
        print(f"  {i}. 「{name}」 ({ws.max_row}行 x {ws.max_column}列)")

    ws_b = wb_b['テストケース_改善後']
    print(f"\n--- ヘッダー行（2行目）[テストケース_改善後] ---")
    for c in range(2, 22):
        v = cell_str(ws_b, 2, c)
        if v:
            print(f"    列{c}: {v}")

    all_tc_b = extract_testcases(ws_b, 'テストケース_改善後', orig_no_col=10)
    print(f"\n  テストケース数: {len(all_tc_b)}件")

    stats_b = calc_stats(all_tc_b)
    print_stats(stats_b, LABEL_B)

    # 曖昧表現の例
    ambig_examples = [tc for tc in all_tc_b if any(re.search(p, tc['expected']) for p in AMBIGUOUS_PATTERNS)]
    if ambig_examples:
        print(f"\n    -- 曖昧表現の例 --")
        for tc in ambig_examples[:3]:
            matched = [p for p in AMBIGUOUS_PATTERNS if re.search(p, tc['expected'])]
            print(f"      No.{tc['no']}: 「{tc['expected'][:80]}」 (検出: {matched[0] if matched else '-'})")

    # ============================================================
    # 差分比較
    # ============================================================
    print(f"\n\n{'='*80}")
    print(f"【差分比較】")
    print(f"{'='*80}")

    # --- ケース数 ---
    print(f"\n--- 1. ケース数の増減 ---")
    print(f"  {LABEL_A}: {len(all_tc_a)}件（目次1: {len(tc_a_sheet1)}, 追加シート: {len(tc_a_additional)}）")
    print(f"  {LABEL_B}: {len(all_tc_b)}件（1シートに統合）")
    diff = len(all_tc_a) - len(all_tc_b)
    if diff > 0:
        print(f"  -> 参考例が {diff}件 多い")
    elif diff < 0:
        print(f"  -> v7が {abs(diff)}件 多い")
    else:
        print(f"  -> 同数")

    # --- 構造の違い ---
    print(f"\n--- 2. ファイル構造の違い ---")
    print(f"  {LABEL_A}:")
    print(f"    テストケースシート数: {len(tc_sheets_a)}シート（目次1, 目次1続き, 目次2, 目次3）")
    print(f"    「テストケース_改善前」シートも存在（改善前バージョンを保持）")
    print(f"    列10: Autify（テスト自動化ツール用列）")
    print(f"  {LABEL_B}:")
    print(f"    テストケースシート数: 1シート（テストケース_改善後）に統合")
    print(f"    「テストケース_改善前」と「テストケース_改善前 のコピー」を保持")
    print(f"    列10: 元No.（元のテストケース番号の追跡列）")

    # --- 元Noベースの対応付け ---
    print(f"\n--- 3. 元No.ベースの対応付け ---")

    # v7の元Noでグループ化
    v7_by_orig = defaultdict(list)
    for tc in all_tc_b:
        if tc['orig_no']:
            v7_by_orig[tc['orig_no']].append(tc)

    # 参考例_目次1のNoでマップ
    a_by_no = {}
    for tc in tc_a_sheet1:
        if tc['no']:
            if tc['no'] not in a_by_no:
                a_by_no[tc['no']] = []
            a_by_no[tc['no']].append(tc)

    # 数値Noのマッチング
    numeric_orig_nos = [k for k in v7_by_orig.keys() if k.isdigit()]
    new_orig_nos = [k for k in v7_by_orig.keys() if not k.isdigit()]

    matched_a_nos = set()
    matched_count = 0
    split_cases = []  # 元No1つ -> v7複数ケースに分割された例

    for orig_no in sorted(numeric_orig_nos, key=int):
        v7_tcs = v7_by_orig[orig_no]
        a_tcs = a_by_no.get(orig_no, [])
        if a_tcs:
            matched_a_nos.add(orig_no)
            matched_count += 1
            if len(v7_tcs) > len(a_tcs):
                split_cases.append((orig_no, len(a_tcs), len(v7_tcs)))

    unmatched_a = [no for no in a_by_no.keys() if no not in matched_a_nos]

    print(f"  v7の元No（数値）: {len(numeric_orig_nos)}種類 -> 参考例_目次1とマッチ: {matched_count}種類")
    print(f"  v7の元No（NEW-x）: {len(new_orig_nos)}種類 -> v7で新規追加")
    if new_orig_nos:
        for nn in sorted(new_orig_nos):
            tcs = v7_by_orig[nn]
            print(f"    {nn}: {len(tcs)}件 - 「{tcs[0]['expected'][:50]}」")

    if unmatched_a:
        print(f"  参考例_目次1でv7に対応なし: {unmatched_a}")

    # 分割されたケース
    print(f"\n--- 4. ケース分割の状況（元No.1つ -> v7で複数ケースに分割）---")
    total_split_from = 0
    total_split_to = 0
    for orig_no, a_count, b_count in split_cases:
        total_split_from += a_count
        total_split_to += b_count

    print(f"  分割が発生した元No数: {len(split_cases)}種類")
    if split_cases:
        print(f"  分割前合計: {total_split_from}件 -> 分割後合計: {total_split_to}件")
        print(f"  代表例（参考例のケース数 -> v7のケース数）:")
        for orig_no, a_count, b_count in split_cases[:10]:
            a_tcs = a_by_no[orig_no]
            v7_tcs = v7_by_orig[orig_no]
            print(f"    元No.{orig_no}: {a_count}件 -> {b_count}件")
            print(f"      参考例: {', '.join(tc['expected'][:40] for tc in a_tcs)}")
            print(f"      v7:     {', '.join(tc['expected'][:40] for tc in v7_tcs)}")

    # --- 期待値の変更例（同じ元Noで期待値が異なるもの）---
    print(f"\n--- 5. 期待値の変更例（元Noが同じで期待値が異なるケース）---")
    changed_examples = []
    for orig_no in sorted(numeric_orig_nos, key=int):
        a_tcs = a_by_no.get(orig_no, [])
        v7_tcs = v7_by_orig[orig_no]
        if not a_tcs:
            continue

        a_expected_set = set(tc['expected'] for tc in a_tcs)
        v7_expected_set = set(tc['expected'] for tc in v7_tcs)

        only_in_a = a_expected_set - v7_expected_set
        only_in_v7 = v7_expected_set - a_expected_set

        if only_in_a or only_in_v7:
            changed_examples.append((orig_no, a_tcs, v7_tcs, only_in_a, only_in_v7))

    print(f"  期待値に差異がある元No: {len(changed_examples)}種類 / {len(numeric_orig_nos)}種類")

    # 代表例5件
    if changed_examples:
        print(f"\n  代表例（5件）:")
        for i, (orig_no, a_tcs, v7_tcs, only_a, only_v7) in enumerate(changed_examples[:5]):
            print(f"\n  [{i+1}] 元No.{orig_no}")
            print(f"    参考例の期待値 ({len(a_tcs)}件):")
            for tc in a_tcs:
                marker = " [一致]" if tc['expected'] not in only_a else " [変更/削除]"
                print(f"      - {tc['expected'][:80]}{marker}")
            print(f"    v7の期待値 ({len(v7_tcs)}件):")
            for tc in v7_tcs:
                marker = " [一致]" if tc['expected'] not in only_v7 else " [新規/変更]"
                print(f"      - {tc['expected'][:80]}{marker}")

    # --- 手順の変更例 ---
    print(f"\n--- 6. 手順の変更例（元Noが同じで手順が異なるケース）---")
    proc_changed = []
    for orig_no in sorted(numeric_orig_nos, key=int):
        a_tcs = a_by_no.get(orig_no, [])
        v7_tcs = v7_by_orig[orig_no]
        if not a_tcs:
            continue

        # グループの代表手順を比較（最初の直接記載手順同士）
        a_proc = ""
        for tc in a_tcs:
            if tc['procedure']:
                a_proc = tc['procedure']
                break
        v7_proc = ""
        for tc in v7_tcs:
            if tc['procedure']:
                v7_proc = tc['procedure']
                break

        if a_proc and v7_proc and a_proc != v7_proc:
            proc_changed.append((orig_no, a_proc, v7_proc))

    print(f"  手順に差異がある元No: {len(proc_changed)}種類")
    if proc_changed:
        print(f"\n  代表例（5件）:")
        for i, (orig_no, a_proc, v7_proc) in enumerate(proc_changed[:5]):
            a_show = a_proc.replace('\n', ' | ')[:100]
            v7_show = v7_proc.replace('\n', ' | ')[:100]
            print(f"\n  [{i+1}] 元No.{orig_no}")
            print(f"    参考例: {a_show}")
            print(f"    v7    : {v7_show}")

    # --- 参考例の追加シートの内容 ---
    print(f"\n--- 7. 参考例で追加されたシート（v7にないテストケース群）---")
    for sname in ['テストケース_目次1続き', 'テストケース_目次2', 'テストケース_目次3']:
        tcs = [tc for tc in all_tc_a if tc['sheet'] == sname]
        print(f"\n  シート「{sname}」: {len(tcs)}件")
        groups = defaultdict(int)
        for tc in tcs:
            groups[tc['group']] += 1
        for g, cnt in groups.items():
            print(f"    - {g}: {cnt}件")

    # --- v7で新規追加されたケース ---
    print(f"\n--- 8. v7で新規追加されたケース（元No.がNEW-x）---")
    for nn in sorted(new_orig_nos):
        tcs = v7_by_orig[nn]
        print(f"\n  {nn} ({len(tcs)}件):")
        for tc in tcs:
            print(f"    No.{tc['no']}: {tc['expected'][:80]}")

    # ============================================================
    # 統計比較サマリー
    # ============================================================
    print(f"\n\n{'='*80}")
    print(f"【統計比較サマリー】")
    print(f"{'='*80}")

    sa, sb = stats_a, stats_b
    rows = [
        ("テストケース数", f"{sa['total']}件", f"{sb['total']}件", f"{sa['total']-sb['total']:+d}"),
        ("期待値平均文字数", f"{sa['avg_len']:.1f}字", f"{sb['avg_len']:.1f}字", f"{sa['avg_len']-sb['avg_len']:+.1f}"),
        ("期待値適正範囲率", f"{sa['normal']/sa['total']*100:.1f}%", f"{sb['normal']/sb['total']*100:.1f}%", ""),
        ("手順記載率(直接)", f"{sa['proc_direct_rate']:.1f}%", f"{sb['proc_direct_rate']:.1f}%", f"{sa['proc_direct_rate']-sb['proc_direct_rate']:+.1f}"),
        ("手順記載率(継承含)", f"{sa['proc_inherited_rate']:.1f}%", f"{sb['proc_inherited_rate']:.1f}%", f"{sa['proc_inherited_rate']-sb['proc_inherited_rate']:+.1f}"),
        ("番号付き手順率", f"{sa['numbered_rate']:.1f}%", f"{sb['numbered_rate']:.1f}%", f"{sa['numbered_rate']-sb['numbered_rate']:+.1f}"),
        ("「こと」1回率", f"{sa['koto_1_rate']:.1f}%", f"{sb['koto_1_rate']:.1f}%", f"{sa['koto_1_rate']-sb['koto_1_rate']:+.1f}"),
        ("「こと」2回以上率", f"{sa['koto_2plus_rate']:.1f}%", f"{sb['koto_2plus_rate']:.1f}%", f"{sa['koto_2plus_rate']-sb['koto_2plus_rate']:+.1f}"),
        ("曖昧表現率", f"{sa['ambig_rate']:.1f}%", f"{sb['ambig_rate']:.1f}%", f"{sa['ambig_rate']-sb['ambig_rate']:+.1f}"),
        ("良好末尾表現率", f"{sa['good_end_rate']:.1f}%", f"{sb['good_end_rate']:.1f}%", f"{sa['good_end_rate']-sb['good_end_rate']:+.1f}"),
    ]

    print(f"\n{'指標':<22s} {'参考例':>12s} {'v7':>12s} {'差':>8s}")
    print("-" * 58)
    for name, va, vb, d in rows:
        print(f"{name:<22s} {va:>12s} {vb:>12s} {d:>8s}")

    # グループ別ケース数
    print(f"\n\n--- グループ別ケース数 ---")
    ga = defaultdict(int)
    for tc in all_tc_a:
        ga[tc['group']] += 1
    gb = defaultdict(int)
    for tc in all_tc_b:
        gb[tc['group']] += 1
    all_groups = sorted(set(list(ga.keys()) + list(gb.keys())))

    print(f"  {'グループ名':<48s} {'参考例':>5s} {'v7':>5s} {'備考':>10s}")
    print(f"  {'-'*72}")
    for g in all_groups:
        ca = ga.get(g, 0)
        cb = gb.get(g, 0)
        note = ""
        if ca > 0 and cb == 0:
            note = "参考例のみ"
        elif ca == 0 and cb > 0:
            note = "v7のみ"
        elif ca != cb:
            note = f"差{cb-ca:+d}"
        gname = g[:46] if len(g) > 46 else g
        print(f"  {gname:<48s} {ca:>5d} {cb:>5d} {note:>10s}")

    print(f"\n\n{'='*80}")
    print(f"分析完了")
    print(f"{'='*80}")


if __name__ == '__main__':
    main()
