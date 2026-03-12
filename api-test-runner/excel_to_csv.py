"""Excel API仕様書 → CSV 変換スクリプト.

Excel（原本）からCSVを自動生成し、既存の csv_parser.py が読める形式で出力する。
Excelを Single Source of Truth として維持しつつ、既存フローとの互換性を保つ。

Usage:
    python excel_to_csv.py                    # 全Excelを変換（プロダクト別ディレクトリ出力）
    python excel_to_csv.py --file "file.xlsx" # 特定ファイルのみ
    python excel_to_csv.py --dry-run          # 変換対象の一覧表示のみ
    python excel_to_csv.py --clean            # 既存CSV削除後に再生成

出力先:
    document/           ← TOKIUM経費精算（既存テストのデフォルト）
    document/invoicing/ ← TOKIUMインボイス・電子帳簿保存
    document/dencho/    ← TOKIUM電子帳簿保存
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DOCUMENT_DIR = Path(__file__).parent / "document"

# Excel → (CSV名プレフィックス, 出力サブディレクトリ)
EXCEL_CONFIG: dict[str, tuple[str, str]] = {
    "【TOKIUM】標準API仕様書（最新版）.xlsx": (
        "【TOKIUM】標準API仕様書（最新版）",
        "",  # document/ 直下（既存テストのデフォルト）
    ),
    "【TOKIUM電子帳簿保存】API仕様書 （最新版）.xlsx": (
        "【TOKIUM電子帳簿保存】API仕様書（最新版）",
        "dencho",
    ),
    "※最新版※【TOKIUMインボイス・電子帳簿保存】標準API仕様書（企業マスタの追加）.xlsx": (
        "【TOKIUMインボイス・電子帳簿保存】標準API仕様書（最新版）",
        "invoicing",
    ),
}

# シート名が数字で始まるもののみAPI仕様シートとして変換対象
API_SHEET_PATTERN = re.compile(r"^\d+")


def _is_api_sheet(name: str) -> bool:
    """API仕様シートかどうか判定."""
    return bool(API_SHEET_PATTERN.match(name))


def convert_sheet_to_csv(
    ws,
    output_path: Path,
    max_col: int = 28,
) -> int:
    """1シートをCSVファイルに変換. 出力行数を返す."""
    row_count = 0
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(max_col=max_col, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            writer.writerow(cells)
            row_count += 1
    return row_count


def convert_excel(
    excel_path: Path,
    output_dir: Path,
    prefix: str,
    dry_run: bool = False,
) -> list[dict]:
    """Excel内の全API仕様シートをCSVに変換."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    results = []

    output_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name in wb.sheetnames:
        if not _is_api_sheet(sheet_name):
            continue

        csv_filename = f"{prefix} - {sheet_name}.csv"
        csv_path = output_dir / csv_filename

        if dry_run:
            results.append({
                "sheet": sheet_name,
                "csv": csv_filename,
                "status": "dry-run",
            })
            continue

        ws = wb[sheet_name]
        row_count = convert_sheet_to_csv(ws, csv_path)
        results.append({
            "sheet": sheet_name,
            "csv": csv_filename,
            "rows": row_count,
            "status": "created",
        })

    wb.close()
    return results


def find_excel_files(target: str | None = None) -> list[Path]:
    """変換対象のExcelファイルを取得."""
    if target:
        path = DOCUMENT_DIR / target
        if not path.exists():
            print(f"ファイルが見つかりません: {path}", file=sys.stderr)
            sys.exit(1)
        return [path]
    return sorted(DOCUMENT_DIR.glob("*.xlsx"))


def clean_generated_csvs(base_dir: Path) -> int:
    """変換生成されたCSVを削除."""
    count = 0
    for _, (prefix, subdir) in EXCEL_CONFIG.items():
        output_dir = base_dir / subdir if subdir else base_dir
        for csv_file in output_dir.glob("*.csv"):
            if csv_file.name.startswith(prefix):
                csv_file.unlink()
                count += 1
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel API仕様書 → CSV 変換")
    parser.add_argument("--file", help="変換対象のExcelファイル名")
    parser.add_argument("--dry-run", action="store_true", help="変換対象の一覧表示のみ")
    parser.add_argument("--clean", action="store_true", help="既存CSV削除後に再生成")
    args = parser.parse_args()

    excel_files = find_excel_files(args.file)

    if not excel_files:
        print("Excelファイルが見つかりません。", file=sys.stderr)
        sys.exit(1)

    # --clean: 変換生成CSVを削除
    if args.clean and not args.dry_run:
        deleted = clean_generated_csvs(DOCUMENT_DIR)
        print(f"削除: {deleted} ファイル")

    total_sheets = 0
    for excel_path in excel_files:
        config = EXCEL_CONFIG.get(excel_path.name)
        if config is None:
            print(f"スキップ（マッピング未定義）: {excel_path.name}", file=sys.stderr)
            continue

        prefix, subdir = config
        output_dir = DOCUMENT_DIR / subdir if subdir else DOCUMENT_DIR

        print(f"\n{'[DRY-RUN] ' if args.dry_run else ''}変換: {excel_path.name}")
        if subdir:
            print(f"  出力先: document/{subdir}/")
        results = convert_excel(excel_path, output_dir, prefix, args.dry_run)

        for r in results:
            rows = f" ({r['rows']}行)" if "rows" in r else ""
            print(f"  {r['sheet']:40s} → {r['csv']}{rows}")
            total_sheets += 1

    print(f"\n合計: {total_sheets} シート{'（dry-run）' if args.dry_run else '変換完了'}")


if __name__ == "__main__":
    main()
